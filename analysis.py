"""
================================================================================
THESIS ANALYSIS SCRIPT
Capital Adequacy and Liquidity Resilience of Fiat-Backed Stablecoins
A Quarterly Stress-Testing Analysis (2020-2025)
================================================================================

PURPOSE
-------
This is the single growing analysis script for the thesis. Each work block
(Mon-Tue, Wed-Thu, etc.) adds a clearly-marked section. Future sessions append
new sections; they do not modify previous ones.

HOW TO RUN
----------
1. Place the master Excel in the same directory as this script:
     USDC_USDT_USDP_Basel3_Master_v3.xlsx
2. Run from the command line:
     python analysis.py
3. Outputs are written to ./outputs/ (figures as PNG, tables as CSV).

STRUCTURE
---------
- SECTION 0: Setup, imports, configuration
- SECTION 1: Data loading and the 14-quarter analysis panel
- SECTION 2: Mon-Tue   - Descriptive table and asset composition figures (DONE)
- SECTION 3: Wed-Thu   - RWA computation (TO DO)
- SECTION 4: Week 2    - Interest-rate stress (DONE)
- SECTION 5: Week 2    - Redemption / LCR stress (DONE)
- SECTION 6: Week 3    - Narrow-bank vs MMF classification (DONE)

DATA SOURCE
-----------
The master Excel in the knowledge base is treated as the verified source of
truth. This script does not re-verify the input data; it consumes it as given.

CONVENTIONS
-----------
- All monetary values are in USD.
- "$bn" in printed tables means billions of US dollars.
- Quarter labels follow YYYY-Qn format (e.g., '2023-Q4').
- The analytical scope is Q3 2022 to Q4 2025 (14 quarters).
================================================================================
"""

# ==============================================================================
# SECTION 0: Setup, imports, configuration
# ==============================================================================

import os
import warnings
from pathlib import Path

import pandas as pd
import numpy as np   # runtime fix (Jul 2026): Section 8's top-level RETAINED_EARNINGS uses np.nan
import matplotlib
matplotlib.use('Agg')  # No GUI needed; we write PNGs to disk.
import matplotlib.pyplot as plt

warnings.filterwarnings('ignore')

# --- File paths -------------------------------------------------------------
# Single base folder. The master Excel is the ONLY input file — TB3MS, the
# MSPD bills-outstanding series, and the BTC daily price history are all
# embedded in it as sheets (TB3MS, MSPD_SumSecty, BTC_Daily). The outputs/
# folder is created inside BASE_DIR.
# Defaults to the folder this script lives in, so `git clone` + `python
# analysis.py` works unmodified on any machine. Override with THESIS_DIR if
# the master Excel lives elsewhere.
BASE_DIR = Path(__file__).resolve().parent

# Allow an env-var override (handy for servers / CI) without touching code:
#   set THESIS_DIR=...   (Windows)  /  export THESIS_DIR=...  (mac/Linux)
import os as _os
if _os.environ.get('THESIS_DIR'):
    BASE_DIR = Path(_os.environ['THESIS_DIR'])

EXCEL_PATH = BASE_DIR / 'USDC_USDT_USDP_Basel3_Master_v8.xlsx'
for _fb in ('USDC_USDT_USDP_Basel3_Master_v7.xlsx',
            'USDC_USDT_USDP_Basel3_Master_v5.xlsx'):
    if not EXCEL_PATH.exists():
        EXCEL_PATH = BASE_DIR / _fb
OUT_DIR = BASE_DIR / 'outputs'
OUT_DIR.mkdir(parents=True, exist_ok=True)

# --- Analytical scope ---
# Locked decision: only Q3 2022 to Q4 2025. Earlier quarters exist in the raw
# sheets as descriptive context but are NOT used in RWA, LCR, or stress tests
# because at least one issuer lacks Tier-3 disclosure in those periods.
SCOPE_FIRST_Q = '2022-Q3'
SCOPE_LAST_Q  = '2025-Q4'

# --- Plot style ---
plt.rcParams.update({
    'font.size': 9,
    'font.family': 'DejaVu Sans',
    'axes.edgecolor': '#666',
    'axes.spines.top': False,
    'axes.spines.right': False,
})

# --- Helper: convert 'YYYY-Qn' to a sortable integer (Q1 2022 = 8089, etc.) ---
def q_to_num(q):
    y, qq = q.split('-Q')
    return int(y) * 4 + int(qq)


# ==============================================================================
# SECTION 1: Data loading and panel construction
# ------------------------------------------------------------------------------
# We load the Comparison_Quarterly sheet, which already pairs USDC and USDT
# quarter-end values side by side. USDP is loaded separately from USDP_RAW
# because it is used only for the classification spectrum (RQ4), not for the
# main quarterly comparison.
# ==============================================================================

def load_panel():
    """
    Load Comparison_Quarterly and restrict it to the locked scope window.

    Returns
    -------
    panel : pd.DataFrame
        14 rows (one per quarter from Q3 2022 to Q4 2025), all USDC and USDT
        balance-sheet columns populated. This is the analytical backbone for
        RQ1, RQ2, and RQ3.
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Master Excel not found.\n"
            f"  Looked for: {EXCEL_PATH}\n"
            f"  cwd is    : {Path.cwd()}\n"
            f"  Fix: place 'USDC_USDT_USDP_Basel3_Master_v4.xlsx' in the same "
            f"folder as this script/notebook and restart the kernel."
        )
    comp = pd.read_excel(EXCEL_PATH, sheet_name='Comparison_Quarterly')

    # Keep only quarters inside the locked scope.
    lo, hi = q_to_num(SCOPE_FIRST_Q), q_to_num(SCOPE_LAST_Q)
    in_scope = comp['Quarter'].apply(lambda q: lo <= q_to_num(q) <= hi)
    panel = comp[in_scope].copy().reset_index(drop=True)

    return panel


def load_usdp():
    """
    Backwards-compat wrapper — returns all Tier-3 month-end rows in scope.
    Used for RQ4 classification; load_usdp_panel() is used for RWA.
    """
    usdp = pd.read_excel(EXCEL_PATH, sheet_name='USDP_RAW')
    usdp['Report_Date'] = pd.to_datetime(usdp['Report_Date'])
    mask = (
        (usdp['Date_Type'] == 'Month-End')
        & (usdp['Disclosure_Quality_Tier'] == 3)
        & (usdp['Report_Date'] >= '2022-11-01')
    )
    return usdp[mask].copy().reset_index(drop=True)


def load_usdp_panel():
    """
    Load USDP_RAW, filter to Tier-3 month-end rows in scope, aggregate to
    one row per quarter (last month-end per quarter).

    USDP asset columns:
        US_Tbills_Direct  — direct T-bill holdings  (0% RW, Level 1)
        Reverse_Repos     — UST-collateralised repos (0% RW, Level 1)
        Cash_Deposits     — cash at custodian banks  (20% RW)
        MMF_NAV           — always 0 (Paxos holds no MMF units)

    Coverage: 13 of 14 panel quarters.
    Q3 2022 is missing (no Tier-3 row in the source data).

    Returns
    -------
    pd.DataFrame  — one row per available quarter, asset columns + Quarter label.
    """
    raw = pd.read_excel(EXCEL_PATH, sheet_name='USDP_RAW')
    raw['Report_Date'] = pd.to_datetime(raw['Report_Date'])
    raw['Quarter'] = (raw['Report_Date'].dt.year.astype(str)
                      + '-Q'
                      + raw['Report_Date'].dt.quarter.astype(str))
    lo, hi = q_to_num(SCOPE_FIRST_Q), q_to_num(SCOPE_LAST_Q)
    mask = (
        (raw['Date_Type'] == 'Month-End')
        & (raw['Disclosure_Quality_Tier'] == 3)
        & (raw['Quarter'].apply(lambda q: lo <= q_to_num(q) <= hi))
    )
    tier3 = raw[mask].sort_values('Report_Date')
    qend  = tier3.groupby('Quarter').last().reset_index()
    keep  = ['Quarter', 'Report_Date', 'USDP_Circulation', 'Total_Assets',
             'US_Tbills_Direct', 'Reverse_Repos', 'Cash_Deposits', 'Equity_Implied']
    return qend[keep].copy()


# ==============================================================================
# SECTION 2: Mon-Tue - Descriptive table and asset composition figures
# ------------------------------------------------------------------------------
# WHAT THIS SECTION PRODUCES
#   - descriptive_table.csv : the 14-quarter quarterly summary
#   - fig_week1_descriptive.png : 4-panel figure (composition + supply + equity)
#
# WHAT IT ANSWERS
#   This is the descriptive groundwork before RWA. It shows scale, supply
#   trajectory, composition contrast, and equity buffer differences across
#   the two main issuers in the analytical window.
# ==============================================================================

def build_descriptive_table(panel):
    """
    Build the 14-quarter summary table.

    Columns:
      - Quarter
      - USDC supply, total assets, equity %
      - USDT supply, total assets, equity %
      - Market share of USDC (as % of USDC + USDT combined supply)

    All monetary values converted to USD billions for readability.
    """
    rows = []
    for _, r in panel.iterrows():
        uc = (r['USDC_Circulation'] or 0) / 1e9          # USDC supply ($bn)
        ua = (r['USDC_Total_Assets'] or 0) / 1e9         # USDC total assets ($bn)
        ue = ((r['USDC_Total_Assets'] - r['USDC_Circulation'])
              / r['USDC_Circulation'] * 100) if r['USDC_Circulation'] else 0

        tc = (r['USDT_Circulation'] or 0) / 1e9
        ta = (r['USDT_Total_Assets'] or 0) / 1e9
        te = ((r['USDT_Total_Assets'] - r['USDT_Circulation'])
              / r['USDT_Circulation'] * 100) if r['USDT_Circulation'] else 0

        # Market share = USDC's slice of the (USDC + USDT) market.
        market_share = uc / (uc + tc) * 100 if (uc + tc) else 0

        rows.append({
            'Quarter':            r['Quarter'],
            'USDC_Supply_bn':     round(uc, 2),
            'USDC_Assets_bn':     round(ua, 2),
            'USDC_Equity_pct':    round(ue, 3),
            'USDT_Supply_bn':     round(tc, 2),
            'USDT_Assets_bn':     round(ta, 2),
            'USDT_Equity_pct':    round(te, 3),
            'USDC_Market_Share_pct': round(market_share, 1),
        })

    return pd.DataFrame(rows)


def make_figures(panel, usdp_panel=None):
    """
    Build the descriptive figure — three-way (USDC, USDP, USDT) version.

    Panel layout (2 x 3 grid):
      (0,0) USDC reserve composition (with CRF look-through)
      (0,1) USDP reserve composition (narrow-bank: T-bills + repos + cash)
      (0,2) USDT reserve composition (all 7 asset classes)
      (1,0) Supply trajectories — USDC vs USDT (left axis) + USDP (right axis)
      (1,1) Equity buffer % over time, all three issuers
      (1,2) "Off-Basel" assets share (BTC + gold + secured loans / Total)

    USDP is plotted on a secondary axis in panel (1,0) because its scale
    ($0.05–1.0bn) is two orders of magnitude smaller than USDC/USDT ($50–190bn).
    """
    q = panel['Quarter'].tolist()
    x = range(len(q))

    def col(name):
        """Return a column from the panel converted to billions, NaN-safe."""
        return (panel[name].fillna(0) / 1e9).values

    fig, axes = plt.subplots(2, 3, figsize=(20, 10))

    # ----- (0,0): USDC composition -----
    ax = axes[0, 0]
    usdc_tb       = col('USDC_Tbills_in_Fund')
    usdc_repos    = col('USDC_Repos')
    usdc_ext_cash = col('USDC_External_Cash')
    usdc_crfcash  = (col('USDC_MMF_NAV') - usdc_tb - usdc_repos).clip(min=0)

    ax.stackplot(x, usdc_tb, usdc_repos, usdc_crfcash, usdc_ext_cash,
                 labels=['T-bills (in CRF)', 'Repos (in CRF)',
                         'CRF cash', 'External cash'],
                 colors=['#1f4e79', '#2e75b6', '#9dc3e6', '#cfe2f3'])
    ax.set_title('USDC — Reserve Composition (CRF look-through)', fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('USD billions')
    ax.legend(loc='upper left', fontsize=7, framealpha=0.9)
    ax.grid(axis='y', alpha=0.3)

    # ----- (0,1): USDP composition (millions scale, narrow-bank) -----
    ax = axes[0, 1]
    if usdp_panel is not None and len(usdp_panel) > 0:
        # Align USDP to the 14-quarter panel; NaN where data is missing
        usdp_by_q = usdp_panel.set_index('Quarter')
        usdp_tb   = (usdp_by_q.reindex(q)['US_Tbills_Direct'].fillna(0) / 1e6).values
        usdp_rp   = (usdp_by_q.reindex(q)['Reverse_Repos'].fillna(0)    / 1e6).values
        usdp_cs   = (usdp_by_q.reindex(q)['Cash_Deposits'].fillna(0)    / 1e6).values
        ax.stackplot(x, usdp_tb, usdp_rp, usdp_cs,
                     labels=['T-bills (direct)', 'Overnight repos', 'Cash at custodians'],
                     colors=['#1f4e79', '#2e75b6', '#cfe2f3'])
        # Annotate the wind-down: composition shifts to ~100% cash by 2024-Q3
        ax.text(0.02, 0.97,
                'Note: USDP shrinks from $0.97bn (Q3 2022)\nto $48m (Q4 2025);\ncomposition runs to ~100% cash by 2024-Q3',
                transform=ax.transAxes, va='top', fontsize=6.5,
                color='#555', style='italic',
                bbox=dict(facecolor='white', alpha=0.75, edgecolor='none', pad=2))
    ax.set_title('USDP — Reserve Composition (narrow bank)', fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('USD millions')
    ax.legend(loc='upper right', fontsize=7, framealpha=0.9)
    ax.grid(axis='y', alpha=0.3)

    # ----- (0,2): USDT composition -----
    ax = axes[0, 2]
    ax.stackplot(x,
                 col('USDT_Tbills_Direct'), col('USDT_Total_Repos'),
                 col('USDT_MMF'),           col('USDT_Cash'),
                 col('USDT_BTC'),           col('USDT_Gold'),
                 col('USDT_Sec_Loans'),
                 labels=['T-bills', 'Repos', 'MMF', 'Cash',
                         'Bitcoin', 'Gold', 'Secured loans'],
                 colors=['#1f4e79', '#2e75b6', '#9dc3e6', '#cfe2f3',
                         '#f7931a', '#d4af37', '#c00000'])
    ax.set_title('USDT — Reserve Composition', fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('USD billions')
    ax.legend(loc='upper left', fontsize=7, framealpha=0.9)
    ax.grid(axis='y', alpha=0.3)

    # ----- (1,0): Supply trajectories (USDC/USDT left axis, USDP right axis) -----
    ax = axes[1, 0]
    ax.plot(x, col('USDC_Circulation'), 'o-', color='#2e75b6',
            label='USDC (left)', lw=2)
    ax.plot(x, col('USDT_Circulation'), 's-', color='#26a17b',
            label='USDT (left)', lw=2)

    # USDP on right axis — millions
    ax_r = ax.twinx()
    if usdp_panel is not None and len(usdp_panel) > 0:
        usdp_circ = (usdp_by_q.reindex(q)['USDP_Circulation'] / 1e6).values
        ax_r.plot(x, usdp_circ, 'D:', color='#70ad47', lw=2, ms=5,
                  label='USDP (right)')
    ax_r.set_ylabel('USDP supply (USD millions)', color='#70ad47', fontsize=9)
    ax_r.tick_params(axis='y', labelcolor='#70ad47')

    if '2023-Q1' in q:
        svb_x = q.index('2023-Q1')
        ax.axvline(svb_x, color='red', ls='--', alpha=0.5, lw=1)
        ax.annotate('SVB depeg\n(Mar 2023)', xy=(svb_x, ax.get_ylim()[1] * 0.92),
                    xytext=(svb_x + 0.2, ax.get_ylim()[1] * 0.85),
                    fontsize=7, color='red')

    ax.set_title('Circulating Supply Trajectory (3-way)', fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('USDC / USDT supply (USD billions)')
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax_r.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, loc='center left', fontsize=7)
    ax.grid(alpha=0.3)

    # ----- (1,1): Equity buffer % -----
    ax = axes[1, 1]
    usdc_eq = ((panel['USDC_Total_Assets'] - panel['USDC_Circulation'])
               / panel['USDC_Circulation'] * 100).values
    usdt_eq = ((panel['USDT_Total_Assets'] - panel['USDT_Circulation'])
               / panel['USDT_Circulation'] * 100).values

    ax.plot(x, usdc_eq, 'o-', color='#2e75b6', label='USDC', lw=2)
    ax.plot(x, usdt_eq, 's-', color='#26a17b', label='USDT', lw=2)

    if usdp_panel is not None and len(usdp_panel) > 0:
        usdp_eq = ((usdp_by_q.reindex(q)['Total_Assets']
                    - usdp_by_q.reindex(q)['USDP_Circulation'])
                   / usdp_by_q.reindex(q)['USDP_Circulation'] * 100).values
        ax.plot(x, usdp_eq, 'D:', color='#70ad47', label='USDP', lw=2, ms=5)

    ax.set_title('Reserve Excess (Assets − Circulation) as % of Circulation',
                 fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('% buffer')
    ax.legend(fontsize=8)
    ax.grid(alpha=0.3)
    ax.axhline(0, color='black', lw=0.5)

    # ----- (1,2): Off-Basel share -----
    ax = axes[1, 2]
    usdc_offbasel = (col('USDC_External_Cash') * 0)  # USDC: zero off-Basel
    usdt_offbasel_share = ((col('USDT_BTC') + col('USDT_Gold') + col('USDT_Sec_Loans'))
                           / col('USDT_Total_Assets') * 100)
    # USDC off-Basel share: zero throughout; plot for completeness
    usdc_offbasel_share = (usdc_offbasel
                           / panel['USDC_Total_Assets'].values * 100)

    bw = 0.35
    xs = [i - bw/2 for i in x]
    xt = [i + bw/2 for i in x]
    ax.bar(xs, usdc_offbasel_share, bw, color='#2e75b6', label='USDC')
    ax.bar(xt, usdt_offbasel_share, bw, color='#26a17b', label='USDT')

    if usdp_panel is not None and len(usdp_panel) > 0:
        # USDP off-Basel share is zero throughout (no BTC/gold/secured loans)
        usdp_offbasel_share = [0] * len(q)
        # Don't bother plotting zeros — note it in text instead
        ax.text(0.02, 0.94,
                'USDC & USDP: 0% off-Basel (no BTC/gold/secured loans)',
                transform=ax.transAxes, va='top', fontsize=7,
                color='#555', style='italic',
                bbox=dict(facecolor='white', alpha=0.85, edgecolor='none', pad=2))

    ax.set_title('"Off-Basel" Asset Share (BTC + gold + secured loans / total)',
                 fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('% of total assets')
    ax.legend(fontsize=8)
    ax.grid(axis='y', alpha=0.3)

    plt.tight_layout()
    return fig


# ==============================================================================
# SECTION 3: Wed-Thu — RWA computation (RQ1)
# ------------------------------------------------------------------------------
# WHAT THIS SECTION PRODUCES
#   - rwa_table.csv       : quarterly RWA by asset class + ratio, both issuers
#   - rwa_evolution.png   : stacked bar chart, RWA by asset class per quarter
#   - rwa_ratio.png       : RWA / total assets over time (risk-density metric)
#
# METHODOLOGY (locked in master Excel Methodology sheet)
#
#   Risk weights applied:
#     0%    : US T-bills (direct), overnight UST-collateralised repos (Level 1)
#     20%   : Cash at banks / external deposits; opaque MMF units (CRF, Option B
#             sensitivity only; primary treatment = look-through to 0% T-bills)
#     100%  : Gold (physical commodity, Basel SA-CCR commodity treatment);
#             secured loans (private, no Basel collateral recognition for stablecoin
#             issuers that are non-bank); "other/residual" USDT category (unknown
#             composition → conservative 100%)
#     1250% : Bitcoin (Basel III SCO60.108, Group 2 unbacked cryptoasset)
#
#   USDC — TWO PARALLEL TREATMENTS:
#     Primary (look-through):
#       CRF decomposed → T-bills (0%) + repos (0%) + CRF residual cash (20%)
#       External cash outside fund (20%)
#     Sensitivity (opaque / Option B):
#       CRF entire NAV treated as opaque MMF unit (20%)
#       External cash outside fund (20%)
#     Note: for Q3 2022, MMF_NAV = 0 (CRF not yet introduced mid-Q4 2022).
#     For that quarter, T-bills are direct (0%) + external cash (20%).
#
#   USDT — single treatment (no look-through available):
#     T-bills direct     (0%)
#     Total repos        (0%)  — all repo categories; Tether discloses as UST-backed
#     MMF units          (20%) — opaque; Tether does not disclose underlying
#     Cash at banks      (20%)
#     Bitcoin            (1250%)
#     Gold               (100%)
#     Secured loans      (100%)
#     Other / residual   (100%) — computed as Total_Assets − sum of named lines;
#                                  includes Other_Investments + Non-US Tbills etc.
#
#   CITES: BIS Basel Framework SCO60.108 (1250% cryptoasset weight);
#          CRE20.31 (0% sovereign); CRE20.39 (20% bank); LCR30.2 (Level 1 HQLA).
# ==============================================================================

# --- Risk weight constants ---
RW_SOVEREIGN    = 0.00      # US T-bills, UST repos
RW_BANK         = 0.20      # Cash at banks, MMF units (opaque)
RW_COMMODITY    = 1.00      # Gold, secured loans, unclassified residual
RW_CRYPTO       = 12.50     # Bitcoin (1250%)


def _usdc_rwa_lookthrough(row):
    """
    Primary USDC RWA: CRF decomposed into constituent holdings.

    Logic:
      - If CRF exists (MMF_NAV > 0):
          tbills_in_fund   → 0%  RW  (Level 1 sovereign)
          repos_in_fund    → 0%  RW  (overnight UST-collateralised)
          crf_residual_cash = max(MMF_NAV - tbills - repos, 0) → 20% RW (cash)
          external_cash    → 20% RW
      - If CRF absent (Q3 2022 and half of Q4 2022 transition):
          tbills_in_fund (direct T-bills) → 0% RW
          external_cash → 20% RW
    """
    tbills   = float(row['USDC_Tbills_in_Fund'])
    repos    = float(row['USDC_Repos'])
    mmf_nav  = float(row['USDC_MMF_NAV'])
    ext_cash = float(row['USDC_External_Cash'])

    if mmf_nav > 0:
        crf_residual = max(mmf_nav - tbills - repos, 0.0)
    else:
        # Pre-CRF: T-bills were held directly (already in USDC_Tbills_in_Fund)
        crf_residual = 0.0

    rwa_tbills   = tbills        * RW_SOVEREIGN   # 0%
    rwa_repos    = repos         * RW_SOVEREIGN   # 0%
    rwa_crf_cash = crf_residual  * RW_BANK        # 20%
    rwa_ext_cash = ext_cash      * RW_BANK        # 20%

    return {
        'USDC_RWA_Tbills':     rwa_tbills,
        'USDC_RWA_Repos':      rwa_repos,
        'USDC_RWA_CRF_Cash':   rwa_crf_cash,
        'USDC_RWA_ExtCash':    rwa_ext_cash,
        'USDC_RWA_Total_Primary': rwa_tbills + rwa_repos + rwa_crf_cash + rwa_ext_cash,
    }


def _usdc_rwa_opaque(row):
    """
    Sensitivity USDC RWA: entire CRF NAV treated as opaque MMF unit (20%).
    """
    mmf_nav  = float(row['USDC_MMF_NAV'])
    tbills   = float(row['USDC_Tbills_in_Fund'])
    ext_cash = float(row['USDC_External_Cash'])

    if mmf_nav > 0:
        # Opaque: whole NAV at 20%; external cash also 20%
        rwa_mmf  = mmf_nav  * RW_BANK
        rwa_cash = ext_cash * RW_BANK
    else:
        # Pre-CRF quarter: direct T-bills still at 0%, cash at 20%
        rwa_mmf  = tbills   * RW_SOVEREIGN
        rwa_cash = ext_cash * RW_BANK

    return {
        'USDC_RWA_MMF_Opaque':   rwa_mmf,
        'USDC_RWA_ExtCash_Sens': rwa_cash,
        'USDC_RWA_Total_Sensitivity': rwa_mmf + rwa_cash,
    }


def _usdt_rwa(row):
    """
    USDT RWA: single treatment across all 7 named lines + residual.
    """
    tbills       = float(row['USDT_Tbills_Direct'])
    repos        = float(row['USDT_Total_Repos'])
    mmf          = float(row['USDT_MMF'])
    cash         = float(row['USDT_Cash'])
    btc          = float(row['USDT_BTC'])
    gold         = float(row['USDT_Gold'])
    sec_loans    = float(row['USDT_Sec_Loans'])
    total_assets = float(row['USDT_Total_Assets'])

    named_sum = tbills + repos + mmf + cash + btc + gold + sec_loans
    other     = max(total_assets - named_sum, 0.0)   # residual at 100%

    rwa_tbills    = tbills    * RW_SOVEREIGN   # 0%
    rwa_repos     = repos     * RW_SOVEREIGN   # 0%
    rwa_mmf       = mmf       * RW_BANK        # 20%
    rwa_cash      = cash      * RW_BANK        # 20%
    rwa_btc       = btc       * RW_CRYPTO      # 1250%
    rwa_gold      = gold      * RW_COMMODITY   # 100%
    rwa_sec_loans = sec_loans * RW_COMMODITY   # 100%
    rwa_other     = other     * RW_COMMODITY   # 100%

    return {
        'USDT_RWA_Tbills':     rwa_tbills,
        'USDT_RWA_Repos':      rwa_repos,
        'USDT_RWA_MMF':        rwa_mmf,
        'USDT_RWA_Cash':       rwa_cash,
        'USDT_RWA_BTC':        rwa_btc,
        'USDT_RWA_Gold':       rwa_gold,
        'USDT_RWA_SecLoans':   rwa_sec_loans,
        'USDT_RWA_Other':      rwa_other,
        'USDT_RWA_Total':
            rwa_tbills + rwa_repos + rwa_mmf + rwa_cash +
            rwa_btc + rwa_gold + rwa_sec_loans + rwa_other,
    }


def _usdp_rwa(row):
    """
    USDP RWA: pure narrow-bank portfolio.
    Three asset lines, two weights:
        US_Tbills_Direct → 0%  (sovereign, CRE20.31)
        Reverse_Repos    → 0%  (overnight UST-collateralised)
        Cash_Deposits    → 20% (bank counterparty, CRE20.39)
    NaN values are treated as zero — Paxos omits lines with zero balances.
    """
    tbills = float(row['US_Tbills_Direct']) if pd.notna(row['US_Tbills_Direct']) else 0.0
    repos  = float(row['Reverse_Repos'])    if pd.notna(row['Reverse_Repos'])    else 0.0
    cash   = float(row['Cash_Deposits'])    if pd.notna(row['Cash_Deposits'])    else 0.0

    rwa_tbills = tbills * RW_SOVEREIGN   # 0%
    rwa_repos  = repos  * RW_SOVEREIGN   # 0%
    rwa_cash   = cash   * RW_BANK        # 20%

    return {
        'USDP_RWA_Tbills': rwa_tbills,
        'USDP_RWA_Repos':  rwa_repos,
        'USDP_RWA_Cash':   rwa_cash,
        'USDP_RWA_Total':  rwa_tbills + rwa_repos + rwa_cash,
    }


def compute_usdp_rwa(usdp_panel):
    """
    Compute RWA for all available USDP quarters and merge onto the full
    14-quarter scope spine.  Missing quarters appear as NaN.

    Returns
    -------
    pd.DataFrame — 14 rows, one per scope quarter.
        USDP_Data_Available : bool — True when Tier-3 data exists.
    """
    all_quarters = [
        '2022-Q3','2022-Q4',
        '2023-Q1','2023-Q2','2023-Q3','2023-Q4',
        '2024-Q1','2024-Q2','2024-Q3','2024-Q4',
        '2025-Q1','2025-Q2','2025-Q3','2025-Q4',
    ]
    spine = pd.DataFrame({'Quarter': all_quarters})

    records = []
    for _, row in usdp_panel.iterrows():
        rec = {'Quarter': row['Quarter']}
        rec.update(_usdp_rwa(row))
        ta = float(row['Total_Assets'])
        rec['USDP_Total_Assets']   = ta
        rec['USDP_Circulation']    = float(row['USDP_Circulation'])
        rec['USDP_Equity_Implied'] = float(row['Equity_Implied'])
        rec['USDP_RWA_Ratio']      = rec['USDP_RWA_Total'] / ta if ta else 0.0
        records.append(rec)

    computed  = pd.DataFrame(records)
    usdp_rwa  = spine.merge(computed, on='Quarter', how='left')
    usdp_rwa['USDP_Data_Available'] = usdp_rwa['USDP_RWA_Total'].notna()
    return usdp_rwa


def make_rwa_ratio_threeway(rwa, usdp_rwa):
    """
    Three-issuer RWA density chart.

    Left axis  : USDC primary (solid) + sensitivity (dashed) + USDP (dotted green)
    Right axis : USDT (red)

    USDP is the narrow-bank benchmark — its near-zero ratio anchors the bottom
    of the left axis and makes the USDC vs USDT divergence visually starker.
    A gap in the USDP line at Q3 2022 reflects the missing Tier-3 row.
    """
    q = rwa['Quarter'].tolist()
    x = list(range(len(q)))

    fig, ax1 = plt.subplots(figsize=(14, 6))
    ax2 = ax1.twinx()

    usdc_primary     = rwa['USDC_RWA_Ratio_Primary'].values * 100
    usdc_sensitivity = rwa['USDC_RWA_Ratio_Sensitivity'].values * 100
    usdt             = rwa['USDT_RWA_Ratio'].values * 100
    usdp_vals        = (usdp_rwa.set_index('Quarter')
                                .reindex(q)['USDP_RWA_Ratio'] * 100).values

    # USDC shaded band
    ax1.fill_between(x, usdc_primary, usdc_sensitivity,
                     alpha=0.20, color='#2e75b6',
                     label='USDC range (primary–sensitivity)')
    ax1.plot(x, usdc_primary,     'o-',  color='#2e75b6', lw=2,
             label='USDC — primary (look-through)')
    ax1.plot(x, usdc_sensitivity, 's--', color='#1f4e79', lw=1.5,
             label='USDC — sensitivity (opaque CRF)')

    # USDP dotted — NaN gaps handled automatically
    ax1.plot(x, usdp_vals, 'D:', color='#70ad47', lw=2, ms=5,
             label='USDP — narrow bank')

    # USDT right axis
    ax2.plot(x, usdt, '^-', color='#c00000', lw=2, label='USDT')
    ax2.set_ylabel('USDT RWA / Total Assets (%)', color='#c00000', fontsize=9)
    ax2.tick_params(axis='y', labelcolor='#c00000')

    ax1.set_ylabel('USDC / USDP  RWA / Total Assets (%)', fontsize=9)
    ax1.set_xticks(x)
    ax1.set_xticklabels(q, rotation=45, ha='right', fontsize=8)
    ax1.set_title(
        'Basel III RWA Density Ratio (RWA / Total Assets)\n'
        'USDC vs USDP vs USDT — Q3 2022 to Q4 2025',
        fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)

    # SVB annotation
    if '2023-Q1' in q:
        svb_x = q.index('2023-Q1')
        ax1.axvline(svb_x, color='red', ls='--', alpha=0.4, lw=1)
        ax1.annotate('SVB depeg', xy=(svb_x, ax1.get_ylim()[1] * 0.78),
                     fontsize=7, color='red')

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2,
               loc='upper left', fontsize=8, framealpha=0.9)
    plt.tight_layout()
    return fig


def print_usdp_rwa_summary(usdp_rwa):
    """Print USDP RWA console table."""
    print("\nUSDP RWA Summary (narrow-bank benchmark)")
    print("-" * 72)
    print(f"{'Quarter':<10}  {'Circulation':>13}  {'RWA Total':>11}"
          f"  {'RWA Ratio':>10}  {'Data':>6}")
    print("-" * 72)
    for _, r in usdp_rwa.iterrows():
        if r['USDP_Data_Available']:
            circ  = r['USDP_Circulation'] / 1e9
            rwa_m = r['USDP_RWA_Total']   / 1e6
            ratio = r['USDP_RWA_Ratio']   * 100
            print(f"{r['Quarter']:<10}  {circ:>11.3f}bn  {rwa_m:>9.1f}m"
                  f"  {ratio:>9.3f}%  {'yes':>6}")
        else:
            print(f"{r['Quarter']:<10}  {'—':>13}  {'—':>11}"
                  f"  {'—':>10}  {'no':>6}")
    print("-" * 72)


def compute_rwa(panel):
    """
    Compute Basel III RWA for USDC (primary + sensitivity) and USDT for all
    14 panel quarters.

    Returns
    -------
    rwa : pd.DataFrame
        One row per quarter. Contains:
          - Per-asset RWA columns (dollars)
          - USDC_RWA_Total_Primary, USDC_RWA_Total_Sensitivity
          - USDT_RWA_Total
          - USDC_RWA_Ratio_Primary   = USDC RWA (primary) / USDC total assets
          - USDC_RWA_Ratio_Sensitivity
          - USDT_RWA_Ratio           = USDT RWA / USDT total assets
    """
    records = []
    for _, row in panel.iterrows():
        rec = {'Quarter': row['Quarter']}

        usdc_lt = _usdc_rwa_lookthrough(row)
        usdc_op = _usdc_rwa_opaque(row)
        usdt    = _usdt_rwa(row)

        rec.update(usdc_lt)
        rec.update(usdc_op)
        rec.update(usdt)

        usdc_assets = float(row['USDC_Total_Assets'])
        usdt_assets = float(row['USDT_Total_Assets'])

        rec['USDC_Total_Assets'] = usdc_assets
        rec['USDT_Total_Assets'] = usdt_assets
        rec['USDC_Circulation']  = float(row['USDC_Circulation'])
        rec['USDT_Circulation']  = float(row['USDT_Circulation'])

        rec['USDC_RWA_Ratio_Primary']     = (rec['USDC_RWA_Total_Primary']
                                              / usdc_assets if usdc_assets else 0)
        rec['USDC_RWA_Ratio_Sensitivity'] = (rec['USDC_RWA_Total_Sensitivity']
                                              / usdc_assets if usdc_assets else 0)
        rec['USDT_RWA_Ratio']             = (rec['USDT_RWA_Total']
                                              / usdt_assets if usdt_assets else 0)

        records.append(rec)

    return pd.DataFrame(records)


def make_rwa_evolution(rwa):
    """
    Stacked bar chart: RWA by asset class per quarter, USDC and USDT side by side.

    Layout:
      Left  (ax0): USDC — Primary treatment (look-through)
      Centre(ax1): USDC — Sensitivity treatment (opaque CRF)
      Right (ax2): USDT — all asset classes

    Y-axis in USD trillions for USDT (to accommodate BTC 1250% weight);
    separate Y-axis scale for USDC panels (billions).
    """
    q = rwa['Quarter'].tolist()
    x = range(len(q))

    fig, axes = plt.subplots(1, 3, figsize=(18, 7))

    def bn(col):
        return (rwa[col].fillna(0) / 1e9).values

    # ---- USDC Primary ----
    ax = axes[0]
    # Stacked bars: T-bills (0 RWA — not shown), repos (0), CRF cash, ext cash
    bar_crf  = ax.bar(x, bn('USDC_RWA_CRF_Cash'), label='CRF residual cash (20%)',
                      color='#9dc3e6')
    bar_ext  = ax.bar(x, bn('USDC_RWA_ExtCash'), bottom=bn('USDC_RWA_CRF_Cash'),
                      label='External cash (20%)', color='#2e75b6')
    ax.set_title('USDC — RWA: Primary (look-through)', fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('RWA (USD billions)')
    ax.legend(loc='upper left', fontsize=7)
    ax.grid(axis='y', alpha=0.3)
    # Annotate: 0%-weight assets are the invisible majority
    ax.text(0.02, 0.97, 'T-bills & repos: 0% RW\n(not visible in bars)',
            transform=ax.transAxes, va='top', fontsize=7,
            color='#555', style='italic')

    # ---- USDC Sensitivity ----
    ax = axes[1]
    ax.bar(x, bn('USDC_RWA_MMF_Opaque'), label='CRF (opaque MMF, 20%)',
           color='#f4b942')
    ax.bar(x, bn('USDC_RWA_ExtCash_Sens'), bottom=bn('USDC_RWA_MMF_Opaque'),
           label='External cash (20%)', color='#c07000')
    ax.set_title('USDC — RWA: Sensitivity (opaque CRF)', fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('RWA (USD billions)')
    ax.legend(loc='upper left', fontsize=7)
    ax.grid(axis='y', alpha=0.3)

    # ---- USDT ----
    ax = axes[2]
    # Use trillions for USDT because BTC at 1250% creates very large RWA
    def tr(col):
        return (rwa[col].fillna(0) / 1e12).values

    b_other  = tr('USDT_RWA_Other')
    b_sl     = tr('USDT_RWA_SecLoans')
    b_gold   = tr('USDT_RWA_Gold')
    b_btc    = tr('USDT_RWA_BTC')
    b_cash   = tr('USDT_RWA_Cash')
    b_mmf    = tr('USDT_RWA_MMF')

    bot = 0
    for vals, lbl, col in [
        (b_mmf,   'MMF units (20%)',          '#9dc3e6'),
        (b_cash,  'Cash at banks (20%)',       '#2e75b6'),
        (b_gold,  'Gold (100%)',               '#d4af37'),
        (b_sl,    'Secured loans (100%)',      '#c00000'),
        (b_other, 'Other/residual (100%)',     '#888888'),
        (b_btc,   'Bitcoin (1250%)',           '#f7931a'),
    ]:
        ax.bar(x, vals, bottom=bot, label=lbl, color=col)
        bot = bot + vals

    ax.set_title('USDT — RWA: All asset classes', fontweight='bold')
    ax.set_xticks(list(x))
    ax.set_xticklabels(q, rotation=45, ha='right', fontsize=7)
    ax.set_ylabel('RWA (USD trillions)')
    ax.legend(loc='upper left', fontsize=7)
    ax.grid(axis='y', alpha=0.3)
    ax.text(0.02, 0.97, 'T-bills & repos: 0% RW\n(not shown)',
            transform=ax.transAxes, va='top', fontsize=7,
            color='#555', style='italic')

    plt.suptitle('Basel III RWA by Asset Class — Q3 2022 to Q4 2025',
                 fontsize=12, fontweight='bold', y=1.01)
    plt.tight_layout()
    return fig


def make_rwa_ratio(rwa):
    """
    RWA / total assets over time — the central risk-density metric for RQ1.

    Shows USDC primary (look-through) and sensitivity bounds as a shaded band,
    and USDT on the same axes (right-hand scale, given its extreme values).
    """
    q = rwa['Quarter'].tolist()
    x = list(range(len(q)))

    fig, ax1 = plt.subplots(figsize=(13, 6))
    ax2 = ax1.twinx()

    usdc_primary     = rwa['USDC_RWA_Ratio_Primary'].values * 100
    usdc_sensitivity = rwa['USDC_RWA_Ratio_Sensitivity'].values * 100
    usdt             = rwa['USDT_RWA_Ratio'].values * 100

    # USDC band (primary to sensitivity)
    ax1.fill_between(x, usdc_primary, usdc_sensitivity,
                     alpha=0.25, color='#2e75b6',
                     label='USDC range (primary–sensitivity)')
    ax1.plot(x, usdc_primary, 'o-', color='#2e75b6', lw=2,
             label='USDC — primary (look-through)')
    ax1.plot(x, usdc_sensitivity, 's--', color='#1f4e79', lw=1.5,
             label='USDC — sensitivity (opaque CRF)')

    # USDT on right axis
    ax2.plot(x, usdt, '^-', color='#c00000', lw=2, label='USDT')
    ax2.set_ylabel('USDT RWA / Total Assets (%)', color='#c00000', fontsize=9)
    ax2.tick_params(axis='y', labelcolor='#c00000')

    ax1.set_ylabel('USDC RWA / Total Assets (%)', fontsize=9)
    ax1.set_xticks(x)
    ax1.set_xticklabels(q, rotation=45, ha='right', fontsize=8)
    ax1.set_title('Basel III RWA Density Ratio (RWA / Total Assets)\n'
                  'USDC vs USDT — Q3 2022 to Q4 2025',
                  fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)

    # Combine legends
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2,
               loc='upper left', fontsize=8, framealpha=0.9)

    # Annotate SVB
    if '2023-Q1' in q:
        svb_x = q.index('2023-Q1')
        ax1.axvline(svb_x, color='red', ls='--', alpha=0.4, lw=1)
        ax1.annotate('SVB\ndepeg', xy=(svb_x, ax1.get_ylim()[1] * 0.8),
                     xytext=(svb_x + 0.2, ax1.get_ylim()[1] * 0.85),
                     fontsize=7, color='red')

    plt.tight_layout()
    return fig


def print_rwa_summary(rwa):
    """Print a compact console table of key RWA metrics for all 14 quarters."""
    print("\nRWA Summary Table (key metrics)")
    print("-" * 95)
    hdr = (f"{'Quarter':<10}  {'USDC RWA Primary':>16}  {'USDC RWA Sens':>14}"
           f"  {'USDC Ratio P':>12}  {'USDC Ratio S':>12}"
           f"  {'USDT RWA':>14}  {'USDT Ratio':>10}")
    print(hdr)
    print("-" * 95)
    for _, r in rwa.iterrows():
        uc_p  = r['USDC_RWA_Total_Primary']     / 1e9
        uc_s  = r['USDC_RWA_Total_Sensitivity'] / 1e9
        ud    = r['USDT_RWA_Total']             / 1e12
        rp    = r['USDC_RWA_Ratio_Primary']     * 100
        rs    = r['USDC_RWA_Ratio_Sensitivity'] * 100
        rd    = r['USDT_RWA_Ratio']             * 100
        print(f"{r['Quarter']:<10}  {uc_p:>14.2f}bn  {uc_s:>12.2f}bn"
              f"  {rp:>10.2f}%  {rs:>10.2f}%"
              f"  {ud:>12.4f}tn  {rd:>8.2f}%")
    print("-" * 95)


# ==============================================================================
# SECTION 4: Week 2 Mon-Tue — Interest-rate sensitivity (RQ2)
# ------------------------------------------------------------------------------
# WHAT THIS SECTION PRODUCES
#   - mtm_loss_grid.csv      : MtM loss grid (WAM x shock x issuer), $ and %
#   - fig_duration_stress.png: per-issuer heatmaps + loss-vs-equity-buffer panel
#
# METHODOLOGY (duration / mark-to-market stress)
#
#   First-order price approximation for a fixed-income instrument:
#       dP/P  ~=  - Duration * dy
#   For the very short instruments stablecoin issuers hold (T-bills,
#   overnight-to-short UST repos), Macaulay duration ~= time to maturity, so we
#   proxy Duration by the Weighted-Average Maturity (WAM) expressed in years.
#   The dollar loss on the rate-sensitive book is then:
#       MtM_loss($)  =  Duration_years * dy * Sovereign_notional
#   where Sovereign_notional = T-bills (direct / in-fund) + UST repos.
#
#   This is intentionally a FIRST-ORDER (no convexity) stress. For sub-1-year
#   instruments and shocks up to +400bp, the convexity correction is immaterial
#   (well under one basis point of price), so the linear model is adequate and
#   transparent. We state this as a limitation rather than introduce a second-
#   order term the disclosed data cannot support.
#
#   WAM ASSUMPTIONS (rows of the grid):
#       30, 60, 90, 180 days  ->  0.0822, 0.1644, 0.2466, 0.4932 years.
#     Issuers do not publish a single portfolio WAM consistently across the
#     panel, so we sweep a plausible band. 30-90 days brackets the typical
#     T-bill/overnight-repo book; 180 days is a deliberately conservative
#     upper bound (longer than any issuer's disclosed maturity profile).
#
#   YIELD-SHOCK SCENARIOS (columns of the grid):
#       +100, +200, +400 bp parallel upward shifts.
#     Rising rates are the adverse direction for a long-only bond book (prices
#     fall). The historical front-end context for calibration is the FRED
#     curve (DGS1MO/3MO/6MO/1Y) on the Yield_Curve sheet: the 2022-2023
#     hiking cycle moved the 3M bill from ~0.05% (Jan 2022) to ~5.4% (mid 2023),
#     i.e. a realised +500bp+ move at the front end — so a +400bp scenario is
#     severe but not unprecedented within the panel window.
#
#   PORTFOLIO SCOPE PER ISSUER:
#       USDC : CRF + external book -> sovereign notional = T-bills_in_Fund + Repos.
#              (CRF residual cash and external cash carry no duration.)
#       USDT : SOVEREIGN SUBSET ONLY = T-bills_Direct + Total_Repos.
#              Gold, BTC and secured loans are NOT rate instruments; their
#              price risk is a separate scenario analysis (RQ-adjacent), not a
#              duration shock. We flag that USDT's portfolio volatility is
#              dominated by these commodity/crypto lines regardless of rates.
#       USDP : entire portfolio = T-bills_Direct + Reverse_Repos (no commodity
#              or crypto exposure at all; cash deposits carry no duration).
#
#   LIMITATION (documented): the duration model captures ONLY interest-rate
#   price sensitivity of sovereign fixed-income. It does not price commodity
#   (gold) or crypto (BTC) exposure. For USDT those lines dominate volatility
#   and are addressed separately; the rate stress here understates USDT's TOTAL
#   mark-to-market risk by construction and should be read only as the
#   sovereign-book component.
#
#   CITES: Basel III interest-rate-risk-in-the-banking-book principles
#          (BCBS d368, standardised duration/repricing framework); FRED series
#          DGS1MO, DGS3MO, DGS6MO, DGS1 for the historical front-end path.
# ==============================================================================

# --- Stress grid axes ---
WAM_DAYS    = [30, 60, 90, 180]                    # rows
WAM_YEARS   = [d / 365.0 for d in WAM_DAYS]
SHOCKS_BP   = [100, 200, 400]                       # columns (basis points)
SHOCKS_DEC  = [bp / 10000.0 for bp in SHOCKS_BP]    # decimal yield change

# --- USDC group-level capital reference (documented assumption) ---
# USDC reserves are held ~1:1, so the reserve over-collateralisation buffer is
# structurally near-zero and does NOT represent USDC's loss-absorbing capital.
# The relevant cushion is Circle's GROUP-LEVEL shareholder equity, computed from
# SEC EDGAR post-IPO filings (CIK 0001876042), averaging ~4.1% of total assets
# post-IPO. We plot this as a separate reference line in the loss-vs-buffer panel
# so USDC's resilience is not understated by the reserve surplus alone.
# NOTE: this is a group-entity ratio, not ring-fenced against the reserve; it is
# shown for comparability with USDT's consolidated equity, not as a claim that
# the reserve itself carries this buffer.
# (USDC_GROUP_EQUITY_PCT removed — capital is measured at reserve level only;
#  Circle group equity is reported as context in the capital bridge, not used
#  as a loss-absorbing buffer. See compute_capital_bridge.)


def _sovereign_notional(panel, usdp_panel):
    """
    Build the per-issuer rate-sensitive (sovereign fixed-income) notional for
    every quarter in scope, plus the equity buffer used for the loss-vs-equity
    comparison.

    Sovereign notional definitions:
        USDC : USDC_Tbills_in_Fund + USDC_Repos
        USDT : USDT_Tbills_Direct  + USDT_Total_Repos   (sovereign subset only)
        USDP : US_Tbills_Direct    + Reverse_Repos      (read from USDP_RAW)

    Equity_Buffer = Total_Assets - Circulation (over-collateralisation), the
    same convention used in the descriptive table. USDP fields are NaN where no
    Tier-3 quarter exists (Q3 2022).

    Returns
    -------
    pd.DataFrame, one row per quarter:
        Quarter,
        USDC_Sov_Notional, USDT_Sov_Notional, USDP_Sov_Notional,
        USDC_Total_Assets, USDT_Total_Assets, USDP_Total_Assets,
        USDC_Equity_Buffer, USDT_Equity_Buffer, USDP_Equity_Buffer
    """
    recs = []
    # USDP notionals straight from the raw Tier-3 panel (NaN-safe).
    up = usdp_panel.copy()
    up['USDP_Sov_Notional'] = (
        up['US_Tbills_Direct'].fillna(0).astype(float)
        + up['Reverse_Repos'].fillna(0).astype(float)
    )
    up = up.set_index('Quarter')

    for _, r in panel.iterrows():
        q = r['Quarter']

        # --- USDC sovereign book: T-bills in fund + repos in fund ---
        usdc_sov = float(r['USDC_Tbills_in_Fund'] or 0) + float(r['USDC_Repos'] or 0)
        usdc_ta  = float(r['USDC_Total_Assets'] or 0)
        usdc_circ = float(r['USDC_Circulation'] or 0)
        usdc_buf = usdc_ta - usdc_circ

        # --- USDT sovereign subset: direct T-bills + all repos ---
        usdt_sov = float(r['USDT_Tbills_Direct'] or 0) + float(r['USDT_Total_Repos'] or 0)
        usdt_ta  = float(r['USDT_Total_Assets'] or 0)
        usdt_circ = float(r['USDT_Circulation'] or 0)
        usdt_buf = usdt_ta - usdt_circ

        # --- USDP whole portfolio (read from raw Tier-3 panel) ---
        if q in up.index:
            usdp_sov  = float(up.loc[q, 'USDP_Sov_Notional'])
            usdp_ta   = float(up.loc[q, 'Total_Assets'])
            usdp_circ = float(up.loc[q, 'USDP_Circulation'])
            usdp_buf  = usdp_ta - usdp_circ
        else:
            usdp_sov = float('nan'); usdp_ta = float('nan'); usdp_buf = float('nan')

        recs.append({
            'Quarter': q,
            'USDC_Sov_Notional': usdc_sov,
            'USDT_Sov_Notional': usdt_sov,
            'USDP_Sov_Notional': usdp_sov,
            'USDC_Total_Assets': usdc_ta,
            'USDT_Total_Assets': usdt_ta,
            'USDP_Total_Assets': usdp_ta,
            'USDC_Equity_Buffer': usdc_buf,
            'USDT_Equity_Buffer': usdt_buf,
            'USDP_Equity_Buffer': usdp_buf,
        })
    return pd.DataFrame(recs)


def compute_duration_stress(panel, usdp_panel):
    """
    Compute the mark-to-market loss grid for all three issuers.

    For each issuer we use the Q4 2025 (latest) sovereign notional as the
    headline balance sheet for the grid, and also retain the full quarterly
    path of notionals so the loss-vs-equity comparison can be plotted over time.

    Returns
    -------
    grid : pd.DataFrame  (long format, one row per issuer x WAM x shock)
        columns: Issuer, WAM_days, WAM_years, Shock_bp, Sov_Notional_bn,
                 MtM_Loss_bn, MtM_Loss_pct_of_assets, MtM_Loss_pct_of_sov
    notion : pd.DataFrame  (per-quarter sovereign notionals + equity buffers)
    """
    notion = _sovereign_notional(panel, usdp_panel)

    # Headline balance sheet = latest quarter (Q4 2025).
    latest = notion[notion['Quarter'] == SCOPE_LAST_Q].iloc[0]

    issuers = {
        'USDC': ('USDC_Sov_Notional', 'USDC_Total_Assets'),
        'USDT': ('USDT_Sov_Notional', 'USDT_Total_Assets'),
        'USDP': ('USDP_Sov_Notional', 'USDP_Total_Assets'),
    }

    rows = []
    for issuer, (sov_col, ta_col) in issuers.items():
        sov = float(latest[sov_col])
        ta  = float(latest[ta_col])
        for d, yrs in zip(WAM_DAYS, WAM_YEARS):
            for bp, dy in zip(SHOCKS_BP, SHOCKS_DEC):
                loss = yrs * dy * sov          # dollar MtM loss (>0 = loss)
                rows.append({
                    'Issuer': issuer,
                    'WAM_days': d,
                    'WAM_years': round(yrs, 4),
                    'Shock_bp': bp,
                    'Sov_Notional_bn': round(sov / 1e9, 4),
                    'MtM_Loss_bn': round(loss / 1e9, 6),
                    'MtM_Loss_pct_of_assets': round(loss / ta * 100, 5) if ta else float('nan'),
                    'MtM_Loss_pct_of_sov': round(yrs * dy * 100, 5),
                })
    grid = pd.DataFrame(rows)
    return grid, notion


def make_duration_stress_figure(grid, notion):
    """
    Four-panel figure:
      Panels 1-3 : per-issuer heatmap of MtM loss as % of total assets
                   (rows = WAM, cols = yield shock). Latest-quarter balance sheet.
      Panel 4    : loss-vs-equity-buffer comparison at the WORST cell
                   (180-day WAM, +400bp) across the quarterly path — bars show
                   each issuer's worst-case sovereign MtM loss against its
                   over-collateralisation buffer, both as % of circulation,
                   at Q4 2025.
    """
    fig, axes = plt.subplots(1, 4, figsize=(20, 5.2))

    issuer_order = ['USDC', 'USDT', 'USDP']
    issuer_titles = {
        'USDC': 'USDC (CRF sovereign book)',
        'USDT': 'USDT (sovereign subset only)',
        'USDP': 'USDP (whole portfolio)',
    }

    # Shared colour scale across the three heatmaps for honest comparison.
    vmax = grid['MtM_Loss_pct_of_assets'].max()

    for ax, issuer in zip(axes[:3], issuer_order):
        sub = grid[grid['Issuer'] == issuer]
        mat = (sub.pivot(index='WAM_days', columns='Shock_bp',
                         values='MtM_Loss_pct_of_assets')
                  .reindex(index=WAM_DAYS, columns=SHOCKS_BP))
        im = ax.imshow(mat.values, cmap='OrRd', aspect='auto',
                       vmin=0, vmax=vmax)
        ax.set_xticks(range(len(SHOCKS_BP)))
        ax.set_xticklabels([f'+{b}bp' for b in SHOCKS_BP])
        ax.set_yticks(range(len(WAM_DAYS)))
        ax.set_yticklabels([f'{d}d' for d in WAM_DAYS])
        ax.set_xlabel('Yield shock')
        if issuer == 'USDC':
            ax.set_ylabel('Assumed WAM')
        ax.set_title(issuer_titles[issuer], fontsize=10)
        for i in range(mat.shape[0]):
            for j in range(mat.shape[1]):
                v = mat.values[i, j]
                ax.text(j, i, f'{v:.3f}%', ha='center', va='center',
                        fontsize=8,
                        color='white' if v > vmax * 0.6 else '#333')
        ax.spines['top'].set_visible(True)
        ax.spines['right'].set_visible(True)

    fig.colorbar(im, ax=axes[:3], fraction=0.022, pad=0.015,
                 label='MtM loss (% of total assets)')

    # --- Panel 4: worst-case loss vs equity buffer, Q4 2025 ---
    ax4 = axes[3]
    latest = notion[notion['Quarter'] == SCOPE_LAST_Q].iloc[0]
    worst_yrs = max(WAM_YEARS)        # 180 days
    worst_dy  = max(SHOCKS_DEC)       # +400bp

    issuers = ['USDC', 'USDT', 'USDP']
    # Circulation = Total_Assets - Equity_Buffer (buffer defined as TA - Circ).
    circ_map = {
        'USDC': float(latest['USDC_Total_Assets']) - float(latest['USDC_Equity_Buffer']),
        'USDT': float(latest['USDT_Total_Assets']) - float(latest['USDT_Equity_Buffer']),
        'USDP': (float(latest['USDP_Total_Assets']) - float(latest['USDP_Equity_Buffer'])
                 if latest['USDP_Total_Assets'] == latest['USDP_Total_Assets'] else float('nan')),
    }
    sov_map = {
        'USDC': float(latest['USDC_Sov_Notional']),
        'USDT': float(latest['USDT_Sov_Notional']),
        'USDP': float(latest['USDP_Sov_Notional']),
    }
    buf_map = {
        'USDC': float(latest['USDC_Equity_Buffer']),
        'USDT': float(latest['USDT_Equity_Buffer']),
        'USDP': float(latest['USDP_Equity_Buffer']),
    }

    loss_pct, buf_pct = [], []
    for iss in issuers:
        circ = circ_map[iss]
        loss = worst_yrs * worst_dy * sov_map[iss]
        loss_pct.append(loss / circ * 100 if circ else float('nan'))
        buf_pct.append(buf_map[iss] / circ * 100 if circ else float('nan'))

    x = range(len(issuers))
    w = 0.38
    ax4.bar([i - w/2 for i in x], buf_pct, width=w,
            label='Reserve over-collat. buffer', color='#2a9d8f')
    ax4.bar([i + w/2 for i in x], loss_pct, width=w,
            label='Worst-case sovereign MtM loss\n(180d WAM, +400bp)', color='#e76f51')

    # USDC group-level equity shown for CONTEXT ONLY (not counted as token-holder
    # protection). Circle's consolidated equity is legally walled off from the
    # reserve; the reserve surplus (~4-45 bps) is the buffer that actually backs
    # the token. Annotated to make the exclusion explicit rather than implied.
    usdc_i = issuers.index('USDC')
    ax4.annotate('Circle group equity is\nlegally separate (excluded)',
                 xy=(usdc_i, 0.5), fontsize=6.5, color='#777',
                 ha='center', va='bottom', style='italic')

    ax4.set_xticks(list(x))
    ax4.set_xticklabels(issuers)
    ax4.set_ylabel('% of circulation')
    ax4.set_title('Loss vs buffer at Q4 2025', fontsize=10)
    ax4.legend(fontsize=6.5, loc='upper right')
    ax4.spines['top'].set_visible(True)
    ax4.spines['right'].set_visible(True)
    for i in x:
        if buf_pct[i] == buf_pct[i]:
            ax4.text(i - w/2, buf_pct[i], f'{buf_pct[i]:.2f}',
                     ha='center', va='bottom', fontsize=7)
        if loss_pct[i] == loss_pct[i]:
            ax4.text(i + w/2, loss_pct[i], f'{loss_pct[i]:.3f}',
                     ha='center', va='bottom', fontsize=7)

    fig.suptitle('RQ2 — Interest-rate (duration) stress on stablecoin sovereign reserve books',
                 fontsize=13, y=1.02)
    fig.subplots_adjust(wspace=0.45)
    return fig


def print_duration_summary(grid, notion):
    """Console summary of the duration stress for the session log."""
    print("\n  Duration stress grid (latest quarter = %s balance sheet):" % SCOPE_LAST_Q)
    for issuer in ['USDC', 'USDT', 'USDP']:
        sub = grid[grid['Issuer'] == issuer]
        sov = sub['Sov_Notional_bn'].iloc[0]
        worst = sub[(sub['WAM_days'] == 180) & (sub['Shock_bp'] == 400)].iloc[0]
        mild  = sub[(sub['WAM_days'] == 30) & (sub['Shock_bp'] == 100)].iloc[0]
        print(f"    {issuer}: sovereign notional ${sov:.2f}bn | "
              f"mildest (30d,+100bp) {mild['MtM_Loss_pct_of_assets']:.4f}% of assets | "
              f"worst (180d,+400bp) {worst['MtM_Loss_pct_of_assets']:.4f}% of assets "
              f"(${worst['MtM_Loss_bn']:.3f}bn)")


# ------------------------------------------------------------------------------
# SECTION 4 (cont.) — Time-series duration stress across all 14 quarters
# ------------------------------------------------------------------------------
# The grid above fixes the balance sheet at the latest quarter. This block
# instead FIXES THE STRESS and sweeps the QUARTER, so we can see how each
# issuer's rate vulnerability evolved as its reserve book grew/shrank.
#
# For a chosen (WAM, shock) pair we recompute the MtM loss on every quarter's
# sovereign notional, expressed as % of that quarter's total assets. We report
# the headline adverse pair (180-day WAM, +400bp) and a central pair
# (90-day WAM, +200bp) so the figure shows both a conservative ceiling and a
# plausible mid-scenario.
# ------------------------------------------------------------------------------

# Time-series stress pairs: (WAM_days, shock_bp, label)
TS_PAIRS = [
    (180, 400, 'Adverse ceiling (180d, +400bp)'),
    (90,  200, 'Central (90d, +200bp)'),
]


def compute_duration_stress_timeseries(panel, usdp_panel):
    """
    Recompute MtM loss for every quarter in scope, for each TS_PAIRS scenario.

    Returns
    -------
    ts : pd.DataFrame  (long format)
        columns: Quarter, Issuer, WAM_days, Shock_bp, Scenario,
                 Sov_Notional_bn, Total_Assets_bn, MtM_Loss_bn,
                 MtM_Loss_pct_of_assets
        USDP rows are NaN where no Tier-3 quarter exists (Q3 2022).
    """
    notion = _sovereign_notional(panel, usdp_panel)
    issuers = {
        'USDC': ('USDC_Sov_Notional', 'USDC_Total_Assets'),
        'USDT': ('USDT_Sov_Notional', 'USDT_Total_Assets'),
        'USDP': ('USDP_Sov_Notional', 'USDP_Total_Assets'),
    }
    rows = []
    for _, r in notion.iterrows():
        for issuer, (sov_col, ta_col) in issuers.items():
            sov = float(r[sov_col]) if pd.notna(r[sov_col]) else float('nan')
            ta  = float(r[ta_col])  if pd.notna(r[ta_col])  else float('nan')
            for d, bp, label in TS_PAIRS:
                yrs = d / 365.0
                dy  = bp / 10000.0
                loss = yrs * dy * sov
                rows.append({
                    'Quarter': r['Quarter'],
                    'Issuer': issuer,
                    'WAM_days': d,
                    'Shock_bp': bp,
                    'Scenario': label,
                    'Sov_Notional_bn': round(sov / 1e9, 4) if pd.notna(sov) else float('nan'),
                    'Total_Assets_bn': round(ta / 1e9, 4) if pd.notna(ta) else float('nan'),
                    'MtM_Loss_bn': round(loss / 1e9, 6) if pd.notna(loss) else float('nan'),
                    'MtM_Loss_pct_of_assets': round(loss / ta * 100, 5) if (pd.notna(loss) and ta) else float('nan'),
                })
    return pd.DataFrame(rows)


def make_duration_timeseries_figure(ts):
    """
    Two-panel time-series figure (MtM loss as % of total assets, per quarter):
      Panel A : Adverse ceiling (180d, +400bp)
      Panel B : Central (90d, +200bp)
    Each panel plots USDC, USDT, USDP as lines across the 14 quarters.
    USDP shows a gap at Q3 2022 (no Tier-3 row).
    """
    quarters = sorted(ts['Quarter'].unique(), key=q_to_num)
    x = list(range(len(quarters)))
    qidx = {q: i for i, q in enumerate(quarters)}

    colors = {'USDC': '#1f77b4', 'USDT': '#d62728', 'USDP': '#2a9d8f'}
    markers = {'USDC': 'o', 'USDT': 's', 'USDP': '^'}

    scenarios = [p[2] for p in TS_PAIRS]
    fig, axes = plt.subplots(1, 2, figsize=(16, 5.2), sharex=True)

    for ax, scen in zip(axes, scenarios):
        sub = ts[ts['Scenario'] == scen]
        for issuer in ['USDC', 'USDT', 'USDP']:
            iss = sub[sub['Issuer'] == issuer].copy()
            iss['xi'] = iss['Quarter'].map(qidx)
            iss = iss.sort_values('xi')
            ax.plot(iss['xi'], iss['MtM_Loss_pct_of_assets'],
                    marker=markers[issuer], markersize=5, linewidth=1.8,
                    color=colors[issuer], label=issuer)
        ax.set_title(scen, fontsize=11)
        ax.set_xticks(x)
        ax.set_xticklabels(quarters, rotation=45, ha='right', fontsize=8)
        ax.set_ylabel('MtM loss (% of total assets)')
        ax.grid(True, alpha=0.25, linewidth=0.6)
        ax.legend(fontsize=8, loc='upper left')
        ax.set_ylim(bottom=0)

    fig.suptitle('RQ2 — Sovereign MtM loss over time (fixed stress, 14-quarter panel)',
                 fontsize=13, y=1.01)
    fig.tight_layout()
    return fig


# ------------------------------------------------------------------------------
# SECTION 4 (cont.) — Historical rate-shock context from the Yield_Curve sheet
# ------------------------------------------------------------------------------
# Calibration evidence: how large were realised front-end moves inside the
# sample window? We compute the trailing rolling change in the 3-month bill
# (DGS3MO) and report how often a +400bp / 12-month move actually occurred.
# This turns "a +400bp shock is plausible" into a measured claim grounded in
# the issuers' own observation window.
#
# NOTE ON SCOPE: the Yield_Curve sheet spans 2020-2025 only, so this measures
# the 2022-23 hiking cycle directly. Older documented episodes (1979, 1980-81,
# 1994) lie outside the sheet and are cited from the historical record in the
# session write-up; they should be pinned to a long FRED series (e.g. TB3MS)
# before being quoted with exact basis points in the thesis.
# ------------------------------------------------------------------------------

def load_yield_curve():
    """Load the Yield_Curve sheet, parse dates, return a clean daily frame."""
    yc = pd.read_excel(EXCEL_PATH, sheet_name='Yield_Curve')
    yc = yc.rename(columns={
        'DGS1MO (%)': 'DGS1MO', 'DGS3MO (%)': 'DGS3MO',
        'DGS6MO (%)': 'DGS6MO', 'DGS1Y (%)': 'DGS1Y',
    })
    yc['Date'] = pd.to_datetime(yc['Date'])
    return yc.sort_values('Date').reset_index(drop=True)


def compute_historical_shock_context(yc, window_trading_days=252, threshold_bp=400):
    """
    Measure realised trailing rate rises in the 3-month bill.

    Returns a dict of headline statistics for the write-up.
    """
    s = yc.set_index('Date')['DGS3MO'].dropna()
    rolled = (s - s.shift(window_trading_days)).dropna()        # change in %, vs ~1y ago
    over = rolled[rolled >= threshold_bp / 100.0]               # threshold in %-points

    cycle = s['2021-12-01':'2023-12-31']
    stats = {
        'series_span': (s.index.min().date(), s.index.max().date()),
        'cycle_low_pct': float(cycle.min()),
        'cycle_low_date': cycle.idxmin().date(),
        'cycle_high_pct': float(cycle.max()),
        'cycle_high_date': cycle.idxmax().date(),
        'cycle_move_bp': float((cycle.max() - cycle.min()) * 100),
        'max_trailing_12m_bp': float(rolled.max() * 100),
        'max_trailing_12m_date': rolled.idxmax().date(),
        'days_over_threshold': int(len(over)),
        'over_first': over.index.min().date() if len(over) else None,
        'over_last': over.index.max().date() if len(over) else None,
        'window_trading_days': window_trading_days,
        'threshold_bp': threshold_bp,
    }
    return stats, s, rolled


def make_historical_context_figure(s, rolled, threshold_bp=400):
    """
    Two-panel figure:
      A : the 3-month bill level over the sample (2020-2025).
      B : trailing ~12-month change in the 3-month bill, with the +400bp
          stress threshold drawn as a reference line. Shading where the
          realised move met or exceeded the threshold.
    """
    fig, axes = plt.subplots(2, 1, figsize=(13, 6.5), sharex=True)

    axes[0].plot(s.index, s.values, color='#1f3864', linewidth=1.4)
    axes[0].set_ylabel('3M T-bill yield (%)')
    axes[0].set_title('DGS3MO level — 2020 to 2025 (sample window)', fontsize=11)
    axes[0].grid(True, alpha=0.25, linewidth=0.6)

    axes[1].plot(rolled.index, rolled.values * 100, color='#c1121f', linewidth=1.4)
    axes[1].axhline(threshold_bp, color='#333', linestyle='--', linewidth=1.3,
                    label=f'+{threshold_bp}bp stress threshold')
    axes[1].fill_between(rolled.index, threshold_bp, rolled.values * 100,
                         where=(rolled.values * 100 >= threshold_bp),
                         color='#f4a261', alpha=0.5,
                         label='realised \u2265 +%dbp' % threshold_bp)
    axes[1].set_ylabel('Trailing 12m change (bp)')
    axes[1].set_title('Realised trailing ~12-month rise in the 3M bill vs the stress threshold',
                      fontsize=11)
    axes[1].grid(True, alpha=0.25, linewidth=0.6)
    axes[1].legend(fontsize=8, loc='upper left')

    fig.suptitle('RQ2 — Historical front-end rate-shock context (in-sample calibration)',
                 fontsize=13, y=1.00)
    fig.tight_layout()
    return fig


# ------------------------------------------------------------------------------
# SECTION 4 (cont.) — Long-history rate-shock frequency (TB3MS, 1934-present)
# ------------------------------------------------------------------------------
# The in-sample Yield_Curve sheet only spans 2020-2025. To turn the "+400bp is
# a once-a-decade-ish tail" claim into a MEASURED count, we read the long FRED
# 3-month bill series TB3MS (monthly, discount basis, from 1934) and count
# distinct episodes where the trailing 12-month rise met or exceeded +400bp.
#
# BASIS NOTE: TB3MS is quoted on a DISCOUNT basis, whereas the in-sample
# DGS3MO is investment (constant-maturity) basis. The two differ by a small,
# roughly constant wedge. For a FREQUENCY COUNT of large *changes* this wedge
# is immaterial (it nets out in the difference), so the episode count is robust
# to the basis choice even though absolute levels sit a touch lower. Investment
# basis only reaches ~1982 (DGS3MO) and daily discount only to 1954 (DTB3), so
# TB3MS is the only series reaching the 1973 and 1979-81 episodes.
#
# The TB3MS sheet is optional: if it is absent from the workbook, this block
# is skipped and the in-sample calibration alone is used.
# ------------------------------------------------------------------------------

SHOCK_THRESHOLD_BP = 400          # the adverse scenario we benchmark against
EPISODE_GAP_MONTHS = 3            # consecutive months within this gap = one episode


def load_long_bill_history(sheet_name='TB3MS'):
    """
    Load the long monthly 3-month bill series (TB3MS) from the master
    workbook's TB3MS sheet. Auto-detects the rate column and flags whether
    the series is discount or investment basis from its name, so a
    swapped-in DGS3MO sheet would also work.

    Returns (series, basis_label) or (None, None) if the sheet is absent.
    """
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)
    except Exception:
        return None, None
    date_col = [c for c in df.columns if 'date' in c.lower()][0]
    rate_col = [c for c in df.columns if c != date_col][0]
    df[date_col] = pd.to_datetime(df[date_col])
    s = df.set_index(date_col)[rate_col].dropna().sort_index()
    basis = 'investment (CMT)' if 'DGS' in rate_col.upper() else 'discount'
    return s, basis


def count_shock_episodes(s, threshold_bp=SHOCK_THRESHOLD_BP, horizon_months=12,
                         gap_months=EPISODE_GAP_MONTHS):
    """
    Count distinct episodes where the trailing `horizon_months` rise in the
    bill rate met or exceeded `threshold_bp`. Consecutive qualifying months
    within `gap_months` of each other are merged into one episode.

    Returns
    -------
    stats : dict, episodes : list of dicts, rolled : pd.Series (%-points)
    """
    rolled = (s - s.shift(horizon_months)).dropna()
    thr = threshold_bp / 100.0
    over = rolled[rolled >= thr]

    episodes = []
    if len(over):
        dts = list(over.index)
        start = prev = dts[0]
        for d in dts[1:]:
            months_apart = (d.year - prev.year) * 12 + (d.month - prev.month)
            if months_apart > gap_months:
                seg = rolled[start:prev]
                episodes.append({'start': start.date(), 'end': prev.date(),
                                 'peak_bp': round(seg.max() * 100),
                                 'peak_date': seg.idxmax().date(),
                                 'n_months': len(seg)})
                start = d
            prev = d
        seg = rolled[start:prev]
        episodes.append({'start': start.date(), 'end': prev.date(),
                         'peak_bp': round(seg.max() * 100),
                         'peak_date': seg.idxmax().date(),
                         'n_months': len(seg)})

    decade = {}
    for ep in episodes:
        dkey = (ep['peak_date'].year // 10) * 10
        decade[dkey] = decade.get(dkey, 0) + 1

    span_years = (s.index.max() - s.index.min()).days / 365.25
    stats = {
        'span': (s.index.min().date(), s.index.max().date()),
        'span_years': round(span_years, 1),
        'n_months_over': int(len(over)),
        'n_episodes': len(episodes),
        'largest_move_bp': round(rolled.max() * 100),
        'largest_move_date': rolled.idxmax().date(),
        'threshold_bp': threshold_bp,
        'horizon_months': horizon_months,
        'decade_counts': dict(sorted(decade.items())),
        'per_decade_rate': round(len(episodes) / (span_years / 10.0), 2),
    }
    return stats, episodes, rolled


def make_long_history_figure(s, rolled, basis_label, threshold_bp=SHOCK_THRESHOLD_BP):
    """
    Two-panel long-history figure:
      A : the bill level since 1934.
      B : trailing 12-month change vs the +threshold line, with qualifying
          episodes shaded. Defense exhibit for the frequency claim.
    """
    fig, axes = plt.subplots(2, 1, figsize=(13, 6.5), sharex=True)

    axes[0].plot(s.index, s.values, color='#1f3864', linewidth=0.9)
    axes[0].set_ylabel('3M bill rate (%)')
    axes[0].set_title(f'TB3MS \u2014 3-month Treasury bill, {basis_label} basis, '
                      f'{s.index.min().year}\u2013{s.index.max().year}', fontsize=11)
    axes[0].grid(True, alpha=0.25, linewidth=0.6)

    axes[1].plot(rolled.index, rolled.values * 100, color='#c1121f', linewidth=0.9)
    axes[1].axhline(threshold_bp, color='#333', linestyle='--', linewidth=1.3,
                    label=f'+{threshold_bp}bp stress threshold')
    axes[1].fill_between(rolled.index, threshold_bp, rolled.values * 100,
                         where=(rolled.values * 100 >= threshold_bp),
                         color='#f4a261', alpha=0.7,
                         label=f'realised \u2265 +{threshold_bp}bp / 12m')
    axes[1].set_ylabel('Trailing 12m change (bp)')
    axes[1].set_title('Realised trailing 12-month rise vs the stress threshold',
                      fontsize=11)
    axes[1].grid(True, alpha=0.25, linewidth=0.6)
    axes[1].legend(fontsize=8, loc='upper left')

    fig.suptitle('RQ2 \u2014 How often has a +400bp front-end shock actually occurred? '
                 '(TB3MS, ~90-year record)', fontsize=12.5, y=1.00)
    fig.tight_layout()
    return fig




# ==============================================================================
# SECTION 5: Week 2 Wed-Thu — Redemption / LCR liquidity stress (RQ3)
# ------------------------------------------------------------------------------
# WHAT THIS SECTION PRODUCES
#   - lcr_stress_grid.csv         : LCR(%) for 3 issuers x 3 outflow scenarios
#                                   x 2 HQLA treatments, on the latest quarter.
#   - lcr_timeseries.csv          : LCR(%) for all 14 quarters x 3 issuers
#                                   x 3 outflow scenarios x 2 treatments.
#   - fig_lcr_stress.png          : per-issuer heatmap (rows = outflow scenario,
#                                   cols = quarter), cells = LCR%, pass/fail >=100.
#   - fig_lcr_vs_rate_combined.png: the RQ2 x RQ3 intersection — a redemption run
#                                   that FORCES a pre-maturity sale and so
#                                   REALISES the otherwise-paper MtM loss.
#
# METHODOLOGY (Basel III LCR applied analogically)
#
#   The Liquidity Coverage Ratio is:
#       LCR  =  HQLA stock  /  Net Cash Outflow over a 30-day stress
#   with a regulatory pass threshold of LCR >= 100%.
#
#   For a stablecoin issuer the "30-day net cash outflow" has a natural analogue:
#   a redemption run. Holders present tokens for par redemption; the issuer must
#   liquidate reserves to meet them. We do NOT model offsetting inflows (an
#   issuer has no loan book rolling in), so Net Cash Outflow = gross redemptions
#       Net Cash Outflow  =  outflow_rate  x  Circulation
#   tested at three severities:
#       20% / 40% / 60% of circulation redeemed inside the stress window.
#   The 40% figure brackets the realised USDC experience: USDC circulation fell
#   ~$11bn (~26% of supply) in the days around the March 2023 SVB depeg, so a
#   40% run is severe-but-witnessed and 60% is a deliberate tail.
#
#   HQLA STOCK (numerator) — two treatments, mirroring the RQ1 RWA split:
#     PRIMARY  (look-through): the Circle Reserve Fund is decomposed into its
#              holdings. T-bills and overnight UST-collateralised repos are
#              Level 1 HQLA at a 0% haircut. This is the thesis's primary lens
#              and is consistent with how the fund actually invests.
#     SENSITIVITY (opaque MMF): the fund NAV is treated as an MMF unit. MMF
#              units are NOT a named HQLA category in Basel III (LCR30.41-30.45
#              enumerate Level 2B as qualifying RMBS, corporates and equities —
#              not fund units), so as the most conservative defensible proxy we
#              admit the NAV as Level 2B at a 50% haircut. This is an UPPER bound
#              on the penalty for opacity, not a claim that MMF units qualify.
#
#   HQLA ELIGIBILITY BY LINE:
#       T-bills (direct or in-fund)     -> Level 1, 0% haircut.
#       UST repos (overnight)           -> Level 1, 0% haircut (sovereign-coll.).
#       MMF NAV (opaque treatment only) -> Level 2B, 50% haircut.
#       Cash at commercial banks        -> NOT counted as HQLA. Basel grants
#               Level 1 only to central-bank reserves and sovereign claims; a
#               deposit at a commercial/custodian bank is itself a claim that can
#               freeze in a run (exactly the SVB failure mode), so it is excluded
#               from the numerator. This is deliberately conservative and is the
#               single biggest driver of the USDC transition-quarter result.
#       Gold, BTC, secured loans        -> NOT HQLA (excluded from numerator).
#
#   PER-ISSUER SCOPE:
#       USDC : L1 = USDC_Tbills_in_Fund + USDC_Repos (look-through);
#              opaque = USDC_MMF_NAV as L2B@50%. External cash excluded.
#       USDT : L1 = USDT_Tbills_Direct + USDT_Total_Repos; opaque adds
#              USDT_MMF as L2B@50% (zero across the panel, so LT == OP for USDT).
#              USDT_Cash, BTC, Gold, Sec_Loans excluded from HQLA.
#       USDP : narrow bank — L1 = US_Tbills_Direct + Reverse_Repos, whole
#              sovereign book HQLA-eligible; Cash_Deposits excluded. Single
#              treatment (no fund to look through).
#
#   RQ2 x RQ3 INTERSECTION (the binding scenario):
#       MtM loss from a rate shock (RQ2) is PAPER while instruments are held to
#       maturity. It only becomes a REALISED loss if a redemption run forces the
#       issuer to sell sovereign paper before maturity. The combined panel
#       therefore overlays, per quarter, the +400bp/180d MtM loss as a % of HQLA
#       against the post-run HQLA coverage, isolating the one state of the world
#       where duration risk actually bites.
#
#   LIMITATIONS (documented):
#     - No inflow offset and no run-off categorisation by holder type (the
#       disclosures do not support a behavioural split); the outflow is a flat
#       fraction of circulation, which is conservative.
#     - Repo "0% haircut, Level 1" assumes the collateral is delivered/tri-party
#       and instantly re-saleable; a gross repo freeze is not modelled.
#     - Cash exclusion is a binary on/off; a partial-recovery assumption would
#       lift the transition-quarter USDC ratios but is not adopted to keep the
#       numerator strictly Basel-eligible.
#
#   CITES: BCBS Basel III LCR standard, LCR30 (HQLA definition and haircuts),
#          esp. LCR30.41-30.45 on Level 2B eligibility; the March 2023 USDC
#          depeg (SVB exposure) as the empirical run benchmark.
# ==============================================================================

# --- LCR stress axes ---
OUTFLOW_SCENARIOS = [0.20, 0.40, 0.60]      # fraction of circulation redeemed
HQLA_HAIRCUT_L1   = 0.00                     # Level 1 (T-bills, UST repos)
HQLA_HAIRCUT_L2B  = 0.50                     # Level 2B proxy (opaque MMF NAV)
LCR_PASS          = 100.0                    # regulatory pass threshold (%)

# Stress pair used for the RQ2 x RQ3 combined panel (adverse ceiling).
LCR_RATE_WAM_DAYS = 180
LCR_RATE_SHOCK_BP = 400


def _usdc_hqla(row, treatment='lookthrough'):
    """USDC HQLA under either treatment. Returns components + admitted HQLA $."""
    tbills = float(row['USDC_Tbills_in_Fund'] or 0)
    repos  = float(row['USDC_Repos'] or 0)
    mmf    = float(row['USDC_MMF_NAV'] or 0)
    ext    = float(row['USDC_External_Cash'] or 0)
    crf_residual = max(mmf - tbills - repos, 0.0) if mmf > 0 else 0.0
    if treatment == 'lookthrough':
        l1, l2b = tbills + repos, 0.0
        hqla = l1 * (1 - HQLA_HAIRCUT_L1)
    else:  # opaque: entire CRF NAV admitted as Level 2B at 50% haircut
        l1, l2b = 0.0, mmf
        hqla = l2b * (1 - HQLA_HAIRCUT_L2B)
    return {'L1': l1, 'L2B': l2b, 'CRF_Residual': crf_residual,
            'Ext_Cash': ext, 'HQLA': hqla}


def _usdt_hqla(row, treatment='lookthrough'):
    """USDT HQLA. MMF is zero across the panel so LT and OP coincide here."""
    tbills = float(row['USDT_Tbills_Direct'] or 0)
    repos  = float(row['USDT_Total_Repos'] or 0)
    mmf    = float(row['USDT_MMF'] or 0)
    cash   = float(row['USDT_Cash'] or 0)
    l1 = tbills + repos
    if treatment == 'lookthrough':
        l2b = 0.0
        hqla = l1 * (1 - HQLA_HAIRCUT_L1)
    else:
        l2b = mmf
        hqla = l1 * (1 - HQLA_HAIRCUT_L1) + l2b * (1 - HQLA_HAIRCUT_L2B)
    return {'L1': l1, 'L2B': l2b, 'Cash': cash, 'HQLA': hqla}


def _usdp_hqla(row):
    """USDP HQLA — narrow bank, single treatment. NaN lines treated as zero."""
    tbills = float(row['US_Tbills_Direct']) if pd.notna(row['US_Tbills_Direct']) else 0.0
    repos  = float(row['Reverse_Repos'])    if pd.notna(row['Reverse_Repos'])    else 0.0
    cash   = float(row['Cash_Deposits'])    if pd.notna(row['Cash_Deposits'])    else 0.0
    l1 = tbills + repos
    return {'L1': l1, 'Cash': cash, 'HQLA': l1 * (1 - HQLA_HAIRCUT_L1)}


def compute_lcr_timeseries(panel, usdp_panel):
    """
    Long-format LCR table over all 14 scope quarters.

    One row per (Quarter x Outflow_pct). Each row carries every issuer's
    admitted HQLA, net outflow, and LCR% under both treatments (USDP single).

    Returns
    -------
    pd.DataFrame
    """
    up = usdp_panel.set_index('Quarter')
    recs = []
    for _, r in panel.iterrows():
        q = r['Quarter']
        usdc_circ = float(r['USDC_Circulation'] or 0)
        usdt_circ = float(r['USDT_Circulation'] or 0)
        uc_lt = _usdc_hqla(r, 'lookthrough'); uc_op = _usdc_hqla(r, 'opaque')
        ut_lt = _usdt_hqla(r, 'lookthrough'); ut_op = _usdt_hqla(r, 'opaque')
        if q in up.index:
            up_row = up.loc[q]
            usdp_circ = float(up_row['USDP_Circulation'])
            upd = _usdp_hqla(up_row)
        else:
            usdp_circ = float('nan'); upd = {'HQLA': float('nan')}

        for o in OUTFLOW_SCENARIOS:
            usdc_out = usdc_circ * o
            usdt_out = usdt_circ * o
            usdp_out = usdp_circ * o if usdp_circ == usdp_circ else float('nan')
            recs.append({
                'Quarter': q,
                'Outflow_pct': int(round(o * 100)),
                'USDC_HQLA_LT_bn': round(uc_lt['HQLA'] / 1e9, 4),
                'USDC_HQLA_OP_bn': round(uc_op['HQLA'] / 1e9, 4),
                'USDC_Outflow_bn': round(usdc_out / 1e9, 4),
                'USDC_LCR_LT': round(uc_lt['HQLA'] / usdc_out * 100, 3) if usdc_out else float('nan'),
                'USDC_LCR_OP': round(uc_op['HQLA'] / usdc_out * 100, 3) if usdc_out else float('nan'),
                'USDT_HQLA_LT_bn': round(ut_lt['HQLA'] / 1e9, 4),
                'USDT_HQLA_OP_bn': round(ut_op['HQLA'] / 1e9, 4),
                'USDT_Outflow_bn': round(usdt_out / 1e9, 4),
                'USDT_LCR_LT': round(ut_lt['HQLA'] / usdt_out * 100, 3) if usdt_out else float('nan'),
                'USDT_LCR_OP': round(ut_op['HQLA'] / usdt_out * 100, 3) if usdt_out else float('nan'),
                'USDP_HQLA_bn': round(upd['HQLA'] / 1e9, 4) if upd['HQLA'] == upd['HQLA'] else float('nan'),
                'USDP_Outflow_bn': round(usdp_out / 1e9, 4) if (usdp_out == usdp_out) else float('nan'),
                'USDP_LCR': round(upd['HQLA'] / usdp_out * 100, 3)
                            if (usdp_out == usdp_out and usdp_out) else float('nan'),
            })
    return pd.DataFrame(recs)


def compute_lcr_grid(panel, usdp_panel, quarter=None):
    """
    Headline LCR grid on a single quarter (default = latest in scope).

    Long format: one row per Issuer x Outflow x Treatment.
    For USDP the two treatments are identical (no fund to look through); we emit
    a single 'single' treatment row to avoid implying a sensitivity that does
    not exist.
    """
    quarter = quarter or SCOPE_LAST_Q
    ts = compute_lcr_timeseries(panel, usdp_panel)
    snap = ts[ts['Quarter'] == quarter].copy()
    rows = []
    for _, r in snap.iterrows():
        o = r['Outflow_pct']
        rows.append({'Issuer': 'USDC', 'Quarter': quarter, 'Outflow_pct': o,
                     'Treatment': 'lookthrough', 'HQLA_bn': r['USDC_HQLA_LT_bn'],
                     'Outflow_bn': r['USDC_Outflow_bn'], 'LCR_pct': r['USDC_LCR_LT'],
                     'Pass': r['USDC_LCR_LT'] >= LCR_PASS})
        rows.append({'Issuer': 'USDC', 'Quarter': quarter, 'Outflow_pct': o,
                     'Treatment': 'opaque', 'HQLA_bn': r['USDC_HQLA_OP_bn'],
                     'Outflow_bn': r['USDC_Outflow_bn'], 'LCR_pct': r['USDC_LCR_OP'],
                     'Pass': r['USDC_LCR_OP'] >= LCR_PASS})
        rows.append({'Issuer': 'USDT', 'Quarter': quarter, 'Outflow_pct': o,
                     'Treatment': 'lookthrough', 'HQLA_bn': r['USDT_HQLA_LT_bn'],
                     'Outflow_bn': r['USDT_Outflow_bn'], 'LCR_pct': r['USDT_LCR_LT'],
                     'Pass': r['USDT_LCR_LT'] >= LCR_PASS})
        rows.append({'Issuer': 'USDT', 'Quarter': quarter, 'Outflow_pct': o,
                     'Treatment': 'opaque', 'HQLA_bn': r['USDT_HQLA_OP_bn'],
                     'Outflow_bn': r['USDT_Outflow_bn'], 'LCR_pct': r['USDT_LCR_OP'],
                     'Pass': r['USDT_LCR_OP'] >= LCR_PASS})
        rows.append({'Issuer': 'USDP', 'Quarter': quarter, 'Outflow_pct': o,
                     'Treatment': 'single', 'HQLA_bn': r['USDP_HQLA_bn'],
                     'Outflow_bn': r['USDP_Outflow_bn'], 'LCR_pct': r['USDP_LCR'],
                     'Pass': (r['USDP_LCR'] >= LCR_PASS) if r['USDP_LCR'] == r['USDP_LCR'] else False})
    return pd.DataFrame(rows)


def make_lcr_stress_figure(ts):
    """
    Three stacked heatmaps (USDC / USDT / USDP), rows = outflow scenario,
    cols = quarter, cell = LCR% (look-through primary). A bold contour marks
    the LCR = 100% pass/fail boundary; cells failing are annotated.
    """
    import numpy as np
    quarters = list(dict.fromkeys(ts['Quarter']))
    outflows = sorted(ts['Outflow_pct'].unique())
    issuers = [('USDC', 'USDC_LCR_LT'), ('USDT', 'USDT_LCR_LT'), ('USDP', 'USDP_LCR')]

    fig, axes = plt.subplots(3, 1, figsize=(13, 9.5), constrained_layout=True)
    cmap = plt.cm.RdYlGn
    vmin, vmax = 0, 300

    for ax, (issuer, col) in zip(axes, issuers):
        M = np.full((len(outflows), len(quarters)), np.nan)
        for i, o in enumerate(outflows):
            sub = ts[ts['Outflow_pct'] == o].set_index('Quarter')
            for j, q in enumerate(quarters):
                if q in sub.index:
                    M[i, j] = sub.loc[q, col]
        im = ax.imshow(M, aspect='auto', cmap=cmap, vmin=vmin, vmax=vmax)
        ax.set_xticks(range(len(quarters)))
        ax.set_xticklabels(quarters, rotation=45, ha='right', fontsize=8)
        ax.set_yticks(range(len(outflows)))
        ax.set_yticklabels([f'{o}% run' for o in outflows])
        ax.set_title(f'{issuer} — LCR (%) under redemption stress (look-through HQLA)',
                     fontsize=10.5, loc='left')
        for i in range(len(outflows)):
            for j in range(len(quarters)):
                v = M[i, j]
                if np.isnan(v):
                    ax.text(j, i, 'n/a', ha='center', va='center', fontsize=7, color='#888')
                    continue
                fail = v < LCR_PASS
                ax.text(j, i, f'{v:.0f}', ha='center', va='center',
                        fontsize=7.5, weight='bold' if fail else 'normal',
                        color='#7a0000' if fail else '#111')
        # pass/fail contour at LCR=100
        ax.contour(np.where(np.isnan(M), vmax, M), levels=[LCR_PASS],
                   colors='black', linewidths=1.6)
    cbar = fig.colorbar(im, ax=axes, shrink=0.6, pad=0.01)
    cbar.set_label('LCR (%) — green = pass (>=100%), red = fail')
    fig.suptitle('RQ3 — Liquidity Coverage Ratio across 14 quarters x 3 redemption scenarios '
                 '(primary look-through HQLA)', fontsize=13)
    return fig


def make_lcr_vs_rate_combined(panel, usdp_panel, ts):
    """
    RQ2 x RQ3 intersection panel.

    For each quarter and the adverse stress pair (180d / +400bp), we ask the
    binding question: if a 40% redemption run forces sale of the sovereign book
    BEFORE maturity, how large is the realised MtM loss relative to (a) HQLA and
    (b) the equity buffer? This is the only state where RQ2's paper loss bites.

    Two panels:
      (top) post-run residual HQLA coverage of a 40% run (look-through), per
            issuer, with the 100% line.
      (bottom) the +400bp/180d realised MtM loss expressed as % of the equity
            buffer, per issuer — i.e. would the forced-sale loss erode capital.
    """
    import numpy as np
    notion = _sovereign_notional(panel, usdp_panel)
    quarters = list(notion['Quarter'])
    wam_years = LCR_RATE_WAM_DAYS / 365.0
    dy = LCR_RATE_SHOCK_BP / 10000.0

    # 40% run LCR per issuer (look-through)
    run = ts[ts['Outflow_pct'] == 40].set_index('Quarter')
    usdc_lcr = [run.loc[q, 'USDC_LCR_LT'] if q in run.index else np.nan for q in quarters]
    usdt_lcr = [run.loc[q, 'USDT_LCR_LT'] if q in run.index else np.nan for q in quarters]
    usdp_lcr = [run.loc[q, 'USDP_LCR']    if q in run.index else np.nan for q in quarters]

    # Realised MtM loss (forced sale) as % of RESERVE EQUITY (loss-absorbing
    # buffer at the level the token-holder's claim attaches to).
    # Reserve equity = reserve assets - circulation for ALL issuers (consistent
    # with the reserve-level capital methodology). For USDC this is structurally
    # tiny (~4-45 bps), which is itself the finding: near-zero reserve capital
    # means the buffer cannot absorb even a small realised loss — safety rests on
    # asset quality and short duration, not equity.
    def loss_pct_capital(sov_col, cap_series):
        out = []
        for (_, r), cap in zip(notion.iterrows(), cap_series):
            sov = float(r[sov_col]) if r[sov_col] == r[sov_col] else np.nan
            loss = wam_years * dy * sov
            out.append(loss / cap * 100 if (cap and cap == cap) else np.nan)
        return out
    usdc_cap = [float(r['USDC_Equity_Buffer'])
                if r['USDC_Equity_Buffer'] == r['USDC_Equity_Buffer'] else np.nan
                for _, r in notion.iterrows()]
    usdt_cap = [float(r['USDT_Equity_Buffer']) for _, r in notion.iterrows()]
    usdp_cap = [float(r['USDP_Equity_Buffer']) for _, r in notion.iterrows()]
    usdc_lb = loss_pct_capital('USDC_Sov_Notional', usdc_cap)
    usdt_lb = loss_pct_capital('USDT_Sov_Notional', usdt_cap)
    usdp_lb = loss_pct_capital('USDP_Sov_Notional', usdp_cap)

    x = range(len(quarters))
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(13, 8.5), constrained_layout=True)

    ax1.plot(x, usdc_lcr, marker='o', label='USDC', color='#1f77b4')
    ax1.plot(x, usdt_lcr, marker='s', label='USDT', color='#2ca02c')
    ax1.plot(x, usdp_lcr, marker='^', label='USDP', color='#d62728')
    ax1.axhline(LCR_PASS, color='black', ls='--', lw=1, label='LCR = 100% (pass)')
    ax1.set_ylabel('LCR (%), 40% run')
    ax1.set_title('A — Liquidity coverage of a 40% redemption run (look-through HQLA)',
                  fontsize=11, loc='left')
    ax1.set_xticks(list(x)); ax1.set_xticklabels(quarters, rotation=45, ha='right', fontsize=8)
    ax1.legend(fontsize=8, ncol=4)

    ax2.plot(x, usdc_lb, marker='o', label='USDC', color='#1f77b4')
    ax2.plot(x, usdt_lb, marker='s', label='USDT', color='#2ca02c')
    ax2.plot(x, usdp_lb, marker='^', label='USDP', color='#d62728')
    ax2.set_ylabel('Realised MtM loss\n(% of reserve equity)')
    ax2.set_title('B — If the run forces a pre-maturity sale: +400bp / 180d MtM loss as % of reserve equity '
                  '(reserve assets \u2212 circulation, all issuers) — the RQ2 x RQ3 bite',
                  fontsize=10, loc='left')
    ax2.set_xticks(list(x)); ax2.set_xticklabels(quarters, rotation=45, ha='right', fontsize=8)
    ax2.legend(fontsize=8, ncol=3)

    fig.suptitle('RQ2 x RQ3 — A redemption run is what turns paper duration loss into a realised one',
                 fontsize=13)
    return fig


def compute_effective_lcr(panel, usdp_panel, wam_days=LCR_RATE_WAM_DAYS,
                          shock_bp=LCR_RATE_SHOCK_BP):
    """
    RQ2 x RQ3 combined coverage: effective LCR when a redemption run FORCES a
    pre-maturity sale of the sovereign book during a rate shock, so the
    otherwise-paper MtM loss is realised against the HQLA actually sold.

    Effective LCR = [ HQLA x (1 - MtM_loss_fraction) ] / (outflow x Circulation)
        MtM_loss_fraction = (wam_days / 365) * (shock_bp / 10000)   [from RQ2]
        HQLA               = post-regulatory-haircut look-through stock [from RQ3]

    Also reports a 5-year-duration COUNTERFACTUAL on the same balance sheet, to
    show that the intersection is benign ONLY because the real book is short:
    at a 5y WAM the same shock would carve ~20% off the HQLA (the SVB mechanism).

    Returns
    -------
    pd.DataFrame, long format: one row per Quarter x Outflow_pct, with plain and
    effective LCR per issuer plus the 5y counterfactual effective LCR.
    """
    ts = compute_lcr_timeseries(panel, usdp_panel)
    mtm_frac = (wam_days / 365.0) * (shock_bp / 10000.0)
    mtm_frac_5y = (5 * 365 / 365.0) * (shock_bp / 10000.0)  # 5-year WAM counterfactual

    rows = []
    for _, r in ts.iterrows():
        rec = {'Quarter': r['Quarter'], 'Outflow_pct': r['Outflow_pct'],
               'MtM_loss_fraction': round(mtm_frac, 5),
               'MtM_loss_fraction_5y': round(mtm_frac_5y, 5)}
        for iss, hcol, lcol in [('USDC', 'USDC_HQLA_LT_bn', 'USDC_LCR_LT'),
                                ('USDT', 'USDT_HQLA_LT_bn', 'USDT_LCR_LT'),
                                ('USDP', 'USDP_HQLA_bn',   'USDP_LCR')]:
            plain = r[lcol]
            rec[f'{iss}_LCR_plain'] = plain
            rec[f'{iss}_LCR_effective'] = (round(plain * (1 - mtm_frac), 3)
                                           if plain == plain else float('nan'))
            rec[f'{iss}_LCR_effective_5y'] = (round(plain * (1 - mtm_frac_5y), 3)
                                              if plain == plain else float('nan'))
        rows.append(rec)
    return pd.DataFrame(rows)


def make_effective_lcr_figure(eff, outflow_pct=40):
    """
    Combined-coverage figure at a chosen run severity (default 40%).

    Panel A: plain vs effective LCR per issuer across quarters (the realised
             rate-shock penalty is the small gap between the two lines).
    Panel B: the 5-year-duration counterfactual effective LCR — same shock, long
             book — showing where the intersection WOULD break the 100% line
             (the SVB mechanism). Latest-quarter bars per issuer.
    """
    import numpy as np
    sub = eff[eff['Outflow_pct'] == outflow_pct]
    quarters = list(sub['Quarter'])
    x = range(len(quarters))

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5.5),
                                   gridspec_kw={'width_ratios': [2.2, 1]},
                                   constrained_layout=True)

    colors = {'USDC': '#1f77b4', 'USDT': '#2ca02c', 'USDP': '#d62728'}
    for iss in ['USDC', 'USDT', 'USDP']:
        ax1.plot(x, sub[f'{iss}_LCR_plain'], color=colors[iss], lw=1.6,
                 marker='o', ms=4, label=f'{iss} plain (sell at par)')
        ax1.plot(x, sub[f'{iss}_LCR_effective'], color=colors[iss], lw=1.6,
                 ls='--', marker='x', ms=5, alpha=0.85,
                 label=f'{iss} effective (forced sale, +{LCR_RATE_SHOCK_BP}bp)')
    ax1.axhline(LCR_PASS, color='black', ls=':', lw=1.2, label='LCR = 100% (pass)')
    ax1.set_xticks(list(x)); ax1.set_xticklabels(quarters, rotation=45, ha='right', fontsize=8)
    ax1.set_ylabel(f'LCR (%), {outflow_pct}% run')
    ax1.set_title(f'A — Plain vs effective LCR ({outflow_pct}% run; effective realises the '
                  f'{LCR_RATE_WAM_DAYS}d/+{LCR_RATE_SHOCK_BP}bp forced-sale loss)',
                  fontsize=10.5, loc='left')
    ax1.legend(fontsize=7, ncol=2)

    # Panel B: latest-quarter, short book vs 5y counterfactual
    latest = sub.iloc[-1]
    issuers = ['USDC', 'USDT', 'USDP']
    short_vals = [latest[f'{i}_LCR_effective'] for i in issuers]
    long_vals  = [latest[f'{i}_LCR_effective_5y'] for i in issuers]
    xb = np.arange(len(issuers)); w = 0.38
    ax2.bar(xb - w/2, short_vals, w, color=[colors[i] for i in issuers],
            label=f'actual short book (~{LCR_RATE_WAM_DAYS}d)')
    ax2.bar(xb + w/2, long_vals, w, color=[colors[i] for i in issuers], alpha=0.45,
            hatch='//', label='5y-duration counterfactual')
    ax2.axhline(LCR_PASS, color='black', ls=':', lw=1.2)
    ax2.set_xticks(xb); ax2.set_xticklabels(issuers)
    ax2.set_ylabel('Effective LCR (%)')
    ax2.set_title(f'B — {latest["Quarter"]}: why duration discipline matters\n'
                  f'(same +{LCR_RATE_SHOCK_BP}bp shock, short vs 5y book — the SVB mechanism)',
                  fontsize=10, loc='left')
    ax2.legend(fontsize=7.5)
    for xi, v in zip(xb - w/2, short_vals):
        if v == v: ax2.text(xi, v + 4, f'{v:.0f}', ha='center', fontsize=7.5)
    for xi, v in zip(xb + w/2, long_vals):
        if v == v: ax2.text(xi, v + 4, f'{v:.0f}', ha='center', fontsize=7.5,
                            weight='bold' if v < LCR_PASS else 'normal',
                            color='#7a0000' if v < LCR_PASS else '#111')

    fig.suptitle('RQ2 x RQ3 — Combined coverage: the run realises the rate loss, '
                 'but short duration keeps the bite small', fontsize=12.5)
    return fig

# ==============================================================================
# SECTION 5 (cont.) — Discussion scaffold: stablecoin T-bill footprint vs the
# market (the price-taker / fire-sale scaling limitation, RQ3 discussion)
# ------------------------------------------------------------------------------
# The static LCR assumes the issuer is a PRICE-TAKER: it liquidates its sovereign
# book at an exogenously shocked price, with no feedback from the sale itself.
# That holds at current scale because stablecoin T-bill holdings are a low single-
# digit % of the ~$3.6-6.5tn marketable bill market. This figure documents that
# share over time so the discussion can argue, with a number, that the price-taker
# assumption is defensible NOW but weakens as the sector grows toward systemic size
# (Brunnermeier-Pedersen 2009 liquidity spirals; Greenwood-Landier-Thesmar 2015
# common-asset fire-sale spillovers). Source for the denominator: US Treasury
# Monthly Statement of the Public Debt (MSPD), Marketable / Bills, quarter-end.
# ==============================================================================

def load_bills_outstanding(sheet_name='MSPD_SumSecty'):
    """Quarter-end marketable T-bills outstanding ($bn) from the Treasury MSPD
    sheet in the master workbook."""
    df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)
    bills = df[(df['Security Type Description'] == 'Marketable') &
               (df['Security Class Description'] == 'Bills')].copy()
    bills['Record Date'] = pd.to_datetime(bills['Record Date'])
    bills['Quarter'] = (bills['Record Date'].dt.year.astype(str) + '-Q' +
                        bills['Record Date'].dt.quarter.astype(str))
    qb = bills.sort_values('Record Date').groupby('Quarter').last().reset_index()
    qb['Bills_outstanding_bn'] = qb['Total Public Debt Outstanding (in Millions)'] / 1000.0
    lo, hi = q_to_num(SCOPE_FIRST_Q), q_to_num(SCOPE_LAST_Q)
    qb = qb[qb['Quarter'].apply(lambda q: lo <= q_to_num(q) <= hi)]
    return qb[['Quarter', 'Bills_outstanding_bn']].reset_index(drop=True)


def compute_tbill_share(panel, bills_df):
    """Sector (USDC+USDT) direct T-bill holdings as % of total bills outstanding."""
    p = panel.copy()
    p['USDC_Tbills_bn'] = p['USDC_Tbills_in_Fund'].fillna(0) / 1e9
    p['USDT_Tbills_bn'] = p['USDT_Tbills_Direct'].fillna(0) / 1e9
    m = p.merge(bills_df, on='Quarter')
    m['Sector_Tbills_bn'] = m['USDC_Tbills_bn'] + m['USDT_Tbills_bn']
    m['Share_pct'] = m['Sector_Tbills_bn'] / m['Bills_outstanding_bn'] * 100
    return m[['Quarter', 'USDC_Tbills_bn', 'USDT_Tbills_bn', 'Sector_Tbills_bn',
              'Bills_outstanding_bn', 'Share_pct']]


def make_tbill_footprint_figure(share_df):
    """
    Twin-axis: stacked sector T-bill holdings ($bn, left) and the share of total
    marketable bills outstanding (%, right). Documents the scaling trend behind
    the price-taker limitation.
    """
    q = list(share_df['Quarter'])
    x = range(len(q))
    fig, ax1 = plt.subplots(figsize=(12, 5.5), constrained_layout=True)
    ax1.bar(x, share_df['USDT_Tbills_bn'], color='#2ca02c', label='USDT direct T-bills')
    ax1.bar(x, share_df['USDC_Tbills_bn'], bottom=share_df['USDT_Tbills_bn'],
            color='#1f77b4', label='USDC in-fund T-bills')
    ax1.set_ylabel('Sector T-bill holdings ($bn)')
    ax1.set_xticks(list(x)); ax1.set_xticklabels(q, rotation=45, ha='right', fontsize=8)
    ax1.legend(loc='upper left', fontsize=8)

    ax2 = ax1.twinx()
    ax2.plot(x, share_df['Share_pct'], color='#d62728', lw=2, marker='o',
             label='Share of total bills outstanding (%)')
    ax2.set_ylabel('Share of marketable bills outstanding (%)', color='#d62728')
    ax2.tick_params(axis='y', labelcolor='#d62728')
    ax2.set_ylim(0, max(4, share_df['Share_pct'].max() * 1.6))
    ax2.legend(loc='upper right', fontsize=8)

    ax1.set_title('Stablecoin T-bill footprint vs the market — small but rising '
                  '(price-taker assumption holds now, weakens at scale)',
                  fontsize=12, loc='left')
    return fig

def print_lcr_summary(grid, ts):
    """Console summary: headline grid + first fail per issuer/treatment."""
    print("\n  Headline LCR grid (latest quarter, %):")
    piv = grid.pivot_table(index=['Issuer', 'Treatment'],
                           columns='Outflow_pct', values='LCR_pct')
    print(piv.round(1).to_string())

    print("\n  Pass/fail at the 100% threshold (latest quarter):")
    for _, r in grid.iterrows():
        flag = 'PASS' if r['Pass'] else 'FAIL'
        print(f"    {r['Issuer']:4s} {r['Treatment']:11s} {r['Outflow_pct']}% run: "
              f"LCR {r['LCR_pct']:.1f}%  [{flag}]")

    print("\n  Lowest LCR across the full 14-quarter panel (look-through, 60% run):")
    worst = ts[ts['Outflow_pct'] == 60]
    for issuer, col in [('USDC', 'USDC_LCR_LT'), ('USDT', 'USDT_LCR_LT'), ('USDP', 'USDP_LCR')]:
        sub = worst.dropna(subset=[col])
        if len(sub):
            mn = sub.loc[sub[col].idxmin()]
            tag = 'FAIL' if mn[col] < LCR_PASS else 'pass'
            print(f"    {issuer}: min {mn[col]:.1f}% at {mn['Quarter']}  [{tag}]")




# ==============================================================================
# SECTION 6: Week 3 — Narrow-bank vs MMF regulatory classification (RQ4)
# ------------------------------------------------------------------------------
# WHAT THIS SECTION PRODUCES
#   - classification_scorecard.csv : per-issuer sub-scores + blended spectrum
#                                    score (0 = pure narrow-bank, 100 = MMF-like)
#                                    with a sensitivity band.
#   - capital_bridge.csv           : Basel-analogous 7% capital test
#                                    (RWA x 7% required vs loss-absorbing equity).
#   - fig_classification_spectrum.png : 1-D positioning figure + axis breakdown.
#
# THE THROUGH-LINE
#   RQ1 (RWA density), RQ2 (rate/MtM), RQ3 (LCR) and the asset-mix / disclosure
#   flags are collapsed onto a single narrow-bank <-> MMF axis. The argument the
#   thesis has been building since Section 2 — that ASSET COMPOSITION and
#   DISCLOSURE, not size, drive the regulatory profile — is made quantitative
#   here. USDP anchors the narrow-bank end (whole book HQLA, ~0 rate risk in the
#   representative period); USDT sits at the MMF-like end (off-Basel BTC/gold/
#   secured-loans + coarser disclosure); USDC sits between, drifting toward the
#   narrow-bank end post-CRF.
#
# REGULATORY HOOK
#   Maps onto the GENIUS Act reserve-quality regime, narrow-bank / payment-
#   stablecoin proposals, and SEC 2a-7 money-market-fund reform: a narrow-bank
#   issuer would satisfy a 100%-HQLA reserve rule trivially, whereas an MMF-like
#   issuer would face fund-style liquidity-fee/gate and capital questions.
#
# METHODOLOGY (locked, documented — NOT data-snooped)
#   Five axes, each normalised to 0..100 against fixed regulatory anchors, then
#   blended with fixed weights. Full-sample (14-quarter) means are used for the
#   continuous axes so the USDP wind-down (book shrinks to bank cash, lifting its
#   RWA *ratio* mechanically) does not distort a single-quarter snapshot. The
#   wind-down effect is instead surfaced explicitly as the USDP sensitivity band.
# ==============================================================================

# --- Spectrum axis anchors: (narrow_bank_value, mmf_like_value) ---
# Chosen from regulatory intuition and fixed in advance. A value at the
# narrow-bank anchor scores 0; at the MMF anchor scores 100; linear in between,
# clipped to [0, 100].
CLASS_ANCHORS = {
    'rwa':  (0.00,  0.75),   # RWA density: 0% = pure sovereign; 75% ~ USDT ceiling
    'mtm':  (0.00,  0.03),   # adverse MtM loss (% assets): 0 = no rate risk; 3% = material
    'lcr':  (250.0, 50.0),   # LCR @40% run (INVERTED): 250% -> 0; 50% -> 100
    'risk': (0.00,  0.25),   # off-Basel asset share: 0 = none; 25% = clearly fund-like
    'disc': (0.00,  1.00),   # disclosure opacity: 0 = fully itemised HQLA; 1 = opaque
}

# --- Blend weights (sum to 1.0). Asset composition (rwa + risk) carries the most
#     weight because the thesis through-line is that composition is the driver. ---
CLASS_WEIGHTS = {'rwa': 0.25, 'mtm': 0.15, 'lcr': 0.20, 'risk': 0.25, 'disc': 0.15}

# --- Disclosure-opacity scores (line-item granularity of the HQLA breakdown) ---
#   USDP : Tier-3 Withum/KPMG examinations, fully itemised        -> 0.10
#   USDC : Tier-3 Circle attestations, itemised but CRF look-through
#          relies on fund holdings reports (one step removed)     -> 0.15
#   USDT : consolidated BDO CRR, coarser line items, residual
#          'other investments', no full look-through              -> 0.55
CLASS_DISCLOSURE = {'USDC': 0.15, 'USDT': 0.55, 'USDP': 0.10}

# --- Basel-analogous capital test ---
CAPITAL_REQ_PCT = 0.07   # 7% of RWA (Basel III Tier-1-style minimum, analogical)
# ---------------------------------------------------------------------------
# RESERVE-LEVEL capital methodology (corrected).
# A token-holder's legal claim attaches to the RESERVE, not the group. Group
# equity (Circle's SEC-filed stockholders' equity; Tether's consolidated equity
# incl. Tether Investments) is legally walled off from the reserve and is NOT
# counted as token-holder protection. Capital adequacy is therefore measured at
# reserve level only, where reserve equity = reserve assets - circulation.
# This is confirmed by Circle's own consolidated balance sheet, which segregates
# "cash ... for the benefit of stablecoin holders" (reserve assets) against
# "deposits from stablecoin holders" (the redemption claim); the gap is the
# reserve surplus (~$0.19bn on ~$44bn at FY2024 = ~4bps).
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Group-equity reference figures (CONTEXT ONLY — not used in the reserve-level
# capital test). Sourced from primary disclosures:
#   USDC: Circle Internet Group consolidated stockholders' equity (SEC S-1/10-Q).
#         FY2024 $0.57bn; FY2025 $3.33bn (post-IPO Jun-2025). Pre-IPO equity is
#         distorted by redeemable convertible preferred + accumulated deficit.
#   USDT: consolidated group equity incl. Tether Investments (BTC mining / AI /
#         energy). FY2023 $7.0bn -> FY2024 $20.0bn (driven by $13.7bn FY24 profit).
# These are shown alongside the reserve ratio to demonstrate why group capital
# is excluded, NOT to compute the adequacy test.
# ---------------------------------------------------------------------------
USDC_GROUP_EQUITY_REF_BN = {'2024-Q4': 0.571, '2025-Q4': 3.331}
USDT_GROUP_EQUITY_REF_BN = {'2023-Q4': 7.005, '2024-Q4': 20.030}


def _classification_axis_values(panel, usdp_panel, rwa, usdp_rwa, ts, lcr_ts):
    """
    Collapse RQ1-RQ3 + asset-mix + disclosure into raw per-issuer axis values,
    using full-sample (14-quarter) means for the continuous axes.

    Returns
    -------
    dict[issuer] -> dict of raw axis values, plus 'rwa_alt' for the band.
    """
    adv = ts['Scenario'] == TS_PAIRS[0][2]   # adverse ceiling (180d, +400bp)

    def mean_mtm(iss):
        s = ts[(ts['Issuer'] == iss) & adv]['MtM_Loss_pct_of_assets'].dropna()
        return s.mean() / 100.0               # back to fraction of assets

    def mean_lcr(col):
        return lcr_ts[lcr_ts['Outflow_pct'] == 40][col].dropna().mean()

    usdt_risk = ((panel['USDT_BTC'] + panel['USDT_Gold'] + panel['USDT_Sec_Loans'])
                 / panel['USDT_Total_Assets']).mean()

    avail = usdp_rwa[usdp_rwa['USDP_Data_Available']]

    return {
        'USDC': {
            'rwa':     rwa['USDC_RWA_Ratio_Primary'].mean(),
            'rwa_alt': rwa['USDC_RWA_Ratio_Sensitivity'].mean(),   # opaque-MMF band
            'mtm':     mean_mtm('USDC'),
            'lcr':     mean_lcr('USDC_LCR_LT'),
            'risk':    0.0,
            'disc':    CLASS_DISCLOSURE['USDC'],
        },
        'USDT': {
            'rwa':     rwa['USDT_RWA_Ratio'].mean(),
            'rwa_alt': rwa['USDT_RWA_Ratio'].mean(),                # no alt treatment
            'mtm':     mean_mtm('USDT'),
            'lcr':     mean_lcr('USDT_LCR_LT'),
            'risk':    usdt_risk,
            'disc':    CLASS_DISCLOSURE['USDT'],
        },
        'USDP': {
            'rwa':     avail['USDP_RWA_Ratio'].mean(),
            # Band low = representative pre-wind-down (2022-23); shows USDP's
            # structural narrow-bank density before the cash migration lifts it.
            'rwa_alt': avail[avail['Quarter'].isin(
                          ['2022-Q3','2022-Q4','2023-Q1','2023-Q2','2023-Q3','2023-Q4']
                       )]['USDP_RWA_Ratio'].mean(),
            'mtm':     mean_mtm('USDP'),
            'lcr':     mean_lcr('USDP_LCR'),
            'risk':    0.0,
            'disc':    CLASS_DISCLOSURE['USDP'],
        },
    }


def _norm_axis(val, key):
    """Normalise a raw axis value to 0..100 against its fixed anchor pair."""
    lo, hi = CLASS_ANCHORS[key]
    if hi == lo:
        return 0.0
    return float(min(max((val - lo) / (hi - lo) * 100.0, 0.0), 100.0))


def compute_classification_scorecard(panel, usdp_panel, rwa, usdp_rwa, ts, lcr_ts):
    """
    Build the narrow-bank <-> MMF classification scorecard.

    Returns
    -------
    pd.DataFrame, one row per issuer:
        sub_rwa, sub_mtm, sub_lcr, sub_risk, sub_disc  : 0..100 sub-scores
        Score      : blended spectrum score (0 = narrow-bank, 100 = MMF-like)
        Score_alt  : score under the alternative RWA treatment (band edge)
        Band_lo/hi : sensitivity band
        Class      : discrete label (Narrow-bank / Hybrid / MMF-like)
    """
    raw = _classification_axis_values(panel, usdp_panel, rwa, usdp_rwa, ts, lcr_ts)
    rows = []
    for iss in ['USDP', 'USDC', 'USDT']:        # ordered narrow -> MMF
        r = raw[iss]
        subs = {k: _norm_axis(r[k], k) for k in CLASS_ANCHORS}
        score = sum(CLASS_WEIGHTS[k] * subs[k] for k in CLASS_WEIGHTS)

        subs_alt = dict(subs)
        subs_alt['rwa'] = _norm_axis(r['rwa_alt'], 'rwa')
        score_alt = sum(CLASS_WEIGHTS[k] * subs_alt[k] for k in CLASS_WEIGHTS)

        band_lo, band_hi = min(score, score_alt), max(score, score_alt)
        label = ('Narrow-bank' if band_hi < 33 else
                 'MMF-like'    if band_lo >= 50 else 'Hybrid')

        rows.append({
            'Issuer': iss,
            'raw_rwa_pct':  round(r['rwa'] * 100, 3),
            'raw_mtm_pct':  round(r['mtm'] * 100, 4),
            'raw_lcr_pct':  round(r['lcr'], 1),
            'raw_risk_pct': round(r['risk'] * 100, 2),
            'raw_disc':     r['disc'],
            'sub_rwa':  round(subs['rwa'], 1),
            'sub_mtm':  round(subs['mtm'], 1),
            'sub_lcr':  round(subs['lcr'], 1),
            'sub_risk': round(subs['risk'], 1),
            'sub_disc': round(subs['disc'], 1),
            'Score':     round(score, 1),
            'Score_alt': round(score_alt, 1),
            'Band_lo':   round(band_lo, 1),
            'Band_hi':   round(band_hi, 1),
            'Class':     label,
        })
    return pd.DataFrame(rows)


def compute_capital_bridge(panel, usdp_panel, rwa, usdp_rwa, quarter=None):
    """
    RESERVE-LEVEL capital adequacy test (corrected methodology).

    For each issuer: reserve equity = reserve assets - circulation (the surplus
    that legally protects token-holders). Risk-weighted equity ratio =
    reserve equity / reserve RWA. Compared to a Basel-analogous 7% minimum.

    Key finding this surfaces: stablecoins are structurally NEAR-ZERO-CAPITAL.
    USDC's reserve surplus is ~4-45 bps of assets (the yield is swept to Circle
    as reserve income, not retained); USDT's is a thicker ~3-5% (it over-reserves
    and temporarily retains profit before dividending it up); USDP's is a thin
    narrow-bank cushion. The point is NOT a pass/fail verdict but that the
    Basel capital lens is near-vacuous for these entities — the binding
    constraints are asset composition (RQ1) and duration (RQ2), not equity.

    Returns a per-issuer DataFrame for the given quarter (default: latest).
    """
    quarter = quarter or SCOPE_LAST_Q
    p  = panel[panel['Quarter'] == quarter].iloc[0]
    r  = rwa[rwa['Quarter'] == quarter].iloc[0]
    rows = []

    def add(issuer, res_assets, circ, rwa_total, group_ref=None):
        eq  = res_assets - circ
        ratio = eq / rwa_total * 100 if rwa_total else float('nan')
        req = rwa_total * CAPITAL_REQ_PCT
        rows.append({
            'Issuer': issuer, 'Quarter': quarter,
            'Reserve_Equity_bn': round(eq / 1e9, 3),
            'Reserve_Equity_bps': round(eq / res_assets * 1e4, 1) if res_assets else None,
            'RWA_bn': round(rwa_total / 1e9, 3),
            'RW_Equity_Ratio_pct': round(ratio, 2),
            'Required_7pct_bn': round(req / 1e9, 4),
            'Coverage_x': round(eq / req, 2) if req else float('nan'),
            'Meets_7pct': eq >= req,
            'Group_Equity_ref_bn': group_ref,   # context only
        })

    add('USDC', p['USDC_Total_Assets'], p['USDC_Circulation'],
        r['USDC_RWA_Total_Primary'], USDC_GROUP_EQUITY_REF_BN.get(quarter))
    add('USDT', p['USDT_Total_Assets'], p['USDT_Circulation'],
        r['USDT_RWA_Total'], USDT_GROUP_EQUITY_REF_BN.get(quarter))
    up  = usdp_panel[usdp_panel['Quarter'] == quarter]
    ur  = usdp_rwa[usdp_rwa['Quarter'] == quarter]
    if len(up) and len(ur) and ur.iloc[0]['USDP_Data_Available']:
        add('USDP', up.iloc[0]['Total_Assets'], up.iloc[0]['USDP_Circulation'],
            ur.iloc[0]['USDP_RWA_Total'])
    return pd.DataFrame(rows)


def compute_capital_bridge_timeseries(panel, usdp_panel, rwa, usdp_rwa):
    """14-quarter reserve-level risk-weighted equity ratio for all issuers."""
    frames = [compute_capital_bridge(panel, usdp_panel, rwa, usdp_rwa, q)
              for q in panel['Quarter']]
    return pd.concat(frames, ignore_index=True)


def make_capital_ratio_figure(cap_ts):
    """
    Reserve-level risk-weighted equity ratio over time, three issuers, with the
    7% Basel-analogous line. Simple, single-message figure (per supervisor
    feedback to cut density).
    """
    import numpy as np
    fig, ax = plt.subplots(figsize=(9, 4.6))
    colors = {'USDC': '#1565c0', 'USDT': '#c62828', 'USDP': '#2e7d32'}
    for iss in ['USDC', 'USDT', 'USDP']:
        d = cap_ts[cap_ts['Issuer'] == iss]
        ax.plot(d['Quarter'], d['RW_Equity_Ratio_pct'], marker='o', label=iss,
                color=colors[iss], linewidth=1.8)
    ax.axhline(CAPITAL_REQ_PCT * 100, color='#555', ls='--', lw=1,
               label='7% Basel-analogous minimum')
    ax.set_ylabel('Reserve risk-weighted equity ratio (%)')
    ax.set_title('Reserve-level capital is thin for all issuers\n'
                 '(reserve equity \u00f7 reserve RWA)', fontsize=11, loc='left')
    ax.legend(frameon=False, ncol=2, fontsize=9)
    ax.grid(axis='y', alpha=0.25)
    plt.setp(ax.get_xticklabels(), rotation=45, ha='right', fontsize=8)
    fig.tight_layout()
    return fig


def load_tether_equity_profit():
    """
    Load the Tether equity & profit series from the master Excel
    'Tether_Equity_Profit' sheet (sourced from BDO ISAE-3000 attestations).

    Reserve = token-issuing entities (the buffer backing the token).
    Group   = consolidated incl. Tether Investments (BTC mining / AI / energy).
    Pre-2024 equity/profit is not disclosed (reserves-only reports) -> NaN.

    Returns a tidy DataFrame indexed by quarter, all values in USD millions.
    """
    df = pd.read_excel(EXCEL_PATH, sheet_name='Tether_Equity_Profit', skiprows=3)
    df = df[df['Quarter'].astype(str).str.match(r'\d{4}-Q\d')].copy()
    for c in ['Reserve_Equity_m', 'Reserve_Profit_Q_m', 'Reserve_Profit_YTD_m',
              'Group_Equity_m', 'Group_Profit_Q_m', 'Group_Profit_YTD_m']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    return df.reset_index(drop=True)


def compute_tether_capital_ratio(tep, rwa):
    """
    Tether risk-weighted equity ratio at BOTH levels, per quarter:
      reserve-level = reserve equity / reserve RWA  (the token-holder buffer)
      group-level   = group equity   / reserve RWA  (context; legally separate)
    Answers the supervisor's headline request directly, with the reserve series
    as primary. Returns a DataFrame in % terms where data exists.
    """
    out = []
    for _, r in tep.iterrows():
        q = r['Quarter']
        rr = rwa[rwa['Quarter'] == q]
        rwabn = rr['USDT_RWA_Total'].iloc[0] / 1e9 if len(rr) else float('nan')
        res_eq = r['Reserve_Equity_m'] / 1e3 if pd.notna(r['Reserve_Equity_m']) else float('nan')
        grp_eq = r['Group_Equity_m'] / 1e3 if pd.notna(r['Group_Equity_m']) else float('nan')
        out.append({
            'Quarter': q,
            'RWA_bn': round(rwabn, 1) if rwabn == rwabn else None,
            'Reserve_Equity_bn': round(res_eq, 2) if res_eq == res_eq else None,
            'Reserve_RW_Equity_pct': round(res_eq / rwabn * 100, 1) if (res_eq == res_eq and rwabn) else None,
            'Group_Equity_bn': round(grp_eq, 2) if grp_eq == grp_eq else None,
            'Group_RW_Equity_pct': round(grp_eq / rwabn * 100, 1) if (grp_eq == grp_eq and rwabn) else None,
            'Reserve_Profit_Q_bn': round(r['Reserve_Profit_Q_m'] / 1e3, 2) if pd.notna(r['Reserve_Profit_Q_m']) else None,
            'Group_Profit_Q_bn': round(r['Group_Profit_Q_m'] / 1e3, 2) if pd.notna(r['Group_Profit_Q_m']) else None,
        })
    return pd.DataFrame(out)


def print_tether_capital_ratio(tcr, tep):
    print("\n  Tether risk-weighted equity ratio (reserve = primary; group = context):")
    for _, r in tcr.iterrows():
        if r['Reserve_RW_Equity_pct'] is None or pd.isna(r['Reserve_RW_Equity_pct']):
            print(f"    {r['Quarter']}: equity/profit not disclosed (reserves-only report)")
            continue
        grp = (f"  | group {r['Group_RW_Equity_pct']:.1f}%"
               if (r['Group_RW_Equity_pct'] is not None and pd.notna(r['Group_RW_Equity_pct'])) else "")
        prof = (f"  | reserve profit ${r['Reserve_Profit_Q_bn']:.1f}bn"
                if (r['Reserve_Profit_Q_bn'] is not None and pd.notna(r['Reserve_Profit_Q_bn'])) else "")
        print(f"    {r['Quarter']}: reserve RW-equity {r['Reserve_RW_Equity_pct']:.1f}%{grp}{prof}")
    # the headline dividend story for FY2024
    fy = tep[tep['Quarter'].isin(['2024-Q1','2024-Q2','2024-Q3','2024-Q4'])]
    res_profit = fy['Reserve_Profit_Q_m'].sum(skipna=True) / 1e3
    print(f"\n    FY2024: reserve entities earned ~${res_profit:.1f}bn but reserve equity rose only "
          f"$5.2bn -> $7.1bn — profit was dividended up to the group, not retained in the reserve.")


def load_circle_equity_profit():
    """
    Load the Circle/USDC equity & profit series from the master Excel
    'Circle_Equity_Profit' sheet (sourced from SEC filings: S-1 audited
    2022-2024, 10-Q 2025).

    All figures are GROUP/consolidated (Circle Internet Group, CIK 0001876042).
    The USDC reserve runs ~1:1, so reserve-level equity is structurally
    near-zero every quarter and is NOT a meaningful retained buffer — the yield
    is swept to the parent as "reserve income". Net income is therefore a parent
    number, included as context (the mirror of Tether's group column).

    Returns a tidy DataFrame, USD millions.
    """
    df = pd.read_excel(EXCEL_PATH, sheet_name='Circle_Equity_Profit', skiprows=3)
    df = df[df['Quarter_tag'].astype(str).str.match(r'\d{4}-Q\d')].copy()
    for c in ['Group_Equity_m', 'Net_Income_m', 'Reserve_Income_m',
              'Total_Revenue_m', 'Reserve_Surplus_m']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    return df.reset_index(drop=True)


def print_circle_equity_profit(cep):
    """Console summary of Circle's group equity/profit and the reserve-passthrough point."""
    print("\n  Circle / USDC equity & profit (GROUP level, from SEC filings):")
    for _, r in cep.iterrows():
        eq = (f"equity ${r['Group_Equity_m']/1e3:.2f}bn"
              if pd.notna(r['Group_Equity_m']) else "equity n/d")
        ni = (f"net income ${r['Net_Income_m']:.0f}m"
              if pd.notna(r['Net_Income_m']) else "NI n/d")
        ri = (f"reserve income ${r['Reserve_Income_m']/1e3:.2f}bn"
              if pd.notna(r['Reserve_Income_m']) else "")
        print(f"    {r['Period']:8s}: {eq}  |  {ni}" + (f"  |  {ri}" if ri else ""))
    # the passthrough point
    fy24 = cep[cep['Period'] == 'FY2024']
    if len(fy24):
        r = fy24.iloc[0]
        print(f"\n    FY2024: Circle earned ${r['Reserve_Income_m']/1e3:.2f}bn of reserve income but kept only "
              f"${r['Net_Income_m']:.0f}m as net income — the reserve yield is swept to the parent, "
              f"not retained in the reserve (reserve surplus was just ${r['Reserve_Surplus_m']:.0f}m).")
    print("    => Like Tether, USDC's reserve retains ~0 equity; the mechanism differs "
          "(Circle passes yield through; Tether dividends it out) but the result is the same.")


def print_capital_bridge(bridge, cap_ts=None):
    print("\n  Reserve-level capital bridge (latest quarter):")
    for _, r in bridge.iterrows():
        grp = (f"  [group ref ${r['Group_Equity_ref_bn']:.2f}bn]"
               if pd.notna(r['Group_Equity_ref_bn']) else "")
        print(f"    {r['Issuer']}: reserve equity ${r['Reserve_Equity_bn']:.3f}bn "
              f"({r['Reserve_Equity_bps']:.0f}bps)  RWA ${r['RWA_bn']:.1f}bn  "
              f"RW-equity ratio {r['RW_Equity_Ratio_pct']:.2f}%  "
              f"({r['Coverage_x']:.2f}x of 7%){grp}")
    if cap_ts is not None:
        print("\n  Reserve RW-equity ratio range (14q):")
        for iss in ['USDC', 'USDT', 'USDP']:
            d = cap_ts[cap_ts['Issuer'] == iss]['RW_Equity_Ratio_pct'].dropna()
            if len(d):
                print(f"    {iss}: {d.min():.2f}% \u2013 {d.max():.2f}% "
                      f"(mean {d.mean():.2f}%)")


def make_classification_spectrum(scorecard):
    """
    Two-panel positioning figure:
      Panel A : the narrow-bank <-> MMF spectrum with each issuer placed by its
                blended score and a horizontal sensitivity band.
      Panel B : grouped bars of the five sub-scores per issuer (what drives the
                placement).
    """
    import numpy as np
    fig, (axA, axB) = plt.subplots(
        2, 1, figsize=(10, 7.2), gridspec_kw={'height_ratios': [1, 1.5]})

    colors = {'USDP': '#2e7d32', 'USDC': '#1565c0', 'USDT': '#c62828'}

    # ---- Panel A: spectrum ----
    axA.set_xlim(0, 100); axA.set_ylim(0, 1)
    axA.axhline(0.5, color='#bbb', lw=1, zorder=1)
    # zone shading
    axA.axvspan(0, 33, color='#2e7d32', alpha=0.06)
    axA.axvspan(33, 50, color='#9e9e9e', alpha=0.06)
    axA.axvspan(50, 100, color='#c62828', alpha=0.06)
    for x, lab in [(16, 'NARROW-BANK'), (41.5, 'HYBRID'), (75, 'MMF-LIKE')]:
        axA.text(x, 0.88, lab, ha='center', va='center', fontsize=9,
                 color='#555', weight='bold')
    for _, r in scorecard.iterrows():
        c = colors[r['Issuer']]
        axA.plot([r['Band_lo'], r['Band_hi']], [0.5, 0.5], color=c, lw=7,
                 solid_capstyle='round', alpha=0.35, zorder=2)
        axA.scatter([r['Score']], [0.5], s=180, color=c, zorder=3,
                    edgecolor='white', linewidth=1.5)
        axA.annotate(f"{r['Issuer']}\n{r['Score']:.0f}",
                     (r['Score'], 0.5), xytext=(0, 22), textcoords='offset points',
                     ha='center', va='bottom', fontsize=9, weight='bold', color=c)
    axA.set_yticks([])
    axA.set_xlabel('Narrow-bank  \u2190  classification score  \u2192  MMF-like')
    axA.set_title('RQ4 — Stablecoin regulatory classification spectrum '
                  '(0 = narrow-bank, 100 = MMF-like)', fontsize=10, weight='bold')
    for s in ['top', 'right', 'left']:
        axA.spines[s].set_visible(False)

    # ---- Panel B: sub-score breakdown ----
    axes_keys = ['sub_rwa', 'sub_mtm', 'sub_lcr', 'sub_risk', 'sub_disc']
    axes_lab  = ['RWA\ndensity', 'Rate/MtM\nrisk', 'Liquidity\n(inv. LCR)',
                 'Off-Basel\nassets', 'Disclosure\nopacity']
    x = np.arange(len(axes_keys)); w = 0.26
    for i, iss in enumerate(['USDP', 'USDC', 'USDT']):
        row = scorecard[scorecard['Issuer'] == iss].iloc[0]
        vals = [row[k] for k in axes_keys]
        axB.bar(x + (i - 1) * w, vals, w, label=iss, color=colors[iss], alpha=0.9)
    axB.set_xticks(x); axB.set_xticklabels(axes_lab, fontsize=8)
    axB.set_ylim(0, 100); axB.set_ylabel('sub-score (0\u2013100)')
    axB.set_title('What drives the placement — five normalised axes',
                  fontsize=10, weight='bold')
    axB.legend(frameon=False, ncol=3, loc='upper center')
    axB.grid(axis='y', alpha=0.25)

    fig.tight_layout()
    return fig


def print_classification_summary(scorecard):
    """Console summary of the scorecard."""
    print("\n  Classification scorecard (0 = narrow-bank, 100 = MMF-like):")
    cols = ['Issuer', 'sub_rwa', 'sub_mtm', 'sub_lcr', 'sub_risk', 'sub_disc',
            'Score', 'Band_lo', 'Band_hi', 'Class']
    print(scorecard[cols].to_string(index=False))




# ------------------------------------------------------------------------------
# SECTION 6 (cont.): Weight robustness & sensitivity analysis (RQ4)
# ------------------------------------------------------------------------------
# WHY THIS EXISTS
#   The blended score depends on a weight vector that is, ultimately, a judgement
#   call. The OECD/JRC Handbook on Constructing Composite Indicators (Nardo et
#   al. 2005/2008) makes robustness analysis on the weighting and aggregation
#   choices a REQUIRED step, not an optional one: a composite indicator must be
#   shown to be decomposable and its rankings shown to survive plausible changes
#   in the weights. This block supplies that evidence. It re-scores the three
#   issuers under (a) equal weights, (b) composition-only weights, (c) a
#   leave-one-axis-out family, and (d) a Monte-Carlo sweep over the entire weight
#   simplex (Dirichlet-uniform), reporting how often the ordinal classification
#   survives.
#
# REGULATORY PRECEDENT FOR THE WEIGHTING APPROACH
#   The supervisory analogue is the CAMELS / UFIRS composite rating: a small set
#   of theory-chosen components combined into one supervisory score, where the
#   relative weight on each component is set by judgement and "some components may
#   be given more weight than others depending on the situation" rather than by a
#   fixed public formula. Our theory-based weighting (composition > liquidity >
#   rate ~ disclosure) is the same species of construct; the robustness sweep is
#   what disciplines the discretion.
# ------------------------------------------------------------------------------

# Alternative weighting schemes for the robustness table.
CLASS_WEIGHT_SCHEMES = {
    'Base (theory)':      {'rwa': 0.25, 'mtm': 0.15, 'lcr': 0.20, 'risk': 0.25, 'disc': 0.15},
    'Equal':              {'rwa': 0.20, 'mtm': 0.20, 'lcr': 0.20, 'risk': 0.20, 'disc': 0.20},
    'Composition-only':   {'rwa': 0.50, 'mtm': 0.00, 'lcr': 0.00, 'risk': 0.50, 'disc': 0.00},
}

# Classification bucket thresholds (0..100 spectrum).
CLASS_NB_CEIL  = 33.0   # below -> narrow-bank zone
CLASS_MMF_FLOOR = 50.0  # at/above -> MMF-like zone


def _subscores_from_scorecard(scorecard):
    """Pull the five 0..100 sub-scores per issuer out of a built scorecard."""
    keys = ['rwa', 'mtm', 'lcr', 'risk', 'disc']
    return {r['Issuer']: {k: r['sub_' + k] for k in keys}
            for _, r in scorecard.iterrows()}


def _blend(subs_iss, weights):
    return sum(weights[k] * subs_iss[k] for k in weights)


def compute_weight_sensitivity(scorecard, n_mc=20000, seed=0):
    """
    Robustness of the classification to the weight vector.

    Returns
    -------
    schemes_df : pd.DataFrame
        Issuer scores under each named weighting scheme + leave-one-out family.
    mc_summary : dict
        Monte-Carlo (Dirichlet-uniform over the 5-simplex) survival rates of the
        key ordinal/threshold claims.
    """
    import numpy as np
    subs = _subscores_from_scorecard(scorecard)
    issuers = ['USDP', 'USDC', 'USDT']
    keys = ['rwa', 'mtm', 'lcr', 'risk', 'disc']

    # --- named schemes ---
    rows = []
    for name, W in CLASS_WEIGHT_SCHEMES.items():
        rows.append({'Scheme': name,
                     **{iss: round(_blend(subs[iss], W), 1) for iss in issuers}})
    # --- leave-one-out family (drop an axis, renormalise the rest) ---
    base = CLASS_WEIGHTS
    for drop in keys:
        W = {k: (base[k] if k != drop else 0.0) for k in keys}
        s = sum(W.values()); W = {k: v / s for k, v in W.items()}
        rows.append({'Scheme': f'Drop {drop}',
                     **{iss: round(_blend(subs[iss], W), 1) for iss in issuers}})
    schemes_df = pd.DataFrame(rows)

    # --- Monte-Carlo over the whole simplex ---
    rng = np.random.default_rng(seed)
    usdt_top = usdc_nb = usdp_nb = usdt_mmf = 0
    order_intact = 0  # USDP<USDT and USDC<USDT and (USDC,USDP) both below MMF floor
    for _ in range(n_mc):
        w = rng.dirichlet(np.ones(len(keys)))
        W = dict(zip(keys, w))
        sc = {iss: _blend(subs[iss], W) for iss in issuers}
        if sc['USDT'] == max(sc.values()):                       usdt_top += 1
        if sc['USDC'] < CLASS_NB_CEIL:                           usdc_nb += 1
        if sc['USDP'] < CLASS_NB_CEIL:                           usdp_nb += 1
        if sc['USDT'] >= CLASS_MMF_FLOOR:                        usdt_mmf += 1
        if (sc['USDT'] > sc['USDC'] and sc['USDT'] > sc['USDP']
                and sc['USDC'] < CLASS_MMF_FLOOR and sc['USDP'] < CLASS_MMF_FLOOR):
            order_intact += 1
    mc_summary = {
        'n': n_mc,
        'usdt_highest_pct':    round(usdt_top / n_mc * 100, 1),
        'usdc_narrowbank_pct': round(usdc_nb / n_mc * 100, 1),
        'usdp_narrowbank_pct': round(usdp_nb / n_mc * 100, 1),
        'usdt_mmf_pct':        round(usdt_mmf / n_mc * 100, 1),
        'ordinal_intact_pct':  round(order_intact / n_mc * 100, 1),
    }
    return schemes_df, mc_summary


def make_weight_sensitivity_figure(scorecard, schemes_df, mc_summary):
    """
    Two-panel robustness figure:
      Panel A : issuer score under each weighting scheme (base, equal,
                composition-only, leave-one-out) — shows ordering is stable.
      Panel B : Monte-Carlo survival rates of the key classification claims.
    """
    import numpy as np
    fig, (axA, axB) = plt.subplots(1, 2, figsize=(11, 4.6),
                                   gridspec_kw={'width_ratios': [1.6, 1]})
    colors = {'USDP': '#2e7d32', 'USDC': '#1565c0', 'USDT': '#c62828'}

    # Panel A
    schemes = schemes_df['Scheme'].tolist()
    y = np.arange(len(schemes))[::-1]
    for iss in ['USDP', 'USDC', 'USDT']:
        axA.scatter(schemes_df[iss], y, s=70, color=colors[iss], label=iss,
                    zorder=3, edgecolor='white', linewidth=1)
    axA.axvspan(0, CLASS_NB_CEIL, color='#2e7d32', alpha=0.06)
    axA.axvspan(CLASS_MMF_FLOOR, 100, color='#c62828', alpha=0.06)
    axA.axvline(CLASS_NB_CEIL, color='#bbb', lw=0.8, ls='--')
    axA.axvline(CLASS_MMF_FLOOR, color='#bbb', lw=0.8, ls='--')
    axA.set_yticks(y); axA.set_yticklabels(schemes, fontsize=8)
    axA.set_xlim(0, 100); axA.set_xlabel('classification score')
    axA.set_title('Issuer score under alternative weightings', fontsize=10, weight='bold')
    axA.legend(frameon=False, ncol=3, loc='lower right', fontsize=8)
    axA.grid(axis='x', alpha=0.2)

    # Panel B
    labels = ['USDT\nhighest', 'USDC\nnarrow-bank', 'USDP\nnarrow-bank',
              'USDT\nMMF-like', 'Full ordinal\nclaim']
    vals = [mc_summary['usdt_highest_pct'], mc_summary['usdc_narrowbank_pct'],
            mc_summary['usdp_narrowbank_pct'], mc_summary['usdt_mmf_pct'],
            mc_summary['ordinal_intact_pct']]
    barcols = ['#c62828', '#1565c0', '#2e7d32', '#c62828', '#1f3864']
    bars = axB.bar(range(len(vals)), vals, color=barcols, alpha=0.9)
    axB.axhline(90, color='#888', lw=0.8, ls='--')
    axB.text(len(vals) - 0.4, 91, '90%', fontsize=7, color='#888', va='bottom', ha='right')
    for b, v in zip(bars, vals):
        axB.text(b.get_x() + b.get_width() / 2, v + 1.5, f'{v:.0f}',
                 ha='center', fontsize=8, weight='bold')
    axB.set_xticks(range(len(labels))); axB.set_xticklabels(labels, fontsize=7.5)
    axB.set_ylim(0, 105); axB.set_ylabel('% of random weightings')
    axB.set_title(f'Claim survival over {mc_summary["n"]:,} random weight vectors',
                  fontsize=9.5, weight='bold')
    axB.grid(axis='y', alpha=0.2)

    fig.tight_layout()
    return fig


def print_weight_sensitivity(schemes_df, mc_summary):
    print("\n  Weight-sensitivity — scores under alternative schemes:")
    print(schemes_df.to_string(index=False))
    print(f"\n  Monte-Carlo robustness ({mc_summary['n']:,} Dirichlet-uniform weightings):")
    print(f"    USDT scores highest:         {mc_summary['usdt_highest_pct']}%")
    print(f"    USDC stays narrow-bank(<33): {mc_summary['usdc_narrowbank_pct']}%")
    print(f"    USDP stays narrow-bank(<33): {mc_summary['usdp_narrowbank_pct']}%")
    print(f"    USDT stays MMF-like(>=50):   {mc_summary['usdt_mmf_pct']}%")
    print(f"    Full ordinal claim intact:   {mc_summary['ordinal_intact_pct']}%")




# ==============================================================================
# SECTION 7: Data-driven structural-event detection (validation)
# ------------------------------------------------------------------------------
# WHY THIS EXISTS
#   The narrative "structural events" used throughout (SVB, FTX, the CRF launch,
#   the 2022-23 hiking cycle) were originally hand-picked from press coverage.
#   That invites the fair objection: "where do these come from?" This section
#   removes the hand-picking. It detects events PURELY FROM THE TIME SERIES —
#   abnormal moves in (i) peg price, (ii) token supply, and (iii) short rates —
#   and then shows the auto-detected set coincides with the known crises. The
#   known events thus VALIDATE the detector rather than being assumed; the
#   detector in turn shows the event set is not cherry-picked.
#
# METHOD (transparent, threshold-based — no look-ahead)
#   - Depeg event   : |price - 1| exceeds DEPEG_BP on a given day.
#   - Supply shock  : N-day % change in circulating supply below SUPPLY_DROP_PCT
#                     (redemption run) — sign-agnostic version also flags surges.
#   - Rate shock    : rolling 1y change in the 3M yield exceeds RATE_EVENT_BP.
#   Events are clustered to the worst observation per coin per calendar month so
#   a single multi-day episode is reported once.
# ==============================================================================

DEPEG_BP        = 50      # |price-1| > 0.50% = a depeg day
SUPPLY_WIN_D    = 7       # window for supply-contraction detection (days)
SUPPLY_DROP_PCT = -0.08   # > 8% supply drop over the window = redemption shock
RATE_EVENT_BP   = 300     # rolling 1y rise in 3M yield over 300bp = rate event


def _load_daily_supply():
    import openpyxl
    """Daily USDC/USDT price + supply over the thesis scope from Supply_Daily."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Supply_Daily']
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df['Date'] = pd.to_datetime(df['Date'])
    lo = pd.Timestamp(f"{SCOPE_FIRST_Q[:4]}-07-01")   # scope start (Q3)
    df = df[(df['Date'] >= pd.Timestamp('2022-07-01')) &
            (df['Date'] <= pd.Timestamp('2025-12-31'))].copy()
    for c in ['USDC_Price', 'USDC_Supply', 'USDT_Price', 'USDT_Supply']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    return df.sort_values('Date').reset_index(drop=True)


def _load_daily_yields():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Yield_Curve']
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df['Date'] = pd.to_datetime(df['Date'])
    df['DGS3MO (%)'] = pd.to_numeric(df['DGS3MO (%)'], errors='coerce')
    return df[df['Date'] >= pd.Timestamp('2022-07-01')].sort_values('Date')


def detect_structural_events():
    """
    Detect events from the data. Returns a tidy DataFrame:
        Date, Type, Coin, Magnitude, Detail
    sorted chronologically, one row per (type, coin, month) episode.
    """
    df = _load_daily_supply()
    out = []

    # --- depegs ---
    for coin in ['USDC', 'USDT']:
        dev = (df[f'{coin}_Price'] - 1.0).abs()
        hit = df.loc[dev > DEPEG_BP / 1e4, ['Date']].copy()
        hit['mag'] = dev[dev > DEPEG_BP / 1e4].values
        if len(hit):
            hit['ym'] = hit['Date'].dt.to_period('M')
            worst = hit.loc[hit.groupby('ym')['mag'].idxmax()]
            for _, r in worst.iterrows():
                out.append({'Date': r['Date'], 'Type': 'Depeg', 'Coin': coin,
                            'Magnitude': round(r['mag'] * 100, 2),
                            'Detail': f"price off peg by {r['mag']*100:.1f}%"})

    # --- supply contractions (redemption runs) ---
    for coin in ['USDC', 'USDT']:
        pct = df[f'{coin}_Supply'].pct_change(SUPPLY_WIN_D)
        hit = df.loc[pct < SUPPLY_DROP_PCT, ['Date']].copy()
        hit['mag'] = pct[pct < SUPPLY_DROP_PCT].values
        if len(hit):
            hit['ym'] = hit['Date'].dt.to_period('M')
            worst = hit.loc[hit.groupby('ym')['mag'].idxmin()]
            for _, r in worst.iterrows():
                out.append({'Date': r['Date'], 'Type': 'Supply shock', 'Coin': coin,
                            'Magnitude': round(r['mag'] * 100, 1),
                            'Detail': f"supply {r['mag']*100:.0f}% over {SUPPLY_WIN_D}d"})

    # --- rate shock ---
    yd = _load_daily_yields()
    yd = yd.copy(); yd['chg'] = yd['DGS3MO (%)'].diff(252) * 100
    rate_hit = yd[yd['chg'] > RATE_EVENT_BP]
    if len(rate_hit):
        peak = rate_hit.loc[rate_hit['chg'].idxmax()]
        out.append({'Date': peak['Date'], 'Type': 'Rate shock', 'Coin': 'Market',
                    'Magnitude': round(peak['chg'], 0),
                    'Detail': f"3M yield +{peak['chg']:.0f}bp over 1y"})

    ev = pd.DataFrame(out).sort_values('Date').reset_index(drop=True)
    return ev


# Known crises for the validation cross-check (labels only — NOT used to detect).
KNOWN_EVENTS = {
    '2022-11': 'FTX collapse',
    '2023-02': 'Paxos/BUSD wind-down order',
    '2023-03': 'SVB / Signature / Silvergate failures; USDC depeg',
}


def make_event_timeline_figure(events):
    """Timeline of detected events with the known-crisis labels overlaid."""
    import matplotlib.dates as mdates
    fig, ax = plt.subplots(figsize=(11, 3.4))
    type_y = {'Depeg': 3, 'Supply shock': 2, 'Rate shock': 1}
    type_c = {'Depeg': '#c62828', 'Supply shock': '#1565c0', 'Rate shock': '#6a1b9a'}
    coin_m = {'USDC': 'o', 'USDT': 's', 'Market': 'D'}
    for _, r in events.iterrows():
        ax.scatter(r['Date'], type_y[r['Type']], s=90,
                   color=type_c[r['Type']], marker=coin_m.get(r['Coin'], 'o'),
                   edgecolor='white', linewidth=1, zorder=3)
    # known-crisis bands
    for ym, lab in KNOWN_EVENTS.items():
        d = pd.Timestamp(ym + '-15')
        ax.axvline(d, color='#999', lw=0.8, ls='--', zorder=1)
        ax.text(d, 3.6, lab, rotation=0, fontsize=7.5, ha='center', color='#555')
    ax.set_yticks([1, 2, 3]); ax.set_yticklabels(['Rate shock', 'Supply shock', 'Depeg'])
    ax.set_ylim(0.4, 4.0); ax.set_xlabel('')
    ax.set_title('Data-driven structural events vs. known crises (markers = detected; '
                 'dashed = known)', fontsize=10, weight='bold')
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    plt.setp(ax.get_xticklabels(), rotation=45, ha='right', fontsize=7.5)
    # legend for coin markers
    from matplotlib.lines import Line2D
    leg = [Line2D([0],[0],marker='o',color='w',markerfacecolor='#444',label='USDC',markersize=8),
           Line2D([0],[0],marker='s',color='w',markerfacecolor='#444',label='USDT',markersize=8),
           Line2D([0],[0],marker='D',color='w',markerfacecolor='#444',label='Market',markersize=8)]
    ax.legend(handles=leg, frameon=False, ncol=3, loc='lower right', fontsize=8)
    ax.grid(axis='x', alpha=0.2)
    fig.tight_layout()
    return fig


def print_event_summary(events):
    print("\n  Data-driven structural events (detected from price/supply/rates):")
    for _, r in events.iterrows():
        print(f"    {r['Date'].date()}  {r['Type']:13s} {r['Coin']:6s} {r['Detail']}")
    # validation line
    months = set(events['Date'].dt.to_period('M').astype(str))
    hit = [lab for ym, lab in KNOWN_EVENTS.items() if ym in months]
    print(f"\n  Validation: detected events coincide with {len(hit)}/{len(KNOWN_EVENTS)} "
          f"known crises: {', '.join(hit)}")


# ==============================================================================
# MAIN ENTRY POINT
# ==============================================================================

# ==============================================================================
# SECTION 8: Supervisor additions — bank benchmark, BTC/gold price shocks,
#            retained-earnings attribution, joint (triple-negative) stress
# ------------------------------------------------------------------------------
# METHODOLOGY
#   BANK BENCHMARK (RQ1 extension). Two US Basel III banks bracket the banking
#   spectrum on the same 14-quarter window: BNY Mellon (custody model, the
#   closest real-bank analogue to a narrow bank; historically a USDC reserve
#   custodian) and JPMorgan (universal G-SIB extreme). Standardized CET1
#   headline ratios per quarter, hardcoded from each quarter's earnings release
#   (SEC 8-K / IR PDF); later capital tables show final revisions of +-0.1pp.
#
#   PRICE SHOCKS (RQ2/RQ3 extension). Instantaneous L = notional x |dp| on
#   USDT's disclosed BTC and gold lines, single and combined, each quarter,
#   against reserve equity (= total assets - circulation). Pre-2023 zeros are
#   a DISCLOSURE artifact (separate CRR categories only from Q1 2023).
#   USDC/USDP hold zero of both -> structural immunity is itself a result.
#
#   RETAINED EARNINGS (RQ1 extension; Acharya, Gujral, Kulkarni & Shin 2022,
#   J. Financial Crises 4(2); orig. NBER WP 16896). Implied distribution_t =
#   E_{t-1} + Profit_t - E_t; retention = dE / Profit. Reserve level only.
#
#   JOINT STRESS (scenario formation). Monthly BTC return, gold return and
#   d(3M yield). Gold: long spliced series (LBMA/datahub pre-2019, Macrotrends
#   2019+, Gold_Monthly_Long sheet); gold price fixed pre-1971 (Bretton
#   Woods), so correlation analysis uses the post-Aug-1971 free-float era.
#   "Triple-negative" = BTC<0, gold<0, dy>0 (bill prices down). The worst
#   observed episodes are replayed on the latest USDT book (rate leg at 90d
#   WAM on the sovereign notional).
# ==============================================================================

BANK_CET1 = {  # Quarter -> (JPM_CET1_pct, JPM_RWA_bn, BNY_CET1_pct)
    '2022-Q3': (12.5, 1682, 10.0), '2022-Q4': (13.2, 1700, 11.2),
    '2023-Q1': (13.8, 1700, 11.0), '2023-Q2': (13.8, 1700, 11.1),
    '2023-Q3': (14.3, 1700, 11.4), '2023-Q4': (15.0, 1700, 11.6),
    '2024-Q1': (15.0, 1700, 10.8), '2024-Q2': (15.3, 1743, 11.4),
    '2024-Q3': (15.3, 1783, 11.9), '2024-Q4': (15.7, 1800, 11.2),
    '2025-Q1': (15.4, 1815, 11.5), '2025-Q2': (15.0, 1886, 11.5),
    '2025-Q3': (14.8, 1935, 11.7), '2025-Q4': (14.5, 2000, 11.9),
}  # Sources: JPM & BNY quarterly earnings releases 3Q22-4Q25 (SEC 8-K / IR)

PRICE_SHOCKS = [0.20, 0.30, 0.40, 0.50]

RETAINED_EARNINGS = [
    # (issuer, year, E_open_m, profit_m, E_close_m) — reserve level, $m
    # Tether: BDO reports + tether.io attestation announcements (v6 fill);
    # Circle: S-1 audited FY2024 / 10-Qs (reserve income swept to parent).
    ('USDT', 'FY2023',  960.0,  6200.0, 5203.0),
    ('USDT', 'FY2024', 5203.0, 12139.0, 7087.0),
    ('USDT', 'FY2025', 7087.0, 10106.0, 6338.0),
    ('USDC', 'FY2024',  191.2,  1661.1,  191.2),
    ('USDC', 'FY2025',  191.2, np.nan,   155.3),
]


def compute_bank_benchmark(cap_ts):
    """Bank CET1 vs issuer reserve-level RW-equity ratio, quarterly panel."""
    rows = []
    for q, (jc, jr, bc) in sorted(BANK_CET1.items()):
        rec = {'Quarter': q, 'JPM_CET1_pct': jc, 'JPM_RWA_bn': jr,
               'BNY_CET1_pct': bc}
        sub = cap_ts[cap_ts['Quarter'] == q]
        for iss in ['USDC', 'USDT', 'USDP']:
            v = sub[sub['Issuer'] == iss]['RW_Equity_Ratio_pct']
            rec[f'{iss}_RWeq_pct'] = round(float(v.iloc[0]), 2) if len(v) else np.nan
        rows.append(rec)
    return pd.DataFrame(rows)


def make_bank_benchmark_figure(bb):
    fig, ax = plt.subplots(figsize=(11, 5.5))
    x = range(len(bb))
    ax.plot(x, bb['JPM_CET1_pct'], marker='o', lw=2, label='JPMorgan CET1 (Std)')
    ax.plot(x, bb['BNY_CET1_pct'], marker='s', lw=2, label='BNY Mellon CET1')
    for iss, mk in [('USDC', '^'), ('USDT', 'v'), ('USDP', 'd')]:
        ax.plot(x, bb[f'{iss}_RWeq_pct'], marker=mk, lw=1.4, ls='--',
                label=f'{iss} reserve RW-equity')
    ax.axhline(7, color='crimson', lw=1, ls=':', label='7% analogical minimum')
    ax.set_xticks(list(x))
    ax.set_xticklabels(bb['Quarter'], rotation=45, ha='right')
    ax.set_ylabel('% of RWA')
    ax.set_title('Basel III bank CET1 vs stablecoin reserve-level RW-equity, '
                 'Q3 2022 - Q4 2025')
    ax.legend(fontsize=8, ncol=2)
    fig.tight_layout()
    return fig


def compute_price_shock_panel(panel):
    """USDT BTC/gold shock grid across the analytical window."""
    rows = []
    for _, r in panel.iterrows():
        q = r['Quarter']
        btc  = float(r.get('USDT_BTC', 0) or 0) / 1e9
        gold = float(r.get('USDT_Gold', 0) or 0) / 1e9
        eq   = (float(r['USDT_Total_Assets']) - float(r['USDT_Circulation'])) / 1e9
        for s in PRICE_SHOCKS:
            lb, lg = btc * s, gold * s
            rows.append({
                'Quarter': q, 'Shock_pct': int(s * 100),
                'BTC_Notional_bn': round(btc, 3),
                'Gold_Notional_bn': round(gold, 3),
                'Reserve_Equity_bn': round(eq, 3),
                'Loss_BTC_only_bn': round(lb, 3),
                'Buffer_after_BTC_bn': round(eq - lb, 3),
                'Loss_Gold_only_bn': round(lg, 3),
                'Buffer_after_Gold_bn': round(eq - lg, 3),
                'Loss_Combined_bn': round(lb + lg, 3),
                'Buffer_after_Combined_bn': round(eq - lb - lg, 3),
                'Breakeven_Combined_pct': (round(eq / (btc + gold) * 100, 1)
                                           if btc + gold > 0 else np.nan)})
    return pd.DataFrame(rows)


def make_price_shock_figure(sp):
    """Left: break-even trajectory (falling = deteriorating resilience).
       Right: latest-quarter buffer after each shock size."""
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    be = sp.drop_duplicates('Quarter').dropna(subset=['Breakeven_Combined_pct'])
    axes[0].plot(be['Quarter'], be['Breakeven_Combined_pct'],
                 marker='o', color='darkred')
    axes[0].set_title('USDT combined BTC+gold break-even shock')
    axes[0].set_ylabel('buffer-exhausting |shock| (%)')
    axes[0].tick_params(axis='x', rotation=45)
    axes[0].invert_yaxis()
    last = sp[sp['Quarter'] == sp['Quarter'].max()]
    w = 0.25
    x = np.arange(len(last))
    axes[1].bar(x - w, last['Buffer_after_BTC_bn'], w, label='after BTC-only')
    axes[1].bar(x,     last['Buffer_after_Gold_bn'], w, label='after gold-only')
    axes[1].bar(x + w, last['Buffer_after_Combined_bn'], w, label='after combined')
    axes[1].axhline(0, color='k', lw=0.8)
    axes[1].set_xticks(x)
    axes[1].set_xticklabels([f"-{s}%" for s in last['Shock_pct']])
    axes[1].set_title(f"Buffer after shock, {sp['Quarter'].max()} ($bn)")
    axes[1].legend(fontsize=8)
    fig.tight_layout()
    return fig


def compute_retention_attribution():
    rows = []
    for iss, yr, e0, p, e1 in RETAINED_EARNINGS:
        ok = not any(np.isnan([e0, p, e1]))
        rows.append({'Issuer': iss, 'Year': yr, 'Equity_Open_m': e0,
                     'Profit_m': p, 'Equity_Close_m': e1,
                     'Implied_Distribution_m': round(e0 + p - e1, 0) if ok else np.nan,
                     'Retention_pct': round((e1 - e0) / p * 100, 1) if ok and p else np.nan})
    return pd.DataFrame(rows)


def load_gold_long():
    """Long gold series: Gold_Monthly_Long (v7) if present, else Gold_Monthly."""
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = 'Gold_Monthly_Long' if 'Gold_Monthly_Long' in wb.sheetnames else 'Gold_Monthly'
    rows = list(wb[sheet].iter_rows(values_only=True))
    if sheet == 'Gold_Monthly_Long':
        hdr_i = next(i for i, r in enumerate(rows) if r[0] == 'Date')
        df = pd.DataFrame(rows[hdr_i + 1:], columns=[str(c) for c in rows[hdr_i]])
        s = pd.Series(df['Gold_USD_oz'].astype(float).values,
                      index=pd.to_datetime(df['Date']) + pd.offsets.MonthEnd(0))
    else:
        df = pd.DataFrame(rows[1:], columns=[str(c) for c in rows[0]])
        s = pd.Series(df['Gold_Price_USD_per_oz'].astype(float).values,
                      index=pd.to_datetime(df['Date']) + pd.offsets.MonthEnd(0))
    return s.sort_index()


def compute_joint_stress(panel, sov_wam_days=90):
    """Correlations, triple-negative scan (BTC era), 1971+ gold-rates dual
    scan, and replay of worst episodes on the latest USDT book."""
    gold = load_gold_long()
    btc = pd.read_excel(EXCEL_PATH, sheet_name='BTC_Daily')[['Date', 'BTC_Price_USD']]
    btc = btc.set_index('Date')['BTC_Price_USD'].sort_index().resample('ME').last()
    tb = pd.read_excel(EXCEL_PATH, sheet_name='TB3MS')
    tb['date'] = pd.to_datetime(tb['observation_date']) + pd.offsets.MonthEnd(0)
    tb = tb.set_index('date')['TB3MS'].astype(float)

    # BTC-era monthly panel
    dm = pd.DataFrame({'btc': btc.pct_change(), 'gold': gold.pct_change(),
                       'dy_bp': tb.diff() * 100}).dropna()
    # quarterly
    dq = pd.DataFrame({'btc': btc.resample('QE').last().pct_change(),
                       'gold': gold.resample('QE').last().pct_change(),
                       'dy_bp': tb.resample('QE').last().diff() * 100}).dropna()
    dq['triple_negative'] = (dq.btc < 0) & (dq.gold < 0) & (dq.dy_bp > 0)

    # 1971+ free-float gold-rates panel (pre-BTC context)
    g71 = pd.DataFrame({'gold': gold.resample('QE').last().pct_change(),
                        'dy_bp': tb.resample('QE').last().diff() * 100}).dropna()
    g71 = g71[g71.index >= '1971-12-31']
    dual = g71[(g71.gold < -0.05) & (g71.dy_bp > 50)].sort_values('gold')

    # replay on latest USDT book
    last = panel.iloc[-1]
    btc_n  = float(last.get('USDT_BTC', 0) or 0) / 1e9
    gold_n = float(last.get('USDT_Gold', 0) or 0) / 1e9
    eq     = (float(last['USDT_Total_Assets']) - float(last['USDT_Circulation'])) / 1e9
    sov    = (float(last['USDT_Tbills_Direct']) + float(last['USDT_Total_Repos'])) / 1e9
    wamf   = sov_wam_days / 365

    def replay(btc_r, gold_r, dy_bp):
        return (btc_n * max(-btc_r, 0) + gold_n * max(-gold_r, 0)
                + sov * wamf * max(dy_bp, 0) / 10000)

    replays = []
    for d, r in dq[dq['triple_negative']].sort_values('btc').iterrows():
        L = replay(r.btc, r.gold, r.dy_bp)
        replays.append({'Episode': f'{d.year}-Q{d.quarter}', 'Type': 'triple (BTC era)',
                        'BTC_ret': round(r.btc, 3), 'Gold_ret': round(r.gold, 3),
                        'dy_bp': round(r.dy_bp, 0), 'Loss_bn': round(L, 2),
                        'Buffer_after_bn': round(eq - L, 2)})
    for d, r in dual.head(6).iterrows():
        L = replay(0, r.gold, r.dy_bp)  # no BTC leg pre-2013
        replays.append({'Episode': f'{d.year}-Q{d.quarter}', 'Type': 'dual (gold+rates)',
                        'BTC_ret': np.nan, 'Gold_ret': round(r.gold, 3),
                        'dy_bp': round(r.dy_bp, 0), 'Loss_bn': round(L, 2),
                        'Buffer_after_bn': round(eq - L, 2)})
    replays = pd.DataFrame(replays)

    summary = {
        'sample': f'{dm.index.min().date()} -> {dm.index.max().date()} ({len(dm)}m)',
        'corr_btc_gold': round(dm.btc.corr(dm.gold), 3),
        'corr_btc_dy': round(dm.btc.corr(dm.dy_bp), 3),
        'corr_gold_dy': round(dm.gold.corr(dm.dy_bp), 3),
        'corr_gold_dy_1971plus': round(
            gold.pct_change().to_frame('g').join(tb.diff().mul(100).rename('dy'))
                .dropna().loc['1971-08-31':].corr().loc['g', 'dy'], 3),
        'triple_months': int(((dm.btc < 0) & (dm.gold < 0) & (dm.dy_bp > 0)).sum()),
        'n_months': len(dm),
        'triple_quarters': int(dq['triple_negative'].sum()),
        'n_quarters': len(dq),
        'reserve_equity_bn': round(eq, 2)}
    return dm, dq, replays, summary


def make_joint_stress_figure(dm, dq):
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    axes[0].scatter(dm['dy_bp'], dm['gold'] * 100, s=18, alpha=0.7)
    axes[0].axhline(0, lw=0.6, color='k')
    axes[0].axvline(0, lw=0.6, color='k')
    axes[0].set_xlabel('d 3M yield (bp/month)')
    axes[0].set_ylabel('gold return (%)')
    axes[0].set_title(f"Gold vs rates, monthly "
                      f"(corr {dm.gold.corr(dm.dy_bp):+.2f})")
    colors = np.where(dq['triple_negative'], 'crimson', 'grey')
    axes[1].bar(range(len(dq)), dq['btc'] * 100, color=colors)
    step = max(1, len(dq) // 18)
    ticks = list(range(0, len(dq), step))
    axes[1].set_xticks(ticks)
    axes[1].set_xticklabels([f'{dq.index[i].year}-Q{dq.index[i].quarter}'
                             for i in ticks], rotation=60, fontsize=6)
    axes[1].set_ylabel('BTC quarterly return (%)')
    axes[1].set_title('Triple-negative quarters highlighted (BTC↓, gold↓, rates↑)')
    fig.tight_layout()
    return fig





def compute_combined_stress_timeseries(panel, rate_shocks_bp=(200, 400),
                                       sov_wam_days=90,
                                       replay=(0.573, 0.069, 105)):
    """Combined BTC+gold+rate stress for EVERY panel quarter (USDT).

    Grid: price shocks (PRICE_SHOCKS) x rate shocks (bp), plus a per-quarter
    replay of the worst observed triple-negative episode (2022-Q2 defaults:
    BTC -57.3%, gold -6.9%, +105bp). USDC/USDP hold no BTC/gold, so their
    combined stress reduces to the RQ2 rate leg and is not duplicated here.
    Pre-2023 caveat: BTC/gold zeros are a disclosure artifact (separate CRR
    categories only from Q1 2023), so early-quarter losses are UNDERSTATED.
    """
    wamf = sov_wam_days / 365
    rows = []
    for _, r in panel.iterrows():
        q = r['Quarter']
        btc  = float(r.get('USDT_BTC', 0) or 0) / 1e9
        gold = float(r.get('USDT_Gold', 0) or 0) / 1e9
        eq   = (float(r['USDT_Total_Assets']) - float(r['USDT_Circulation'])) / 1e9
        sov  = (float(r.get('USDT_Tbills_Direct', 0) or 0)
                + float(r.get('USDT_Total_Repos', 0) or 0)) / 1e9
        rep_loss = (btc * replay[0] + gold * replay[1]
                    + sov * wamf * replay[2] / 10000)
        for s in PRICE_SHOCKS:
            for dy in rate_shocks_bp:
                lb, lg = btc * s, gold * s
                lr = sov * wamf * dy / 10000
                lt = lb + lg + lr
                rows.append({
                    'Quarter': q, 'Price_Shock_pct': int(s * 100),
                    'Rate_Shock_bp': dy,
                    'BTC_Notional_bn': round(btc, 3),
                    'Gold_Notional_bn': round(gold, 3),
                    'Sov_Notional_bn': round(sov, 3),
                    'Reserve_Equity_bn': round(eq, 3),
                    'Loss_BTC_bn': round(lb, 3), 'Loss_Gold_bn': round(lg, 3),
                    'Loss_Rate_bn': round(lr, 3), 'Loss_Total_bn': round(lt, 3),
                    'Buffer_after_bn': round(eq - lt, 3),
                    'Survives': bool(eq - lt > 0),
                    'Replay_2022Q2_Loss_bn': round(rep_loss, 3),
                    'Replay_2022Q2_Buffer_after_bn': round(eq - rep_loss, 3)})
    return pd.DataFrame(rows)


def make_combined_stress_figure(cs):
    """Left: buffer after each price shock at +400bp, per quarter.
       Right: buffer after the 2022-Q2 replay vs unstressed equity."""
    fig, axes = plt.subplots(1, 2, figsize=(12.5, 5))
    qs = cs['Quarter'].drop_duplicates().tolist()
    x = range(len(qs))
    for s in sorted(cs['Price_Shock_pct'].unique()):
        sub = cs[(cs['Price_Shock_pct'] == s) & (cs['Rate_Shock_bp'] == 400)]
        axes[0].plot(x, sub['Buffer_after_bn'], marker='o', ms=3,
                     label=f'-{s}% prices +400bp')
    axes[0].axhline(0, color='k', lw=1)
    axes[0].set_xticks(list(x))
    axes[0].set_xticklabels(qs, rotation=45, ha='right', fontsize=7)
    axes[0].set_ylabel('buffer after shock ($bn)')
    axes[0].set_title('USDT combined stress timeline (price + rate legs)')
    axes[0].legend(fontsize=7)
    rep = cs.drop_duplicates('Quarter')
    axes[1].plot(x, rep['Reserve_Equity_bn'], marker='o', ms=3,
                 label='reserve equity (unstressed)')
    axes[1].plot(x, rep['Replay_2022Q2_Buffer_after_bn'], marker='s', ms=3,
                 color='darkred', label='buffer after 2022-Q2 replay')
    axes[1].axhline(0, color='k', lw=1)
    axes[1].set_xticks(list(x))
    axes[1].set_xticklabels(qs, rotation=45, ha='right', fontsize=7)
    axes[1].set_title('Worst observed episode replayed on each quarter')
    axes[1].legend(fontsize=8)
    fig.tight_layout()
    return fig


# ==============================================================================

def main():
    print("=" * 70)
    print("Thesis analysis script")
    print("=" * 70)
    print(f"Input  : {EXCEL_PATH.name}")
    print(f"Output : {OUT_DIR}")
    print(f"Scope  : {SCOPE_FIRST_Q} to {SCOPE_LAST_Q}")
    print()

    # ---- Load (USDC/USDT panel + USDP panel, the latter feeds 3-way figures) ----
    panel      = load_panel()
    usdp_panel = load_usdp_panel()
    print(f"Loaded {len(panel)} quarters (USDC/USDT) + {len(usdp_panel)} USDP quarters")

    # ---- SECTION 2: Mon-Tue deliverables ----
    print("\n" + "-" * 70)
    print("Mon-Tue: Descriptive table and composition figures (3-way)")
    print("-" * 70)

    table = build_descriptive_table(panel)
    table_path = OUT_DIR / 'descriptive_table.csv'
    table.to_csv(table_path, index=False)
    print(f"\n  Saved {table_path.name}")
    print("\n  Preview (first and last rows):")
    print(table.head(2).to_string(index=False))
    print("  ...")
    print(table.tail(2).to_string(index=False))

    fig = make_figures(panel, usdp_panel=usdp_panel)
    fig_path = OUT_DIR / 'fig_week1_descriptive.png'
    fig.savefig(fig_path, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"\n  Saved {fig_path.name}")

    # ---- Headline numbers for the session summary ----
    q4_2025 = panel[panel['Quarter'] == '2025-Q4'].iloc[0]
    print("\n" + "-" * 70)
    print("Headline numbers at Q4 2025 (USD billions)")
    print("-" * 70)
    print(f"  USDC supply: {q4_2025['USDC_Circulation']/1e9:6.2f}")
    print(f"  USDT supply: {q4_2025['USDT_Circulation']/1e9:6.2f}")
    print(f"  USDC market share: {q4_2025['USDC_Circulation']/(q4_2025['USDC_Circulation']+q4_2025['USDT_Circulation'])*100:5.1f}%")
    print(f"  USDT off-Basel assets (BTC + gold + secured loans): "
          f"${(q4_2025['USDT_BTC']+q4_2025['USDT_Gold']+q4_2025['USDT_Sec_Loans'])/1e9:.1f}bn")
    print(f"  USDC off-Basel assets: $0.0bn (100% sovereign / MMF / cash)")

    # ---- SECTION 3: Wed-Thu RWA computation ----
    print("\n" + "-" * 70)
    print("Wed-Thu: RWA computation (RQ1)")
    print("-" * 70)

    rwa = compute_rwa(panel)

    # Save CSV
    rwa_path = OUT_DIR / 'rwa_table.csv'
    rwa.to_csv(rwa_path, index=False)
    print(f"\n  Saved {rwa_path.name}  ({len(rwa)} rows)")

    # Print summary table
    print_rwa_summary(rwa)

    # Headline numbers
    r_q3_2022 = rwa[rwa['Quarter'] == '2022-Q3'].iloc[0]
    r_q4_2025 = rwa[rwa['Quarter'] == '2025-Q4'].iloc[0]
    print("\n  Key RWA findings:")
    print(f"    USDC RWA ratio range (primary):     "
          f"{rwa['USDC_RWA_Ratio_Primary'].min()*100:.2f}% – "
          f"{rwa['USDC_RWA_Ratio_Primary'].max()*100:.2f}%")
    print(f"    USDC RWA ratio range (sensitivity): "
          f"{rwa['USDC_RWA_Ratio_Sensitivity'].min()*100:.2f}% – "
          f"{rwa['USDC_RWA_Ratio_Sensitivity'].max()*100:.2f}%")
    print(f"    USDT RWA ratio range:               "
          f"{rwa['USDT_RWA_Ratio'].min()*100:.2f}% – "
          f"{rwa['USDT_RWA_Ratio'].max()*100:.2f}%")
    print(f"\n    USDT BTC RWA at Q4 2025: "
          f"${r_q4_2025['USDT_RWA_BTC']/1e9:.1f}bn  "
          f"(BTC notional: ${q4_2025['USDT_BTC']/1e9:.1f}bn × 1250%)")
    print(f"    USDT total RWA at Q4 2025: "
          f"${r_q4_2025['USDT_RWA_Total']/1e12:.3f}tn  "
          f"({r_q4_2025['USDT_RWA_Ratio']*100:.1f}% of total assets)")
    print(f"    USDC total RWA primary at Q4 2025: "
          f"${r_q4_2025['USDC_RWA_Total_Primary']/1e9:.2f}bn  "
          f"({r_q4_2025['USDC_RWA_Ratio_Primary']*100:.2f}% of total assets)")

    # ---- USDP RWA (narrow-bank benchmark) — uses usdp_panel loaded earlier ----
    usdp_rwa   = compute_usdp_rwa(usdp_panel)

    usdp_path  = OUT_DIR / 'usdp_rwa_table.csv'
    usdp_rwa.to_csv(usdp_path, index=False)
    print(f"\n  Saved {usdp_path.name}  ({usdp_rwa['USDP_Data_Available'].sum()} quarters with data)")

    print_usdp_rwa_summary(usdp_rwa)

    # USDP headline
    avail = usdp_rwa[usdp_rwa['USDP_Data_Available']]
    print(f"\n  USDP RWA ratio range (available quarters): "
          f"{avail['USDP_RWA_Ratio'].min()*100:.3f}% – "
          f"{avail['USDP_RWA_Ratio'].max()*100:.3f}%")
    print(f"  USDP Q4 2025: circulation=${avail.iloc[-1]['USDP_Circulation']/1e6:.1f}m  "
          f"RWA ratio={avail.iloc[-1]['USDP_RWA_Ratio']*100:.3f}%")

    # Figures
    fig_evo = make_rwa_evolution(rwa)
    evo_path = OUT_DIR / 'rwa_evolution.png'
    fig_evo.savefig(evo_path, dpi=130, bbox_inches='tight')
    plt.close(fig_evo)
    print(f"\n  Saved {evo_path.name}")

    fig_ratio = make_rwa_ratio(rwa)
    ratio_path = OUT_DIR / 'rwa_ratio.png'
    fig_ratio.savefig(ratio_path, dpi=130, bbox_inches='tight')
    plt.close(fig_ratio)
    print(f"  Saved {ratio_path.name}")

    # Three-way chart (USDC + USDP + USDT)
    fig_3way = make_rwa_ratio_threeway(rwa, usdp_rwa)
    threeway_path = OUT_DIR / 'rwa_ratio_threeway.png'
    fig_3way.savefig(threeway_path, dpi=130, bbox_inches='tight')
    plt.close(fig_3way)
    print(f"  Saved {threeway_path.name}")

    # ---- SECTION 4: Week 2 Mon-Tue — Interest-rate sensitivity (RQ2) ----
    print("\n" + "-" * 70)
    print("Week 2 Mon-Tue: Interest-rate (duration) stress (RQ2)")
    print("-" * 70)

    grid, notion = compute_duration_stress(panel, usdp_panel)

    grid_path = OUT_DIR / 'mtm_loss_grid.csv'
    grid.to_csv(grid_path, index=False)
    print(f"\n  Saved {grid_path.name}  ({len(grid)} rows: 3 issuers x "
          f"{len(WAM_DAYS)} WAM x {len(SHOCKS_BP)} shocks)")

    print_duration_summary(grid, notion)

    fig_dur = make_duration_stress_figure(grid, notion)
    dur_path = OUT_DIR / 'fig_duration_stress.png'
    fig_dur.savefig(dur_path, dpi=130, bbox_inches='tight')
    plt.close(fig_dur)
    print(f"\n  Saved {dur_path.name}")

    # ---- Time-series duration stress (all 14 quarters, fixed stress) ----
    ts = compute_duration_stress_timeseries(panel, usdp_panel)
    ts_path = OUT_DIR / 'mtm_loss_timeseries.csv'
    ts.to_csv(ts_path, index=False)
    print(f"  Saved {ts_path.name}  ({len(ts)} rows: 14 quarters x 3 issuers x "
          f"{len(TS_PAIRS)} scenarios)")

    fig_ts = make_duration_timeseries_figure(ts)
    ts_fig_path = OUT_DIR / 'fig_duration_timeseries.png'
    fig_ts.savefig(ts_fig_path, dpi=130, bbox_inches='tight')
    plt.close(fig_ts)
    print(f"  Saved {ts_fig_path.name}")

    # Time-series headline: adverse-ceiling peak per issuer
    adverse = ts[ts['Scenario'] == TS_PAIRS[0][2]]
    print("\n  Adverse-ceiling (180d,+400bp) peak MtM loss across quarters:")
    for issuer in ['USDC', 'USDT', 'USDP']:
        iss = adverse[adverse['Issuer'] == issuer].dropna(subset=['MtM_Loss_pct_of_assets'])
        if len(iss):
            pk = iss.loc[iss['MtM_Loss_pct_of_assets'].idxmax()]
            print(f"    {issuer}: peak {pk['MtM_Loss_pct_of_assets']:.4f}% of assets "
                  f"at {pk['Quarter']} (${pk['MtM_Loss_bn']:.3f}bn)")

    # ---- Historical rate-shock context from the Yield_Curve sheet ----
    yc = load_yield_curve()
    hstats, s3m, rolled = compute_historical_shock_context(yc)
    print("\n  Historical front-end context (DGS3MO, in-sample 2020-2025):")
    print(f"    2022-23 cycle: {hstats['cycle_low_pct']:.2f}% ({hstats['cycle_low_date']}) "
          f"-> {hstats['cycle_high_pct']:.2f}% ({hstats['cycle_high_date']}) "
          f"= +{hstats['cycle_move_bp']:.0f}bp")
    print(f"    Peak trailing-12m rise: +{hstats['max_trailing_12m_bp']:.0f}bp "
          f"on {hstats['max_trailing_12m_date']}")
    print(f"    Trading days with trailing-12m rise >= +{hstats['threshold_bp']}bp: "
          f"{hstats['days_over_threshold']} "
          f"({hstats['over_first']} to {hstats['over_last']})")

    fig_hist = make_historical_context_figure(s3m, rolled)
    hist_path = OUT_DIR / 'fig_rate_history.png'
    fig_hist.savefig(hist_path, dpi=130, bbox_inches='tight')
    plt.close(fig_hist)
    print(f"  Saved {hist_path.name}")

    # ---- Long-history rate-shock frequency (TB3MS, 1934-present) ----
    long_s, long_basis = load_long_bill_history()
    if long_s is not None:
        lstats, leps, lrolled = count_shock_episodes(long_s)
        print(f"\n  Long-history frequency (TB3MS, {long_basis} basis, "
              f"{lstats['span'][0]}\u2013{lstats['span'][1]}, ~{lstats['span_years']:.0f}y):")
        print(f"    +{lstats['threshold_bp']}bp / {lstats['horizon_months']}m episodes: "
              f"{lstats['n_episodes']}  (~{lstats['per_decade_rate']:.2f} per decade)")
        for ep in leps:
            print(f"      {ep['start']} to {ep['end']}  | peak +{ep['peak_bp']:.0f}bp "
                  f"({ep['peak_date']})  | {ep['n_months']} months")
        print(f"    All-time max trailing-{lstats['horizon_months']}m rise: "
              f"+{lstats['largest_move_bp']:.0f}bp on {lstats['largest_move_date']}")

        fig_long = make_long_history_figure(long_s, lrolled, long_basis)
        long_path = OUT_DIR / 'fig_rate_history_long.png'
        fig_long.savefig(long_path, dpi=130, bbox_inches='tight')
        plt.close(fig_long)
        print(f"  Saved {long_path.name}")

        # Persist the episode table for the thesis.
        ep_df = pd.DataFrame(leps)
        ep_path = OUT_DIR / 'rate_shock_episodes.csv'
        ep_df.to_csv(ep_path, index=False)
        print(f"  Saved {ep_path.name}  ({lstats['n_episodes']} episodes)")
    else:
        print("\n  [long-history] TB3MS sheet not found in the master Excel; "
              "skipping the 1934-present frequency analysis.")

    # ---- SECTION 5: Week 2 Wed-Thu — Redemption / LCR liquidity stress (RQ3) ----
    print("\n" + "-" * 70)
    print("Week 2 Wed-Thu: Redemption / LCR liquidity stress (RQ3)")
    print("-" * 70)

    lcr_grid = compute_lcr_grid(panel, usdp_panel)
    lcr_grid_path = OUT_DIR / 'lcr_stress_grid.csv'
    lcr_grid.to_csv(lcr_grid_path, index=False)
    print(f"\n  Saved {lcr_grid_path.name}  ({len(lcr_grid)} rows: "
          f"latest quarter x issuers x {len(OUTFLOW_SCENARIOS)} outflows x treatments)")

    lcr_ts = compute_lcr_timeseries(panel, usdp_panel)
    lcr_ts_path = OUT_DIR / 'lcr_timeseries.csv'
    lcr_ts.to_csv(lcr_ts_path, index=False)
    print(f"  Saved {lcr_ts_path.name}  ({len(lcr_ts)} rows: 14 quarters x "
          f"{len(OUTFLOW_SCENARIOS)} outflows)")

    print_lcr_summary(lcr_grid, lcr_ts)

    fig_lcr = make_lcr_stress_figure(lcr_ts)
    lcr_fig_path = OUT_DIR / 'fig_lcr_stress.png'
    fig_lcr.savefig(lcr_fig_path, dpi=130, bbox_inches='tight')
    plt.close(fig_lcr)
    print(f"\n  Saved {lcr_fig_path.name}")

    fig_comb = make_lcr_vs_rate_combined(panel, usdp_panel, lcr_ts)
    comb_path = OUT_DIR / 'fig_lcr_vs_rate_combined.png'
    fig_comb.savefig(comb_path, dpi=130, bbox_inches='tight')
    plt.close(fig_comb)
    print(f"  Saved {comb_path.name}")

    # ---- RQ2 x RQ3 combined effective coverage (forced-sale realised loss) ----
    eff = compute_effective_lcr(panel, usdp_panel)
    eff_path = OUT_DIR / 'lcr_effective_coverage.csv'
    eff.to_csv(eff_path, index=False)
    print(f"  Saved {eff_path.name}  ({len(eff)} rows)")
    print("\n  Combined coverage at 40% run, +400bp/180d (effective vs plain LCR):")
    e40 = eff[eff['Outflow_pct'] == 40]
    latest_e = e40[e40['Quarter'] == SCOPE_LAST_Q].iloc[0]
    for iss in ['USDC', 'USDT', 'USDP']:
        p = latest_e[f'{iss}_LCR_plain']; ef = latest_e[f'{iss}_LCR_effective']
        e5 = latest_e[f'{iss}_LCR_effective_5y']
        if p == p:
            print(f"    {iss}: plain {p:.1f}%  ->  effective {ef:.1f}%  "
                  f"(short book); 5y-counterfactual {e5:.1f}%")
    fig_eff = make_effective_lcr_figure(eff, outflow_pct=40)
    eff_fig_path = OUT_DIR / 'fig_lcr_effective.png'
    fig_eff.savefig(eff_fig_path, dpi=130, bbox_inches='tight')
    plt.close(fig_eff)
    print(f"  Saved {eff_fig_path.name}")

    # ---- Discussion scaffold: stablecoin T-bill footprint vs the market ----
    try:
        bills_df = load_bills_outstanding()
    except Exception as e:
        bills_df = None
        print(f"\n  [tbill-footprint] MSPD_SumSecty sheet unavailable ({e}); "
              f"skipping the T-bill footprint discussion figure.")
    if bills_df is not None:
        share_df = compute_tbill_share(panel, bills_df)
        share_path = OUT_DIR / 'stablecoin_tbill_share.csv'
        share_df.to_csv(share_path, index=False)
        print(f"  Saved {share_path.name}  (sector T-bill share of total bills)")
        print("\n  Stablecoin (USDC+USDT) T-bill share of total marketable bills:")
        print(f"    {share_df.iloc[0]['Quarter']}: {share_df.iloc[0]['Share_pct']:.2f}%  "
              f"(${share_df.iloc[0]['Sector_Tbills_bn']:.0f}bn)")
        print(f"    {share_df.iloc[-1]['Quarter']}: {share_df.iloc[-1]['Share_pct']:.2f}%  "
              f"(${share_df.iloc[-1]['Sector_Tbills_bn']:.0f}bn)")
        fig_fp = make_tbill_footprint_figure(share_df)
        fp_path = OUT_DIR / 'fig_tbill_footprint.png'
        fig_fp.savefig(fp_path, dpi=130, bbox_inches='tight')
        plt.close(fig_fp)
        print(f"  Saved {fp_path.name}")

    # ---- SECTION 6: Week 3 — Narrow-bank vs MMF classification (RQ4) ----
    print("\n" + "-" * 70)
    print("Week 3: Narrow-bank vs MMF regulatory classification (RQ4)")
    print("-" * 70)

    scorecard = compute_classification_scorecard(
        panel, usdp_panel, rwa, usdp_rwa, ts, lcr_ts)
    sc_path = OUT_DIR / 'classification_scorecard.csv'
    scorecard.to_csv(sc_path, index=False)
    print(f"\n  Saved {sc_path.name}  ({len(scorecard)} issuers x "
          f"{len(CLASS_ANCHORS)} axes)")

    # ---- Reserve-level capital bridge (corrected: reserve equity / reserve RWA) ----
    bridge = compute_capital_bridge(panel, usdp_panel, rwa, usdp_rwa)
    bridge_path = OUT_DIR / 'capital_bridge.csv'
    bridge.to_csv(bridge_path, index=False)
    print(f"  Saved {bridge_path.name}  ({len(bridge)} issuers, reserve level)")

    cap_ts = compute_capital_bridge_timeseries(panel, usdp_panel, rwa, usdp_rwa)
    cap_ts_path = OUT_DIR / 'capital_bridge_timeseries.csv'
    cap_ts.to_csv(cap_ts_path, index=False)
    print(f"  Saved {cap_ts_path.name}  ({len(cap_ts)} rows)")
    print_capital_bridge(bridge, cap_ts)

    fig_cap = make_capital_ratio_figure(cap_ts)
    cap_fig_path = OUT_DIR / 'fig_capital_ratio.png'
    fig_cap.savefig(cap_fig_path, dpi=130, bbox_inches='tight')
    plt.close(fig_cap)
    print(f"  Saved {cap_fig_path.name}")

    # ---- Tether equity & profit (supervisor's headline request) ----
    tep = load_tether_equity_profit()
    tcr = compute_tether_capital_ratio(tep, rwa)
    tcr_path = OUT_DIR / 'tether_capital_ratio.csv'
    tcr.to_csv(tcr_path, index=False)
    print(f"  Saved {tcr_path.name}  ({tcr['Reserve_RW_Equity_pct'].notna().sum()} quarters with equity data)")
    print_tether_capital_ratio(tcr, tep)

    # ---- Circle / USDC equity & profit (group context; SEC filings) ----
    cep = load_circle_equity_profit()
    cep_path = OUT_DIR / 'circle_equity_profit.csv'
    cep.to_csv(cep_path, index=False)
    print(f"\n  Saved {cep_path.name}  ({cep['Group_Equity_m'].notna().sum()} quarters with group equity)")
    print_circle_equity_profit(cep)

    print_classification_summary(scorecard)

    fig_spec = make_classification_spectrum(scorecard)
    spec_path = OUT_DIR / 'fig_classification_spectrum.png'
    fig_spec.savefig(spec_path, dpi=130, bbox_inches='tight')
    plt.close(fig_spec)
    print(f"\n  Saved {spec_path.name}")

    # ---- Weight robustness / sensitivity (OECD-JRC required step) ----
    schemes_df, mc_summary = compute_weight_sensitivity(scorecard)
    ws_path = OUT_DIR / 'classification_weight_sensitivity.csv'
    schemes_df.to_csv(ws_path, index=False)
    print(f"  Saved {ws_path.name}")
    print_weight_sensitivity(schemes_df, mc_summary)

    fig_ws = make_weight_sensitivity_figure(scorecard, schemes_df, mc_summary)
    ws_fig_path = OUT_DIR / 'fig_weight_sensitivity.png'
    fig_ws.savefig(ws_fig_path, dpi=130, bbox_inches='tight')
    plt.close(fig_ws)
    print(f"  Saved {ws_fig_path.name}")

    # ---- SECTION 7: Data-driven structural-event detection (validation) ----
    print("\n" + "-" * 70)
    print("Data-driven structural-event detection")
    print("-" * 70)
    events = detect_structural_events()
    ev_path = OUT_DIR / 'structural_events_detected.csv'
    events.to_csv(ev_path, index=False)
    print(f"\n  Saved {ev_path.name}  ({len(events)} events)")
    print_event_summary(events)
    fig_ev = make_event_timeline_figure(events)
    ev_fig_path = OUT_DIR / 'fig_structural_events.png'
    fig_ev.savefig(ev_fig_path, dpi=130, bbox_inches='tight')
    plt.close(fig_ev)
    print(f"  Saved {ev_fig_path.name}")

    # ---- SECTION 8: Supervisor additions ----
    print("\n" + "-" * 70)
    print("Section 8: bank benchmark, price shocks, retention, joint stress")
    print("-" * 70)
    bb = compute_bank_benchmark(cap_ts)
    bb.to_csv(OUT_DIR / 'bank_benchmark.csv', index=False)
    fig_bb = make_bank_benchmark_figure(bb)
    fig_bb.savefig(OUT_DIR / 'fig_bank_benchmark.png', dpi=130, bbox_inches='tight')
    plt.close(fig_bb)
    print(f"  Saved bank_benchmark.csv + figure "
          f"(JPM {bb['JPM_CET1_pct'].iloc[-1]}%, BNY {bb['BNY_CET1_pct'].iloc[-1]}% "
          f"vs USDT {bb['USDT_RWeq_pct'].iloc[-1]}% in {bb['Quarter'].iloc[-1]})")

    sp = compute_price_shock_panel(panel)
    sp.to_csv(OUT_DIR / 'btc_gold_shock_panel.csv', index=False)
    fig_sp = make_price_shock_figure(sp)
    fig_sp.savefig(OUT_DIR / 'fig_btc_gold_shocks.png', dpi=130, bbox_inches='tight')
    plt.close(fig_sp)
    be_last = sp.dropna(subset=['Breakeven_Combined_pct'])
    if len(be_last):
        print(f"  Saved btc_gold_shock_panel.csv + figure "
              f"(combined break-even {be_last['Breakeven_Combined_pct'].iloc[-1]}% "
              f"in {be_last['Quarter'].iloc[-1]})")

    ra = compute_retention_attribution()
    ra.to_csv(OUT_DIR / 'retained_earnings_attribution.csv', index=False)
    print("  Saved retained_earnings_attribution.csv")
    for _, r in ra.iterrows():
        if not np.isnan(r['Retention_pct']):
            print(f"    {r['Issuer']} {r['Year']}: retention {r['Retention_pct']}% "
                  f"(distributed ${r['Implied_Distribution_m']:,.0f}m)")

    dm, dq, replays, jsum = compute_joint_stress(panel)
    dm.round(5).to_csv(OUT_DIR / 'btc_gold_rates_monthly.csv')
    dq.round(5).to_csv(OUT_DIR / 'btc_gold_rates_quarterly.csv')
    replays.to_csv(OUT_DIR / 'joint_stress_replays.csv', index=False)
    fig_js = make_joint_stress_figure(dm, dq)
    fig_js.savefig(OUT_DIR / 'fig_joint_stress.png', dpi=130, bbox_inches='tight')
    plt.close(fig_js)
    print(f"  Joint stress ({jsum['sample']}): corr BTCxGold {jsum['corr_btc_gold']}, "
          f"GoldxDy {jsum['corr_gold_dy']}; triple-negative "
          f"{jsum['triple_months']}/{jsum['n_months']} months, "
          f"{jsum['triple_quarters']}/{jsum['n_quarters']} quarters")
    if len(replays):
        w = replays.sort_values('Buffer_after_bn').iloc[0]
        print(f"    Worst replay {w['Episode']} ({w['Type']}): loss "
              f"${w['Loss_bn']}bn -> buffer after ${w['Buffer_after_bn']}bn "
              f"of ${jsum['reserve_equity_bn']}bn")

    cs = compute_combined_stress_timeseries(panel)
    cs.to_csv(OUT_DIR / 'combined_stress_timeseries.csv', index=False)
    fig_cs = make_combined_stress_figure(cs)
    fig_cs.savefig(OUT_DIR / 'fig_combined_stress_timeline.png', dpi=130,
                   bbox_inches='tight')
    plt.close(fig_cs)
    fails = cs[~cs['Survives']]
    frontier = (fails.sort_values(['Price_Shock_pct', 'Rate_Shock_bp'])
                     .groupby('Quarter').first())
    print(f"  Saved combined_stress_timeseries.csv ({len(cs)} rows) + figure")
    for q in cs['Quarter'].drop_duplicates():
        if q in frontier.index:
            f = frontier.loc[q]
            print(f"    {q}: mildest failing scenario -{f['Price_Shock_pct']}% "
                  f"/ +{f['Rate_Shock_bp']}bp (buffer {f['Buffer_after_bn']})")
        else:
            print(f"    {q}: survives all scenarios up to -50% / +400bp")

    # Headline through-line
    usdp_s = scorecard[scorecard['Issuer'] == 'USDP'].iloc[0]
    usdc_s = scorecard[scorecard['Issuer'] == 'USDC'].iloc[0]
    usdt_s = scorecard[scorecard['Issuer'] == 'USDT'].iloc[0]
    print("\n  RQ4 through-line:")
    print(f"    USDP {usdp_s['Score']:.0f} ({usdp_s['Class']})  ->  "
          f"USDC {usdc_s['Score']:.0f} ({usdc_s['Class']})  ->  "
          f"USDT {usdt_s['Score']:.0f} ({usdt_s['Class']})")
    print("    Asset composition + disclosure (not size) drive the spread.")

    print("\nDone.")


if __name__ == '__main__':
    main()


# ==============================================================================
# SECTION 9: Publication figures for the written thesis (Chapters 4-7)
# ------------------------------------------------------------------------------
# WHAT THIS SECTION PRODUCES
#   outputs/thesis_figures/
#     fig_4_1_rwa_density.png         -> Section 4.2,   after Table 4.2
#     fig_4_2_bank_corridor.png       -> Section 4.3.2, after Table 4.4
#     fig_5_1_rate_shock_history.png  -> Section 5.2,   after Table 5.1
#     fig_6_1_lcr_40pct.png           -> Section 6.3.1, after Table 6.2
#     fig_6_2_combined_stress.png     -> Section 6.3.4, after Table 6.4
#     fig_6_3_tbill_share.png         -> Section 6.2   (footprint paragraph)
#     fig_7_1_classification.png      -> Section 7.3.2, after Table 7.3
#
# DESIGN RULES (locked)
#   - Reads ONLY locked outputs: the Stablecoin Reserve Panel (the master
#     Excel at EXCEL_PATH, cited in the thesis as Infante, 2026) and the CSVs
#     written by Sections 3-8 into OUT_DIR. This section recomputes NOTHING.
#     The only arithmetic is ratio-of-locked-values (RWA / total assets) and
#     the trailing-12-month difference of raw TB3MS -- the documented episode-
#     detection input of Section 4; the episodes themselves are read from
#     rate_shock_episodes.csv.
#   - Two short series are transcribed from locked, thesis-printed tables
#     rather than re-derived: the USDC opaque-treatment density (Table 4.2)
#     and the Q4-2025 mildest-failing-cell label (Table 6.4).
#   - Publication style (serif, 300 dpi, no in-image titles; captions live in
#     the thesis) is applied via rc_context so the working-figure style of
#     Section 0 is untouched.
#   - Missing optional inputs skip the single affected figure with a message;
#     they never raise. Break-even values are printed for verification against
#     Table 6.4.
# ==============================================================================

THESIS_FIG_DIR = OUT_DIR / 'thesis_figures'

_TF_C = {'USDC': '#1f4e8c', 'USDT': '#1a8f6d', 'USDP': '#c17817',
         'red': '#b3392f', 'gray': '#5a5f66', 'band': '#e8eaed'}
_TF_STYLE = {
    'font.family': 'serif', 'font.size': 9, 'axes.labelsize': 9.5,
    'axes.spines.top': False, 'axes.spines.right': False,
    'axes.grid': True, 'grid.color': '#d9dce1', 'grid.linewidth': 0.5,
    'legend.frameon': False, 'legend.fontsize': 8.5,
    'savefig.dpi': 300, 'savefig.bbox': 'tight', 'figure.facecolor': 'white',
}
_TF_QUARTERS = [f'{y}-Q{q}' for y in (2022, 2023, 2024, 2025)
                for q in (1, 2, 3, 4)][2:]          # 2022-Q3 .. 2025-Q4

# Locked transcriptions (see DESIGN RULES above).
_TF_USDC_OPAQUE_DENSITY = [3.88, 15.29] + [20.0] * 12        # Table 4.2
_TF_MILDEST_FAIL_LABEL = '\u221225%; mildest grid cell\n(\u221220% / +400 bp) fails'  # Table 6.4


def _tf_qticks(ax, quarters):
    ax.set_xticks(range(len(quarters)))
    ax.set_xticklabels([f"Q{q.split('-Q')[1]}\n{q[:4]}" for q in quarters],
                       fontsize=7.8)
    ax.set_xlim(-0.5, len(quarters) - 0.5)


def _tf_series(df, issuer, col):
    return (df[df['Issuer'] == issuer].set_index('Quarter')[col]
            .reindex(_TF_QUARTERS).values.astype(float))


def _tf_fig_4_1(cb, mtm, usdp_rwa):
    fig, ax = plt.subplots(figsize=(7.3, 3.9))
    for iss in ('USDT', 'USDP', 'USDC'):
        if iss == 'USDP':
            d = (usdp_rwa.set_index('Quarter')['USDP_RWA_Ratio']
                 .reindex(_TF_QUARTERS) * 100).values
        else:
            d = (_tf_series(cb, iss, 'RWA_bn')
                 / _tf_series(mtm, iss, 'Total_Assets_bn') * 100)
        lbl = iss + (' (look-through)' if iss == 'USDC' else '')
        ax.plot(range(14), d, marker='o', ms=3.5, lw=1.6, color=_TF_C[iss], label=lbl)
    ax.plot(range(14), _TF_USDC_OPAQUE_DENSITY, ls=':', lw=1.4,
            color=_TF_C['USDC'], alpha=0.75, label='USDC opaque (sensitivity)')
    ax.annotate('', xy=(9, 20), xytext=(9, 3.3),
                arrowprops=dict(arrowstyle='<->', color=_TF_C['USDC'], lw=1))
    ax.text(9.25, 11.3, 'look-through\npremium \u2248 17 pp', fontsize=8, color=_TF_C['USDC'])
    ax.set_ylabel('RWA density (RWA / total reserve assets, %)')
    ax.set_ylim(0, 92); _tf_qticks(ax, _TF_QUARTERS)
    ax.legend(loc='upper left', ncol=2)
    fig.savefig(THESIS_FIG_DIR / 'fig_4_1_rwa_density.png'); plt.close(fig)


def _tf_fig_4_2(bank):
    fig, ax = plt.subplots(figsize=(7.3, 3.9))
    x = range(14)
    ax.fill_between(x, bank['BNY'].astype(float), bank['JPM'].astype(float),
                    color=_TF_C['band'], zorder=0,
                    label='US bank CET1 corridor (BNY\u2013JPM)')
    for b in ('JPM', 'BNY'):
        ax.plot(x, bank[b].astype(float), color='#8a8f98', lw=1.1)
        ax.text(13.15, float(bank[b].iloc[-1]), b, fontsize=8,
                color='#6b7078', va='center')
    for iss in ('USDC', 'USDT', 'USDP'):
        ax.plot(x, bank[iss].astype(float), marker='o', ms=3.5, lw=1.6,
                color=_TF_C[iss], label=f'{iss} reserve RW-equity')
    ax.axhline(7, ls='--', lw=1.1, color=_TF_C['red'])
    ax.text(0.05, 7.4, '7% Basel-analogous minimum', fontsize=8, color=_TF_C['red'])
    ax.annotate('USDP Q2 2023: 50.2%', xy=(3, 27.6), xytext=(4.1, 25.5), fontsize=8,
                color=_TF_C['USDP'],
                arrowprops=dict(arrowstyle='->', color=_TF_C['USDP'], lw=0.9))
    ax.set_ylabel('Equity / risk-weighted assets (%)')
    ax.set_ylim(0, 28.5); _tf_qticks(ax, _TF_QUARTERS)
    ax.legend(loc='upper right', ncol=2)
    fig.savefig(THESIS_FIG_DIR / 'fig_4_2_bank_corridor.png'); plt.close(fig)


def _tf_fig_5_1(tb_series, eps):
    d12 = (tb_series.diff(12) * 100).dropna()   # bp; documented detection input
    fig, ax = plt.subplots(figsize=(7.3, 3.7))
    ax.plot(d12.index, d12.values, lw=0.8, color='#4a6b8a')
    ax.axhline(400, ls='--', lw=1.1, color=_TF_C['red'])
    ax.axhline(0, lw=0.6, color='#9aa0a6')
    ax.text(pd.Timestamp('1935-06-01'), 425, '+400 bp threshold',
            fontsize=8, color=_TF_C['red'])
    labels = {1973: '1973\n+465', 1980: '1980\n+572',
              1981: '1981\n+772', 2022: '2022\u201323\n+439'}
    for _, r in eps.iterrows():
        ax.axvspan(r['start'] - pd.DateOffset(months=2),
                   r['end'] + pd.DateOffset(months=2),
                   color=_TF_C['red'], alpha=0.15, lw=0)
        ax.annotate(labels[r['start'].year], xy=(r['peak_date'], r['peak_bp']),
                    xytext=(r['peak_date'], r['peak_bp'] + 60),
                    fontsize=7.8, ha='center', color='#7a2620')
    ax.axvspan(pd.Timestamp('2022-07-01'), pd.Timestamp('2026-01-01'),
               color='#1f4e8c', alpha=0.07, lw=0)
    ax.text(pd.Timestamp('2013-01-01'), -540, 'panel window \u2192',
            fontsize=8, color='#1f4e8c')
    ax.set_ylabel('Trailing-12-month change in 3-month\nT-bill yield (basis points)')
    ax.set_ylim(-620, 900)
    fig.savefig(THESIS_FIG_DIR / 'fig_5_1_rate_shock_history.png'); plt.close(fig)


def _tf_fig_6_1(lcr):
    l40 = lcr[lcr['Outflow_pct'] == 40].set_index('Quarter').reindex(_TF_QUARTERS)
    fig, ax = plt.subplots(figsize=(7.3, 3.9))
    for iss in ('USDC', 'USDT', 'USDP'):
        lbl = iss + (' (look-through)' if iss == 'USDC' else '')
        ax.plot(range(14), l40[f'{iss}_LCR_plain'], marker='o', ms=3.5, lw=1.6,
                color=_TF_C[iss], label=lbl)
    ax.axhline(100, ls='--', lw=1.1, color=_TF_C['red'])
    ax.text(13.35, 106, 'pass\n(100%)', fontsize=8, color=_TF_C['red'],
            va='bottom', ha='center')
    ax.annotate('Q4 2022: 59% \u2014\nfund-transition artifact', xy=(1, 59.1),
                xytext=(1.7, 22), fontsize=8, color=_TF_C['USDC'],
                arrowprops=dict(arrowstyle='->', color=_TF_C['USDC'], lw=0.9))
    ax.annotate('USDP wind-down:\nsovereign book runs off to cash', xy=(8, 2),
                xytext=(6.1, 118), fontsize=8, color=_TF_C['USDP'],
                arrowprops=dict(arrowstyle='->', color=_TF_C['USDP'], lw=0.9))
    ax.set_ylabel('LCR at 40% redemption run (%)')
    ax.set_ylim(0, 248); _tf_qticks(ax, _TF_QUARTERS)
    ax.legend(loc='lower left', bbox_to_anchor=(0.02, 0.02))
    fig.savefig(THESIS_FIG_DIR / 'fig_6_1_lcr_40pct.png'); plt.close(fig)


def _tf_fig_6_2(bg):
    q12 = _TF_QUARTERS[2:]                       # 2023-Q1 onward (first disclosure)
    sleeve = (bg['BTC'] + bg['Gold']).reindex(q12).astype(float)
    buf = bg['Equity'].reindex(q12).astype(float)
    brk = (bg['Breakeven'].reindex(q12).astype(float) * 100)
    fig, ax = plt.subplots(figsize=(7.3, 4.0))
    x = range(12)
    ax.plot(x, sleeve.values, marker='o', ms=4, lw=1.8, color=_TF_C['red'],
            label='Bitcoin + gold sleeve ($bn, left)')
    ax.plot(x, buf.values, marker='s', ms=3.5, lw=1.8, color=_TF_C['USDT'],
            label='Reserve buffer ($bn, left)')
    ax.fill_between(x, buf.values, sleeve.values,
                    where=sleeve.values > buf.values,
                    color=_TF_C['red'], alpha=0.06, lw=0)
    ax.set_ylabel('$ billion'); ax.set_ylim(0, 28)
    ax2 = ax.twinx(); ax2.spines['top'].set_visible(False); ax2.grid(False)
    ax2.plot(x, brk.values, ls='--', marker='D', ms=3.2, lw=1.3, color=_TF_C['gray'],
             label='Break-even combined price shock (|%|, right)')
    ax2.set_ylabel('Break-even shock, absolute (%)', color=_TF_C['gray'])
    ax2.tick_params(axis='y', colors=_TF_C['gray']); ax2.set_ylim(0, 100)
    ax2.annotate('peak resilience:\n\u221286% (Q4 2023)', xy=(3, 86), xytext=(0.4, 68),
                 fontsize=8, color=_TF_C['gray'],
                 arrowprops=dict(arrowstyle='->', color=_TF_C['gray'], lw=0.9))
    ax2.annotate(_TF_MILDEST_FAIL_LABEL, xy=(11, 25), xytext=(7.6, 8),
                 fontsize=8, color=_TF_C['gray'],
                 arrowprops=dict(arrowstyle='->', color=_TF_C['gray'], lw=0.9))
    ax.set_xticks(x)
    ax.set_xticklabels([f"Q{q.split('-Q')[1]}\n{q[:4]}" for q in q12], fontsize=7.8)
    h1, l1 = ax.get_legend_handles_labels(); h2, l2 = ax2.get_legend_handles_labels()
    ax.legend(h1 + h2, l1 + l2, loc='upper left', fontsize=8.2)
    fig.savefig(THESIS_FIG_DIR / 'fig_6_2_combined_stress.png'); plt.close(fig)
    return dict(zip(q12, brk.round(0).values))


def _tf_fig_6_3(share):
    sh = share.set_index('Quarter').reindex(_TF_QUARTERS)
    fig, ax = plt.subplots(figsize=(7.3, 3.6))
    ax.bar(range(14), sh['Sector_Tbills_bn'], color='#8fa8c4', width=0.62,
           label='USDC + USDT direct T-bills ($bn, left)')
    ax.set_ylabel('$ billion'); ax.set_ylim(0, 155)
    ax2 = ax.twinx(); ax2.spines['top'].set_visible(False); ax2.grid(False)
    ax2.plot(range(14), sh['Share_pct'], marker='o', ms=3.5, lw=1.6, color=_TF_C['red'],
             label='Share of marketable bills outstanding (%, right)')
    ax2.set_ylabel('% of bills outstanding', color=_TF_C['red'])
    ax2.tick_params(axis='y', colors=_TF_C['red']); ax2.set_ylim(0, 3.4)
    _tf_qticks(ax, _TF_QUARTERS)
    h1, l1 = ax.get_legend_handles_labels(); h2, l2 = ax2.get_legend_handles_labels()
    ax.legend(h1 + h2, l1 + l2, loc='upper left', fontsize=8.2)
    fig.savefig(THESIS_FIG_DIR / 'fig_6_3_tbill_share.png'); plt.close(fig)


def _tf_fig_7_1(score, schemes):
    nice = {'Equal': 'Equal weights', 'Drop rwa': 'Drop RWA density',
            'Drop mtm': 'Drop MtM', 'Drop lcr': 'Drop LCR',
            'Drop risk': 'Drop off-Basel', 'Drop disc': 'Drop disclosure'}
    schemes = schemes.copy(); schemes['Scheme'] = schemes['Scheme'].replace(nice)
    sc = schemes.set_index('Scheme')
    fig, (a1, a2) = plt.subplots(2, 1, figsize=(7.3, 4.6), height_ratios=[1, 1.7],
                                 sharex=True, gridspec_kw={'hspace': 0.12})
    for a in (a1, a2):
        a.axvspan(0, 33, color='#e2efe6', lw=0)
        a.axvspan(33, 50, color='#f7f0da', lw=0)
        a.axvspan(50, 100, color='#f7e4e1', lw=0)
        a.set_xlim(0, 100); a.grid(False)
    for xx, txt, col in ((16.5, 'narrow-bank (<33)', '#3e6b4f'),
                         (41.5, 'hybrid', '#8a6d1f'),
                         (75, 'MMF-like (\u226550)', '#8c3a32')):
        a1.text(xx, 1.02, txt, fontsize=8, ha='center', color=col,
                transform=a1.get_xaxis_transform())
    ypos = {'USDC': 0, 'USDP': 1, 'USDT': 2}
    for _, r in score.iterrows():
        y = ypos[r['Issuer']]
        a1.plot([r['Band_lo'], r['Band_hi']], [y, y], lw=4, color=_TF_C[r['Issuer']],
                alpha=0.35, solid_capstyle='round')
        a1.plot(r['Score'], y, 'o', ms=7, color=_TF_C[r['Issuer']])
        a1.text(r['Band_hi'] + 1.5, y, f"{r['Issuer']}  {r['Score']:.1f}",
                fontsize=8.5, va='center', color=_TF_C[r['Issuer']])
    a1.set_ylim(-0.7, 2.7); a1.set_yticks([]); a1.spines['left'].set_visible(False)
    for i, (_, row) in enumerate(sc.iloc[::-1].iterrows()):
        for iss in ('USDP', 'USDC', 'USDT'):
            a2.plot(row[iss], i, 'o', ms=5, color=_TF_C[iss], mec='white', mew=0.5)
    a2.set_yticks(range(len(sc))); a2.set_yticklabels(sc.index[::-1], fontsize=8)
    a2.set_ylim(-0.6, len(sc) - 0.4)
    a2.set_xlabel('Composite score (0 = narrow-bank pole, 100 = MMF pole)')
    for iss in ('USDC', 'USDP', 'USDT'):
        a2.plot([], [], 'o', ms=5, color=_TF_C[iss], label=iss)
    a2.legend(loc='center right', fontsize=8)
    fig.savefig(THESIS_FIG_DIR / 'fig_7_1_classification.png'); plt.close(fig)


def make_thesis_figures():
    """Section 9 entry point: regenerate the seven publication figures."""
    print('\n' + '=' * 78)
    print('SECTION 9: Publication figures for the written thesis')
    print('=' * 78)
    THESIS_FIG_DIR.mkdir(parents=True, exist_ok=True)

    def _csv(name):
        p = OUT_DIR / name
        return pd.read_csv(p) if p.exists() else None

    cb, mtm = _csv('capital_bridge_timeseries.csv'), _csv('mtm_loss_timeseries.csv')
    lcr, usdp_rwa = _csv('lcr_effective_coverage.csv'), _csv('usdp_rwa_table.csv')
    share = _csv('stablecoin_tbill_share.csv')
    score = _csv('classification_scorecard.csv')
    schemes = _csv('classification_weight_sensitivity.csv')
    eps = _csv('rate_shock_episodes.csv')
    if eps is not None:
        for c in ('start', 'end', 'peak_date'):
            eps[c] = pd.to_datetime(eps[c])
    if mtm is not None:
        mtm = mtm.drop_duplicates(['Quarter', 'Issuer'])

    try:
        bank = pd.read_excel(EXCEL_PATH, sheet_name='Bank_Benchmark',
                             skiprows=3, header=0)
        bank.columns = ['Quarter', 'JPM', 'JPM_RWA', 'BNY', 'USDC', 'USDT', 'USDP']
        bank = bank.dropna(subset=['Quarter']).head(14)
    except Exception as e:
        bank = None
        print(f'  [thesis-figs] Bank_Benchmark sheet unavailable ({e}); skipping fig 4.2')
    try:
        bg = pd.read_excel(EXCEL_PATH, sheet_name='BTC_Gold_Shocks',
                           skiprows=4, header=0)
        bg.columns = ['Quarter', 'Shock', 'BTC', 'Gold', 'Equity', 'LossBTC',
                      'BufBTC', 'LossGold', 'BufGold', 'LossComb', 'BufComb',
                      'Breakeven']
        bg = (bg.dropna(subset=['Quarter']).drop_duplicates('Quarter')
                .set_index('Quarter').reindex(_TF_QUARTERS))
    except Exception as e:
        bg = None
        print(f'  [thesis-figs] BTC_Gold_Shocks sheet unavailable ({e}); skipping fig 6.2')

    tb_series, _basis = load_long_bill_history()

    with plt.rc_context(_TF_STYLE):
        if cb is not None and mtm is not None and usdp_rwa is not None:
            _tf_fig_4_1(cb, mtm, usdp_rwa);  print('  Saved fig_4_1_rwa_density.png')
        if bank is not None:
            _tf_fig_4_2(bank);               print('  Saved fig_4_2_bank_corridor.png')
        if tb_series is not None and eps is not None:
            _tf_fig_5_1(tb_series, eps);     print('  Saved fig_5_1_rate_shock_history.png')
        else:
            print('  [thesis-figs] TB3MS sheet or rate_shock_episodes.csv missing; '
                  'skipping fig 5.1')
        if lcr is not None:
            _tf_fig_6_1(lcr);                print('  Saved fig_6_1_lcr_40pct.png')
        if bg is not None:
            brk = _tf_fig_6_2(bg)
            print('  Saved fig_6_2_combined_stress.png')
            print('    break-even (verify vs Table 6.4):',
                  {q: int(v) for q, v in brk.items()})
        if share is not None:
            _tf_fig_6_3(share);              print('  Saved fig_6_3_tbill_share.png')
        if score is not None and schemes is not None:
            _tf_fig_7_1(score, schemes);     print('  Saved fig_7_1_classification.png')

    print(f'  -> {THESIS_FIG_DIR}')


if __name__ == '__main__':
    # Second entry point under the append-only convention: `python analysis.py`
    # runs main() above (Sections 1-8) and then regenerates the thesis figures.
    make_thesis_figures()
